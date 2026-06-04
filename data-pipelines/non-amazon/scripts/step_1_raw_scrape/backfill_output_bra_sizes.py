#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import json
from pathlib import Path
import sys
from typing import Dict, Iterable, List, Sequence, Tuple

SCRIPT_DIR = Path(__file__).resolve().parent
REPO_ROOT = SCRIPT_DIR.parents[3]
if str(SCRIPT_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPT_DIR))

from step1_intake_utils import BRA_SIZE_RE, extract_measurements, normalize_whitespace  # noqa: E402

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover - local system Python may not include openpyxl.
    load_workbook = None


TEXT_COLUMNS: Sequence[str] = (
    "user_comment",
    "review_body",
    "review_text",
    "body",
    "text",
)
SIZE_COLUMNS: Sequence[str] = (
    "size_display",
    "SizeOrdered(generated)",
    "asizemini",
    "ordered_size",
)
OUTPUT_PATTERNS: Sequence[str] = ("*.csv", "*.xlsx")
BRA_COLUMNS: Sequence[str] = (
    "bra_band_in_display",
    "bust_in_number_display",
    "cupsize_display",
)


def default_outputs_root() -> Path:
    return REPO_ROOT / "outputs"


def iter_output_paths(root: Path) -> Iterable[Path]:
    seen = set()
    for pattern in OUTPUT_PATTERNS:
        for path in sorted(root.rglob(pattern)):
            if path.name.startswith("~$") or path in seen:
                continue
            seen.add(path)
            yield path


def normalize_header(value: object) -> str:
    return normalize_whitespace(value).strip()


def first_present(row: Dict[str, str], columns: Sequence[str]) -> str:
    for column in columns:
        value = normalize_whitespace(row.get(column, ""))
        if value:
            return value
    return ""


def extracted_bra_fields(row: Dict[str, str]) -> Dict[str, str]:
    comment = first_present(row, TEXT_COLUMNS)
    size_hint = first_present(row, SIZE_COLUMNS)
    if not comment and not size_hint:
        return {}
    extracted = extract_measurements(comment, size_hint)
    if not extracted.get("cupsize_display"):
        return {}
    if not (extracted.get("bra_band_in_display") or extracted.get("bust_in_number_display")):
        return {}
    if not (BRA_SIZE_RE.search(size_hint) or BRA_SIZE_RE.search(comment)):
        return {}
    return extracted


def apply_bra_fields(row: Dict[str, str], *, repair_existing: bool) -> Tuple[Dict[str, str], int]:
    extracted = extracted_bra_fields(row)
    if not extracted:
        return row, 0

    updates = 0
    band = extracted.get("bra_band_in_display") or extracted.get("bust_in_number_display", "")
    cup = extracted.get("cupsize_display", "")
    replacements = {
        "bra_band_in_display": band,
        "bust_in_number_display": band,
        "cupsize_display": cup,
    }
    for field, value in replacements.items():
        if not value:
            continue
        current = normalize_whitespace(row.get(field, ""))
        if current and (not repair_existing or current == value):
            continue
        row[field] = value
        updates += 1
    return row, updates


def ensure_columns(fieldnames: Sequence[str]) -> List[str]:
    updated = list(fieldnames)
    for column in BRA_COLUMNS:
        if column in updated:
            continue
        if column == "bra_band_in_display" and "bust_in_display" in updated:
            updated.insert(updated.index("bust_in_display") + 1, column)
        elif column == "bust_in_number_display" and "bra_band_in_display" in updated:
            updated.insert(updated.index("bra_band_in_display") + 1, column)
        elif column == "cupsize_display" and "bust_in_number_display" in updated:
            updated.insert(updated.index("bust_in_number_display") + 1, column)
        else:
            updated.append(column)
    return updated


def rewrite_csv(path: Path, *, dry_run: bool, repair_existing: bool) -> Dict[str, int]:
    with path.open(newline="", encoding="utf-8-sig") as handle:
        reader = csv.DictReader(handle)
        if not reader.fieldnames:
            return {"rows": 0, "rows_updated": 0, "fields_updated": 0, "skipped": 1}
        rows = [dict(row) for row in reader]
        fieldnames = ensure_columns(reader.fieldnames)

    rows_updated = 0
    fields_updated = 0
    for row in rows:
        _, updates = apply_bra_fields(row, repair_existing=repair_existing)
        if updates:
            rows_updated += 1
            fields_updated += updates

    if fields_updated and not dry_run:
        with path.open("w", newline="", encoding="utf-8") as handle:
            writer = csv.DictWriter(handle, fieldnames=fieldnames, extrasaction="ignore")
            writer.writeheader()
            for row in rows:
                writer.writerow({field: row.get(field, "") for field in fieldnames})

    return {"rows": len(rows), "rows_updated": rows_updated, "fields_updated": fields_updated, "skipped": 0}


def rewrite_xlsx(path: Path, *, dry_run: bool, repair_existing: bool) -> Dict[str, int]:
    if load_workbook is None:
        return {"rows": 0, "rows_updated": 0, "fields_updated": 0, "skipped": 1}

    workbook = load_workbook(path)
    rows_total = 0
    rows_updated = 0
    fields_updated = 0
    workbook_changed = False

    for worksheet in workbook.worksheets:
        if worksheet.max_row < 2:
            continue
        headers = [normalize_header(worksheet.cell(row=1, column=col).value) for col in range(1, worksheet.max_column + 1)]
        if not any(column in headers for column in TEXT_COLUMNS + SIZE_COLUMNS):
            continue
        headers = ensure_columns(headers)
        for index, header in enumerate(headers, start=1):
            if worksheet.cell(row=1, column=index).value != header:
                worksheet.cell(row=1, column=index).value = header
                workbook_changed = True

        col_index = {header: index for index, header in enumerate(headers, start=1) if header}
        for row_number in range(2, worksheet.max_row + 1):
            row = {
                header: normalize_whitespace(worksheet.cell(row=row_number, column=col).value)
                for header, col in col_index.items()
            }
            rows_total += 1
            _, updates = apply_bra_fields(row, repair_existing=repair_existing)
            if not updates:
                continue
            rows_updated += 1
            fields_updated += updates
            workbook_changed = True
            if not dry_run:
                for field in BRA_COLUMNS:
                    worksheet.cell(row=row_number, column=col_index[field]).value = row.get(field, "")

    if workbook_changed and fields_updated and not dry_run:
        workbook.save(path)

    return {"rows": rows_total, "rows_updated": rows_updated, "fields_updated": fields_updated, "skipped": 0}


def rewrite_path(path: Path, *, dry_run: bool, repair_existing: bool) -> Dict[str, int]:
    if path.suffix.lower() == ".csv":
        return rewrite_csv(path, dry_run=dry_run, repair_existing=repair_existing)
    if path.suffix.lower() == ".xlsx":
        return rewrite_xlsx(path, dry_run=dry_run, repair_existing=repair_existing)
    return {"rows": 0, "rows_updated": 0, "fields_updated": 0, "skipped": 1}


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Backfill missing or truncated bra band/cup fields in Supabase review outputs from deterministic text parsing."
    )
    parser.add_argument("--root", type=Path, default=default_outputs_root(), help="outputs/ root to scan.")
    parser.add_argument("--dry-run", action="store_true", help="Report changes without rewriting files.")
    parser.add_argument(
        "--repair-existing",
        action="store_true",
        help="Repair existing bra-size values when explicit text has a different deterministic band/cup value.",
    )
    parser.add_argument("--report-json", type=Path, help="Optional JSON report path.")
    args = parser.parse_args()

    totals = {"files": 0, "skipped_files": 0, "rows": 0, "rows_updated": 0, "fields_updated": 0}
    changed_files = []
    for path in iter_output_paths(args.root):
        stats = rewrite_path(path, dry_run=args.dry_run, repair_existing=args.repair_existing)
        totals["files"] += 1
        totals["skipped_files"] += stats["skipped"]
        totals["rows"] += stats["rows"]
        totals["rows_updated"] += stats["rows_updated"]
        totals["fields_updated"] += stats["fields_updated"]
        if stats["fields_updated"]:
            changed_files.append({"path": str(path), **stats})

    report = {"totals": totals, "changed_files": changed_files}
    print(json.dumps(totals, indent=2))
    for item in changed_files:
        print(f"{item['path']}: {item['rows_updated']} rows, {item['fields_updated']} fields")
    if args.report_json:
        args.report_json.parent.mkdir(parents=True, exist_ok=True)
        args.report_json.write_text(json.dumps(report, indent=2) + "\n", encoding="utf-8")


if __name__ == "__main__":
    main()
