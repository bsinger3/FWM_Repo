#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import json
import re
from pathlib import Path
import sys
from typing import Dict, Iterable, List, Sequence, Tuple

SCRIPT_DIR = Path(__file__).resolve().parent
REPO_ROOT = SCRIPT_DIR.parents[3]
if str(SCRIPT_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPT_DIR))

from step1_intake_utils import normalize_whitespace, numeric_text, parse_number_text  # noqa: E402

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover - system Python may not have openpyxl.
    load_workbook = None


WEIGHT_RANGE_RE = re.compile(
    r"\b(\d{2,3}(?:\.\d+)?)\s*(?:-|–|—|to)\s*(\d{2,3}(?:\.\d+)?)\s*(lbs?|pounds?|#)?\b",
    re.I,
)
TEXT_COLUMNS: Sequence[str] = (
    "weight_raw",
    "weight_display_display",
    "user_comment",
    "review_body",
    "review_text",
    "body",
    "text",
)
WEIGHT_COLUMNS: Sequence[str] = (
    "weight_raw",
    "weight_display_display",
    "weight_lbs_display",
)
OUTPUT_PATTERNS: Sequence[str] = ("*.csv", "*.xlsx")


def default_outputs_root() -> Path:
    return REPO_ROOT / "outputs"


def iter_csv_paths(root: Path) -> Iterable[Path]:
    patterns = ("*reviews_matching_intake_schema.csv", "*reviews_matching_amazon_schema.csv")
    seen = set()
    for pattern in patterns:
        for path in sorted(root.rglob(pattern)):
            if path in seen:
                continue
            seen.add(path)
            yield path


def iter_output_paths(root: Path) -> Iterable[Path]:
    seen = set()
    for pattern in OUTPUT_PATTERNS:
        for path in sorted(root.rglob(pattern)):
            if path.name.startswith("~$") or path in seen:
                continue
            seen.add(path)
            yield path


def weight_range_from_text(text: str, *, trusted_weight_field: bool = False) -> str:
    for match in WEIGHT_RANGE_RE.finditer(text):
        try:
            low = parse_number_text(match.group(1))
            high = parse_number_text(match.group(2))
        except ValueError:
            continue
        unit = normalize_whitespace(match.group(3))
        context = text[max(0, match.start() - 32) : match.start()].lower()
        if not unit and not trusted_weight_field and not re.search(r"\b(?:weight|weighs?|pounds?|lbs?)\b", context):
            continue
        if 50 <= low < high <= 700 and high - low <= 150:
            return f"{numeric_text(low)}-{numeric_text(high)} lb"
    return ""


def row_weight_range(row: Dict[str, str]) -> str:
    for column in TEXT_COLUMNS:
        value = normalize_whitespace(row.get(column, ""))
        if not value:
            continue
        found = weight_range_from_text(value, trusted_weight_field=column in {"weight_raw", "weight_display_display"})
        if found:
            return found
    return ""


def ensure_columns(fieldnames: Sequence[str]) -> List[str]:
    updated = list(fieldnames)
    for column in WEIGHT_COLUMNS:
        if column in updated:
            continue
        if column == "weight_lbs_display" and "cupsize_display" in updated:
            updated.insert(updated.index("cupsize_display") + 1, column)
        elif column == "weight_display_display" and "search_fts" in updated:
            updated.insert(updated.index("search_fts") + 1, column)
        elif column == "weight_raw" and "height_raw" in updated:
            updated.insert(updated.index("height_raw") + 1, column)
        else:
            updated.append(column)
    return updated


def apply_weight_range(row: Dict[str, str]) -> Tuple[Dict[str, str], int]:
    display_range = row_weight_range(row)
    if not display_range:
        return row, 0

    updates = 0
    for field in ("weight_raw", "weight_display_display"):
        if field in row and normalize_whitespace(row.get(field, "")) != display_range:
            row[field] = display_range
            updates += 1
    if "weight_lbs_display" in row and normalize_whitespace(row.get("weight_lbs_display", "")):
        row["weight_lbs_display"] = ""
        updates += 1
    return row, updates


def rewrite_csv(path: Path, *, dry_run: bool) -> Dict[str, int]:
    with path.open(newline="", encoding="utf-8-sig") as handle:
        reader = csv.DictReader(handle)
        if not reader.fieldnames or not any(column in reader.fieldnames for column in TEXT_COLUMNS):
            return {"rows": 0, "rows_updated": 0, "fields_updated": 0, "skipped": 1}
        fieldnames = ensure_columns(reader.fieldnames)
        rows = [dict(row) for row in reader]
        for row in rows:
            for field in fieldnames:
                row.setdefault(field, "")

    rows_updated = 0
    fields_updated = 0
    for row in rows:
        _, updates = apply_weight_range(row)
        if updates:
            rows_updated += 1
            fields_updated += updates

    if fields_updated and not dry_run:
        with path.open("w", newline="", encoding="utf-8") as handle:
            writer = csv.DictWriter(handle, fieldnames=fieldnames, extrasaction="ignore")
            writer.writeheader()
            for row in rows:
                writer.writerow({field: row.get(field, "") for field in fieldnames})

    return {"rows": len(rows), "rows_updated": rows_updated, "fields_updated": fields_updated, "skipped": 0}


def rewrite_xlsx(path: Path, *, dry_run: bool) -> Dict[str, int]:
    if load_workbook is None:
        return {"rows": 0, "rows_updated": 0, "fields_updated": 0, "skipped": 1}

    workbook = load_workbook(path)
    rows_total = 0
    rows_updated = 0
    fields_updated = 0
    workbook_changed = False

    for worksheet in workbook.worksheets:
        if worksheet.max_row < 2:
            continue
        headers = [normalize_whitespace(worksheet.cell(row=1, column=col).value) for col in range(1, worksheet.max_column + 1)]
        if not any(column in headers for column in TEXT_COLUMNS):
            continue
        headers = ensure_columns(headers)
        for index, header in enumerate(headers, start=1):
            if worksheet.cell(row=1, column=index).value != header:
                worksheet.cell(row=1, column=index).value = header
                workbook_changed = True
        col_index = {header: index for index, header in enumerate(headers, start=1) if header}

        for row_number in range(2, worksheet.max_row + 1):
            row = {
                header: normalize_whitespace(worksheet.cell(row=row_number, column=col).value)
                for header, col in col_index.items()
            }
            rows_total += 1
            _, updates = apply_weight_range(row)
            if not updates:
                continue
            rows_updated += 1
            fields_updated += updates
            workbook_changed = True
            if not dry_run:
                for field in WEIGHT_COLUMNS:
                    worksheet.cell(row=row_number, column=col_index[field]).value = row.get(field, "")

    if workbook_changed and fields_updated and not dry_run:
        workbook.save(path)

    return {"rows": rows_total, "rows_updated": rows_updated, "fields_updated": fields_updated, "skipped": 0}


def rewrite_path(path: Path, *, dry_run: bool) -> Dict[str, int]:
    if path.suffix.lower() == ".csv":
        return rewrite_csv(path, dry_run=dry_run)
    if path.suffix.lower() == ".xlsx":
        return rewrite_xlsx(path, dry_run=dry_run)
    return {"rows": 0, "rows_updated": 0, "fields_updated": 0, "skipped": 1}


def update_summary(csv_path: Path) -> None:
    summary_path = csv_path.with_name(f"{csv_path.stem}_summary.json")
    if not summary_path.exists():
        return
    try:
        payload = json.loads(summary_path.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return
    payload["weight_range_backfilled_at"] = "2026-06-03"
    summary_path.write_text(json.dumps(payload, indent=2) + "\n", encoding="utf-8")


def run(paths: Iterable[Path], *, dry_run: bool, update_summaries: bool) -> Dict[str, object]:
    totals = {"files": 0, "skipped_files": 0, "rows": 0, "rows_updated": 0, "fields_updated": 0}
    changed_files = []
    for path in paths:
        stats = rewrite_path(path, dry_run=dry_run)
        totals["files"] += 1
        totals["skipped_files"] += stats["skipped"]
        totals["rows"] += stats["rows"]
        totals["rows_updated"] += stats["rows_updated"]
        totals["fields_updated"] += stats["fields_updated"]
        if stats["fields_updated"]:
            changed_files.append({"path": str(path), **stats})
            if update_summaries and path.suffix.lower() == ".csv" and not dry_run:
                update_summary(path)
    return {"totals": totals, "changed_files": changed_files}


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Preserve explicit weight ranges as display text and clear numeric weight fields for those rows."
    )
    parser.add_argument("--root", type=Path, default=default_outputs_root(), help="Root to scan.")
    parser.add_argument(
        "--mode",
        choices=("outputs", "raw-scrapes"),
        default="outputs",
        help="Use outputs mode for CSV/XLSX review outputs; raw-scrapes mode for Step 1 merchant CSVs.",
    )
    parser.add_argument("--dry-run", action="store_true", help="Report changes without rewriting files.")
    parser.add_argument("--report-json", type=Path, help="Optional JSON report path.")
    args = parser.parse_args()

    paths = iter_csv_paths(args.root) if args.mode == "raw-scrapes" else iter_output_paths(args.root)
    report = run(paths, dry_run=args.dry_run, update_summaries=args.mode == "raw-scrapes")
    print(json.dumps(report["totals"], indent=2))
    for item in report["changed_files"]:
        print(f"{item['path']}: {item['rows_updated']} rows, {item['fields_updated']} fields")
    if args.report_json:
        args.report_json.parent.mkdir(parents=True, exist_ok=True)
        args.report_json.write_text(json.dumps(report, indent=2) + "\n", encoding="utf-8")


if __name__ == "__main__":
    main()
