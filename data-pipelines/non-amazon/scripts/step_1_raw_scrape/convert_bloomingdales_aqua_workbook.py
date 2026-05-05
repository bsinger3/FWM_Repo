#!/usr/bin/env python3
from __future__ import annotations

import hashlib
import json
import os
import re
import sys
from datetime import datetime, timezone
from pathlib import Path
from typing import Dict, List, Optional, Sequence
from urllib.parse import urlparse, urlunparse

from openpyxl import load_workbook

from step1_intake_utils import (
    ProductContext,
    ReviewImage,
    build_intake_row,
    dedupe_rows,
    normalize_whitespace,
    validate_rows,
    write_intake_csv,
)


ROOT = Path(__file__).resolve().parents[4]
DATA_ROOT = Path(os.environ.get("FWM_DATA_DIR", ROOT.parent / "FWM_Data"))
RETAILER = "bloomingdales_aqua"
SOURCE_SITE = "https://www.bloomingdales.com/"
OUTPUT_DIR = DATA_ROOT / "non-amazon" / "data" / "step_1_raw_scraping_data" / RETAILER
WORKBOOK_PATH = OUTPUT_DIR / "Bloomingdales_Aqua.xlsx"
OUTPUT_CSV = OUTPUT_DIR / f"{RETAILER}_reviews_matching_intake_schema.csv"
SUMMARY_JSON = OUTPUT_DIR / f"{RETAILER}_reviews_matching_intake_schema_summary.json"


def utc_now() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z")


def clean_url(value: object) -> str:
    text = normalize_whitespace(value)
    if not text:
        return ""
    parsed = urlparse(text)
    return urlunparse((parsed.scheme, parsed.netloc, parsed.path, "", parsed.query, ""))


def product_key(value: object) -> str:
    parsed = urlparse(clean_url(value))
    return f"{parsed.path}?{parsed.query}".lower().rstrip("?")


def clean_labeled_value(value: object, label: str) -> str:
    text = normalize_whitespace(value)
    if not text:
        return ""
    text = re.sub(rf"^{re.escape(label)}\s*:?\s*", "", text, flags=re.I)
    return normalize_whitespace(text)


def product_lookup(workbook) -> Dict[str, Dict[str, str]]:
    if "ProductLiks" not in workbook.sheetnames:
        return {}
    sheet = workbook["ProductLiks"]
    rows = sheet.iter_rows(values_only=True)
    headers = [normalize_whitespace(value) for value in next(rows)]
    lookup: Dict[str, Dict[str, str]] = {}
    for row in rows:
        item = {headers[index]: row[index] if index < len(row) else "" for index in range(len(headers))}
        url = clean_url(item.get("Title_URL"))
        if not url:
            continue
        lookup[product_key(url)] = {
            "title": normalize_whitespace(item.get("Title")),
            "url": url,
            "brand": normalize_whitespace(item.get("productbrand")) or "Bloomingdale's",
            "description": normalize_whitespace(item.get("Keywords") or item.get("Description") or item.get("Description1")),
        }
    return lookup


def title_from_url(url: str) -> str:
    path = urlparse(url).path.rstrip("/")
    slug = path.rsplit("/", 1)[-1]
    return normalize_whitespace(slug.replace("-", " ").title())


def context_for(url: str, lookup: Dict[str, Dict[str, str]]) -> ProductContext:
    meta = lookup.get(product_key(url), {})
    title = meta.get("title") or title_from_url(url)
    return ProductContext(
        url=url,
        title=title,
        description=meta.get("description", ""),
        brand=meta.get("brand", "") or "Bloomingdale's",
        category="",
        provider_hints="Bazaarvoice workbook export",
    )


def image_groups(headers: Sequence[str]) -> List[Dict[str, int]]:
    groups: List[Dict[str, int]] = []
    for index, header in enumerate(headers):
        if not header.lower().startswith("bigimage"):
            continue
        suffix = header[len("BigImage") :]
        groups.append(
            {
                "image": index,
                "size": index - 1 if index > 0 and headers[index - 1].lower().startswith("sizeordered") else -1,
                "profile": index + 1 if index + 1 < len(headers) and headers[index + 1].lower().startswith("height") else -1,
                "usual": index + 2 if index + 2 < len(headers) and headers[index + 2].lower().startswith("usualsize") else -1,
                "page": index + 4 if index + 4 < len(headers) and headers[index + 4].lower().startswith("page_url") else -1,
                "suffix": suffix or "1",
            }
        )
    return groups


def row_value(row: Sequence[object], index: int) -> str:
    return normalize_whitespace(row[index]) if 0 <= index < len(row) else ""


def profile_parts(profile_raw: str, usual_raw: str) -> List[str]:
    parts: List[str] = []
    profile = clean_labeled_value(profile_raw, "Usual size")
    usual = clean_labeled_value(usual_raw, "Usual size")
    for value in [profile, usual]:
        if not value:
            continue
        if re.search(r"\b[4-6]\s*(?:'|ft|feet)\s*\d{0,2}", value, re.I):
            parts.append(f"Height: {value}")
        else:
            parts.append(f"Usual size: {value}")
    return parts


def convert() -> tuple[List[Dict[str, str]], Dict[str, object]]:
    started_at = utc_now()
    workbook = load_workbook(WORKBOOK_PATH, read_only=True, data_only=True)
    lookup = product_lookup(workbook)
    sheet = workbook["Sheet1"]
    rows_iter = sheet.iter_rows(values_only=True)
    headers = [normalize_whitespace(value) for value in next(rows_iter)]
    groups = image_groups(headers)
    converted: List[Dict[str, str]] = []
    skipped_blank_image = 0
    skipped_blank_url = 0
    for row_number, row in enumerate(rows_iter, start=2):
        comment = row_value(row, 4)
        for group in groups:
            image_url = row_value(row, group["image"])
            if not image_url or not image_url.startswith("http"):
                skipped_blank_image += 1
                continue
            page_url = clean_url(row_value(row, group["page"]))
            if not page_url:
                skipped_blank_url += 1
                continue
            size = clean_labeled_value(row_value(row, group["size"]), "Size ordered")
            profile = profile_parts(row_value(row, group["profile"]), row_value(row, group["usual"]))
            body = normalize_whitespace(" ".join(part for part in [comment, *profile] if part))
            review_hash = hashlib.md5(f"{page_url}|{image_url}|{body}|{size}".encode("utf-8")).hexdigest()[:12]
            review = ReviewImage(
                image_url=image_url,
                review_id=f"bloomingdales-aqua-{review_hash}",
                review_body=body,
                size_raw=size,
                extra={"product_url": page_url},
            )
            converted.append(build_intake_row(context_for(page_url, lookup), review, utc_now()))
    rows = dedupe_rows(converted)
    rows.sort(key=lambda item: (item.get("product_page_url_display", ""), item.get("original_url_display", ""), item.get("id", "")))
    summary: Dict[str, object] = {
        "site": SOURCE_SITE,
        "retailer": RETAILER,
        "adapter": "local_bazaarvoice_workbook_conversion",
        "source_workbook": str(WORKBOOK_PATH),
        "source_sheets": ["Sheet1", "ProductLiks"],
        "started_at": started_at,
        "finished_at": utc_now(),
        "output_csv": str(OUTPUT_CSV),
        "products_discovered": len(lookup),
        "products_scanned": len(lookup),
        "product_pages_scanned": len(lookup),
        "rows_before_dedupe": len(converted),
        "skipped_blank_image_cells": skipped_blank_image,
        "skipped_blank_url_cells": skipped_blank_url,
        "access_policy": "local_workbook_conversion_only; no_live_site_requests",
        "scrape_scope_status": "completed_from_workbook_review_photo_sheet",
        "full_catalog_scrape_complete": False,
        "seed_scrape_only": False,
        "warnings": [
            "Converted from existing workbook sheets, not a live site scrape.",
            "Workbook contains repeated BigImage/Page_URL groups; rows were deduped by generated review id, product URL, and image URL.",
        ],
    }
    summary.update(validate_rows(rows))
    return rows, summary


def main() -> int:
    rows, summary = convert()
    write_intake_csv(rows, OUTPUT_CSV)
    SUMMARY_JSON.write_text(json.dumps(summary, indent=2), encoding="utf-8")
    print(f"Rows written: {len(rows)}")
    print(f"Qualified rows: {summary.get('rows_with_image_product_size_and_measurement', 0)}")
    print(f"CSV: {OUTPUT_CSV}")
    print(f"Summary: {SUMMARY_JSON}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
