#!/usr/bin/env python3
from __future__ import annotations

import argparse
import hashlib
import json
import re
import time
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Sequence
from urllib.parse import parse_qs, urlencode, urlparse

import openpyxl

from step1_intake_utils import (
    ProductContext,
    ReviewImage,
    build_intake_row,
    classify_clothing_type,
    dedupe_rows,
    fetch_text,
    normalize_whitespace,
    output_paths,
    strip_tags,
    utc_now,
    write_intake_csv,
    write_summary,
)


SITE_ROOT = "https://www.victoriassecret.com"
DOMAIN = "victoriassecret.com"
RETAILER = "victoriassecret_com"
DEFAULT_WORKBOOK = (
    Path(__file__).resolve().parents[4] / "FWM_Data"
    / "non-amazon"
    / "data"
    / "step_1_raw_scraping_data"
    / "vs"
    / "VS.xlsx"
)
SHEET_NAME = "VictoriasSecret_bigImages1"
PRODUCT_LINK_SHEET = "VSprodLinks"
PRODUCT_API = "https://api.victoriassecret.com/products/v37/page/{product_id}"
REVIEWS_API = "https://api.victoriassecret.com/ratings-and-reviews/v4/reviews"
IMAGE_ROOT = "https://www.victoriassecret.com/p/404x539"

SIZE_RE = re.compile(
    r"\b(?:ordered|order(?:ed)?|bought|purchased|got|picked|wearing|wears?|size)\s+"
    r"(?:it\s+)?(?:in\s+)?(?:a\s+|an\s+)?(?:size\s+)?"
    r"(?P<size>XXS|XS|XSP|S|SP|M|MP|L|LP|XL|XLP|XXL|2X|3X|4X|"
    r"x-small|small|medium|large|x-large|xx-large|\d{1,2}(?:/\d{1,2})?)\b",
    re.I,
)
VALID_SIZE_RE = re.compile(
    r"^(?:XXS|XS|XSP|S|SP|M|MP|L|LP|XL|XLP|XXL|2X|3X|4X|"
    r"\d{1,2}(?:/\d{1,2})?|"
    r"(?:28|30|32|34|36|38|40|42|44|46|48|50|52|54)\s*"
    r"(?:AAA|AA|A|B|C|D|DD|DDD|F|G|H|I|J|K))$",
    re.I,
)
HEIGHT_RANGE_RE = re.compile(r"^(?P<f1>\d)'(?P<i1>\d{1,2})-(?:(?P<f2>\d)')?(?P<i2>\d{1,2})$")
HEIGHT_OR_LESS_RE = re.compile(r"^(?P<f>\d)'(?P<i>\d{1,2})orless$", re.I)
HEIGHT_EXACT_RE = re.compile(r"^(?P<f>\d)'(?P<i>\d{1,2})$")


def workbook_rows(path: Path) -> list[tuple[object, ...]]:
    workbook = openpyxl.load_workbook(path, data_only=True)
    worksheet = workbook[SHEET_NAME]
    return list(worksheet.iter_rows(min_row=2, values_only=True))


def workbook_product_urls(path: Path) -> List[str]:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook[PRODUCT_LINK_SHEET]
    urls: List[str] = []
    seen = set()
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        url = clean(row[2] if len(row) > 2 else "")
        if not url or url in seen:
            continue
        seen.add(url)
        urls.append(url)
    return urls


def clean(value: object) -> str:
    return normalize_whitespace(value)


def product_title_from_page_title(page_title: str) -> str:
    title = re.sub(r"^Buy\s+", "", page_title or "", flags=re.I)
    title = re.sub(r"\s+-\s+Order\s+.*$", "", title, flags=re.I)
    title = title.split(",")[0]
    return normalize_whitespace(title)


def product_category_from_page_title(page_title: str) -> str:
    match = re.search(r"\bOrder\s+(.+?)\s+online\b", page_title or "", re.I)
    return normalize_whitespace(match.group(1)) if match else ""


def color_from_page_title(page_title: str) -> str:
    title = re.sub(r"^Buy\s+", "", page_title or "", flags=re.I)
    title = re.sub(r"\s+-\s+Order\s+.*$", "", title, flags=re.I)
    parts = [normalize_whitespace(part) for part in title.split(",")]
    return parts[1] if len(parts) > 1 else ""


def profile_fields(row: tuple[object, ...]) -> Dict[str, str]:
    fields: Dict[str, str] = {}
    for key_index in (11, 13, 15, 17):
        key = clean(row[key_index] if key_index < len(row) else "")
        value = clean(row[key_index + 1] if key_index + 1 < len(row) else "")
        if key and value:
            fields[key.lower()] = value
    return fields


def height_to_inches(value: str) -> tuple[str, str]:
    raw = clean(value).replace(" ", "")
    match = HEIGHT_RANGE_RE.match(raw)
    if match:
        f1 = int(match.group("f1"))
        i1 = int(match.group("i1"))
        f2 = int(match.group("f2") or f1)
        i2 = int(match.group("i2"))
        low = f1 * 12 + i1
        high = f2 * 12 + i2
        return value, str(round((low + high) / 2))
    match = HEIGHT_OR_LESS_RE.match(raw)
    if match:
        return value, str(int(match.group("f")) * 12 + int(match.group("i")))
    match = HEIGHT_EXACT_RE.match(raw)
    if match:
        return value, str(int(match.group("f")) * 12 + int(match.group("i")))
    return "", ""


def ordered_size(text: str) -> str:
    match = SIZE_RE.search(text or "")
    if not match:
        return ""
    size = normalize_whitespace(match.group("size")).upper()
    return {
        "X-SMALL": "XS",
        "SMALL": "S",
        "MEDIUM": "M",
        "LARGE": "L",
        "X-LARGE": "XL",
        "XX-LARGE": "XXL",
    }.get(size, size)


def sanitize_size(value: str) -> str:
    size = normalize_whitespace(value).upper()
    return size if VALID_SIZE_RE.fullmatch(size) else ""


def product_id_from_url(product_url: str) -> str:
    match = re.search(r"/[^/?#]+-catalog/(\d+)", product_url)
    return match.group(1) if match else ""


def product_api_url(product_url: str) -> str:
    parsed = urlparse(product_url)
    query = parse_qs(parsed.query)
    product_id = product_id_from_url(product_url)
    collection_id = (query.get("collectionId") or [""])[0]
    params = {"activeCountry": "US", "isWishlistEnabled": "true"}
    if collection_id:
        params["collectionId"] = collection_id
    return f"{PRODUCT_API.format(product_id=product_id)}?{urlencode(params)}"


def review_api_url(variant_id: str, offset: int, limit: int) -> str:
    params = {
        "filter": f"productid:eq:{variant_id}",
        "include": "authors,products",
        "stats": "reviews",
        "offset": str(offset),
        "limit": str(limit),
        "sort": "submissiontime:desc",
        "activeCountry": "US",
    }
    return f"{REVIEWS_API}?{urlencode(params)}"


def fetch_json(url: str, referer: str) -> Dict[str, object]:
    text = fetch_text(url, accept="application/json,text/plain,*/*", referer=referer)
    return json.loads(text)


def first_variant_id(product: Dict[str, object]) -> str:
    for generic in (product.get("productData") or {}).values():
        for choice in (generic.get("choices") or {}).values():
            for size in (choice.get("availableSizes") or {}).values():
                variant_id = clean(size.get("variantId"))
                if variant_id:
                    return variant_id
    return ""


def image_url(path: str) -> str:
    path = clean(path)
    if not path:
        return ""
    if path.startswith("http"):
        return path
    return f"{IMAGE_ROOT}/{path}.jpg"


def context_for_product(product_url: str, product: Dict[str, object], generic: Dict[str, object], choice: Dict[str, object]) -> ProductContext:
    title = clean(generic.get("shortDescription") or product.get("shortDescription"))
    color = clean(choice.get("label") or choice.get("color"))
    description = strip_tags(generic.get("longDescription") or generic.get("pimLongDescription"))
    return ProductContext(
        url=product_url,
        title=title,
        description=description,
        detail=description,
        category=clean(product.get("categoryDisplay") or product.get("classDisplay")),
        brand=clean(product.get("brandName")) or "Victoria's Secret",
        color=color,
        product_id=clean(product.get("id")),
        provider_hints="victoriassecret_public_product_api",
    )


def model_comment(model: Dict[str, object]) -> str:
    parts = []
    name = clean(model.get("fullName") or model.get("shortName"))
    if name:
        parts.append(f"Catalog model: {name}.")
    if clean(model.get("height")):
        parts.append(f"Height: {clean(model.get('height'))}.")
    band = clean(model.get("band"))
    cup = clean(model.get("cup"))
    if band and cup:
        parts.append(f"Bra size: {band}{cup}.")
    size = clean(model.get("alphaSize") or model.get("size2") or model.get("size1"))
    if size:
        parts.append(f"Wearing size {size}.")
    return " ".join(parts)


def catalog_model_rows(product_url: str, product: Dict[str, object], fetched_at: str) -> List[Dict[str, str]]:
    rows: List[Dict[str, str]] = []
    model_info = product.get("modelInformation") or {}
    for generic_id, generic in (product.get("productData") or {}).items():
        for choice_id, choice in (generic.get("choices") or {}).items():
            context = context_for_product(product_url, product, generic, choice)
            for image in choice.get("images") or []:
                if "onModel" not in clean(image.get("type")):
                    continue
                model_id = clean(image.get("modelID"))
                model = model_info.get(model_id) or {}
                comment = model_comment(model)
                if not comment:
                    continue
                src = image_url(clean(image.get("image")))
                if not src:
                    continue
                size_raw = clean(model.get("band")) + clean(model.get("cup")) if clean(model.get("band")) and clean(model.get("cup")) else clean(model.get("alphaSize") or model.get("size2") or model.get("size1"))
                review = ReviewImage(
                    image_url=src,
                    review_id="vs-model-" + hashlib.md5(f"{generic_id}|{choice_id}|{src}|{comment}".encode("utf-8")).hexdigest()[:18],
                    review_title="Catalog model measurements",
                    review_body=comment,
                    reviewer_name=clean(model.get("fullName") or model.get("shortName")),
                    size_raw=size_raw,
                    extra={
                        "image_source_type": "catalog_model_image",
                        "image_source_detail": "public Victoria's Secret product API on-model image with structured model measurements",
                    },
                )
                row = build_intake_row(context, review, fetched_at)
                rows.append(row)
                break
    return rows


def review_height(review: Dict[str, object]) -> str:
    return clean(((review.get("ContextDataValues") or {}).get("Height") or {}).get("ValueLabel"))


def review_age(review: Dict[str, object]) -> str:
    return clean(((review.get("ContextDataValues") or {}).get("Age") or {}).get("ValueLabel"))


def review_fit(review: Dict[str, object]) -> str:
    return clean(((review.get("SecondaryRatings") or {}).get("Fit") or {}).get("ValueLabel"))


def review_rows(product_url: str, product: Dict[str, object], fetched_at: str, delay: float, review_limit: int) -> tuple[List[Dict[str, str]], Dict[str, object]]:
    variant_id = first_variant_id(product)
    rows: List[Dict[str, str]] = []
    summary = {"variant_id": variant_id, "review_pages": 0, "reviews_seen": 0, "photo_reviews": 0}
    if not variant_id:
        return rows, summary
    offset = 0
    while True:
        payload = fetch_json(review_api_url(variant_id, offset, review_limit), product_url)
        results = payload.get("Results") or []
        includes = payload.get("Includes") or {}
        products = includes.get("Products") or {}
        summary["review_pages"] += 1
        summary["reviews_seen"] += len(results)
        for result in results:
            photos = result.get("Photos") or []
            if not photos:
                continue
            product_record = products.get(clean(result.get("ProductId"))) or {}
            context = ProductContext(
                url=product_url,
                title=clean(product_record.get("Name") or result.get("OriginalProductName") or product.get("shortDescription")),
                description=clean(product_record.get("Description") or ""),
                category=clean(product.get("categoryDisplay") or product.get("classDisplay")),
                brand="Victoria's Secret",
                provider_hints="victoriassecret_public_ratings_reviews_api",
            )
            comment = normalize_whitespace(
                " ".join(
                    part
                    for part in [
                        clean(result.get("ReviewText")),
                        f"Fit: {review_fit(result)}" if review_fit(result) else "",
                        f"Age: {review_age(result)}" if review_age(result) else "",
                        f"Height: {review_height(result)}" if review_height(result) else "",
                    ]
                    if part
                )
            )
            size = ordered_size(" ".join([clean(result.get("Title")), clean(result.get("ReviewText"))]))
            for photo in photos:
                src = clean(((photo.get("Sizes") or {}).get("large") or {}).get("Url")) or clean(((photo.get("Sizes") or {}).get("normal") or {}).get("Url"))
                if not src:
                    continue
                review = ReviewImage(
                    image_url=src,
                    review_id=f"vs-review-{clean(result.get('Id'))}-{clean(photo.get('Id'))}",
                    review_title=clean(result.get("Title")),
                    review_body=comment,
                    reviewer_name=clean(result.get("UserNickname")),
                    date_raw=clean(result.get("SubmissionTime")),
                    size_raw=size,
                    rating=clean(result.get("Rating")),
                    extra={
                        "image_source_type": "customer_review_image",
                        "image_source_detail": "public Victoria's Secret ratings-and-reviews API customer review photo",
                    },
                )
                row = build_intake_row(context, review, fetched_at)
                rows.append(row)
                summary["photo_reviews"] += 1
        offset += len(results)
        if not results or offset >= int(payload.get("TotalResults") or 0):
            break
        if delay:
            time.sleep(delay)
    return rows, summary


def live_scrape(product_urls: Sequence[str], delay: float, review_limit: int, include_catalog_models: bool) -> tuple[List[Dict[str, str]], Dict[str, object]]:
    started_at = utc_now()
    rows: List[Dict[str, str]] = []
    product_summaries: List[Dict[str, object]] = []
    errors: List[str] = []
    for index, product_url in enumerate(product_urls, start=1):
        try:
            product_payload = fetch_json(product_api_url(product_url), product_url)
            product = product_payload.get("product") or {}
            product_rows, review_summary = review_rows(product_url, product, started_at, delay, review_limit)
            model_rows = catalog_model_rows(product_url, product, started_at) if include_catalog_models else []
            rows.extend(product_rows)
            rows.extend(model_rows)
            product_summaries.append(
                {
                    "product_url": product_url,
                    "master_style_id": clean(product.get("id")),
                    "title": clean(product.get("shortDescription")),
                    "customer_photo_rows": len(product_rows),
                    "catalog_model_rows": len(model_rows),
                    **review_summary,
                }
            )
            print(f"[vs {index}/{len(product_urls)}] review_rows={len(product_rows)} model_rows={len(model_rows)} url={product_url}", flush=True)
        except Exception as exc:
            errors.append(f"{product_url}: {exc}")
            print(f"[vs {index}/{len(product_urls)}] error={exc} url={product_url}", flush=True)
        if delay:
            time.sleep(delay)
    rows = dedupe_rows(rows)
    deduped: List[Dict[str, str]] = []
    seen_images = set()
    for row in rows:
        key = (row.get("id"), row.get("original_url_display"))
        if key in seen_images:
            continue
        seen_images.add(key)
        deduped.append(row)
    rows = deduped
    summary = {
        "site": SITE_ROOT,
        "retailer": "vs",
        "adapter": "victoriassecret_public_product_and_reviews_api_from_local_catalog",
        "started_at": started_at,
        "finished_at": utc_now(),
        "source_product_catalog": str(DEFAULT_WORKBOOK),
        "source_product_sheet": PRODUCT_LINK_SHEET,
        "products_discovered": len(product_urls),
        "products_scanned": len(product_urls),
        "product_pages_scanned": len(product_urls),
        "scrape_scope_status": "all_product_urls_from_local_vs_product_catalog_scanned",
        "full_catalog_scrape_complete": False,
        "customer_review_feed_used": True,
        "catalog_model_rows_enabled": include_catalog_models,
        "aggregate_feed_used": False,
        "access_policy": "public_product_and_ratings_reviews_api_only; no_auth_bypass; no_captcha_bypass",
        "product_summaries": product_summaries,
        "errors": errors,
    }
    return rows, summary


def row_id(product_url: str, image_url: str, review_title: str, date_raw: str, reviewer: str) -> str:
    basis = "|".join([product_url, image_url, review_title, date_raw, reviewer])
    return "vs-" + hashlib.md5(basis.encode("utf-8")).hexdigest()[:20]


def converted_row(raw: tuple[object, ...], fetched_at: str) -> Optional[Dict[str, str]]:
    page_title = clean(raw[0] if len(raw) > 0 else "")
    review_title = clean(raw[1] if len(raw) > 1 else "")
    image_url = clean(raw[2] if len(raw) > 2 else "")
    date_raw = clean(raw[5] if len(raw) > 5 else "")
    reviewer = clean(raw[6] if len(raw) > 6 else "")
    review_body = clean(raw[8] if len(raw) > 8 else "")
    product_url = clean(raw[21] if len(raw) > 21 else "") or clean(raw[22] if len(raw) > 22 else "")
    if not image_url or not product_url:
        return None

    fields = profile_fields(raw)
    height_raw, height_in = height_to_inches(fields.get("height", ""))
    context = ProductContext(
        url=product_url,
        title=product_title_from_page_title(page_title),
        description=page_title,
        category=product_category_from_page_title(page_title),
        brand="Victoria's Secret",
        color=color_from_page_title(page_title),
        provider_hints="legacy_bazaarvoice_workbook",
    )
    comment = normalize_whitespace(
        " ".join(
            part
            for part in [
                review_body,
                f"Fit: {fields.get('fit', '')}" if fields.get("fit") else "",
                f"Age: {fields.get('age', '')}" if fields.get("age") else "",
                f"Height: {fields.get('height', '')}" if fields.get("height") else "",
            ]
            if part
        )
    )
    review = ReviewImage(
        image_url=image_url,
        review_id=row_id(product_url, image_url, review_title, date_raw, reviewer),
        review_title=review_title,
        review_body=comment,
        reviewer_name=reviewer,
        date_raw=date_raw,
        size_raw=ordered_size(" ".join([review_title, review_body])),
        extra={
            "image_source_type": "customer_review_image",
            "image_source_detail": "legacy Bazaarvoice workbook customer review image",
        },
    )
    row = build_intake_row(context, review, fetched_at)
    if height_raw and not row.get("height_raw"):
        row["height_raw"] = height_raw
    if height_in and not row.get("height_in_display"):
        row["height_in_display"] = height_in
    row["size_display"] = sanitize_size(row.get("size_display", ""))
    row["clothing_type_id"] = classify_clothing_type(context)
    return row


def main(argv: Optional[Sequence[str]] = None) -> int:
    parser = argparse.ArgumentParser(description="Scrape Victoria's Secret public review photos and catalog model rows.")
    parser.add_argument("--workbook", type=Path, default=DEFAULT_WORKBOOK)
    parser.add_argument("--mode", choices=["live", "workbook"], default="live")
    parser.add_argument("--max-products", type=int, default=0)
    parser.add_argument("--request-delay-seconds", type=float, default=0.4)
    parser.add_argument("--review-limit", type=int, default=100)
    parser.add_argument("--skip-catalog-models", action="store_true")
    args = parser.parse_args(argv)

    started_at = utc_now()
    if args.mode == "live":
        product_urls = workbook_product_urls(args.workbook)
        if args.max_products:
            product_urls = product_urls[: args.max_products]
        rows, live_summary = live_scrape(
            product_urls,
            delay=args.request_delay_seconds,
            review_limit=args.review_limit,
            include_catalog_models=not args.skip_catalog_models,
        )
        started_at = live_summary["started_at"]
        finished_at = live_summary["finished_at"]
        product_summaries = live_summary["product_summaries"]
        errors = live_summary["errors"]
        products_scanned = int(live_summary["products_scanned"])
        adapter = str(live_summary["adapter"])
    else:
        fetched_at = started_at
        raw_rows = workbook_rows(args.workbook)
        rows = [row for raw in raw_rows if (row := converted_row(raw, fetched_at))]
        rows = dedupe_rows(rows)
        finished_at = utc_now()
        product_summaries = []
        errors = []
        products_scanned = len({row.get("product_page_url_display") for row in rows if row.get("product_page_url_display")})
        adapter = "legacy_bazaarvoice_workbook_conversion"
    output_csv, summary_json = output_paths(DOMAIN)
    write_intake_csv(rows, output_csv)
    write_summary(
        summary_json,
        site=SITE_ROOT,
        retailer=DOMAIN,
        rows=rows,
        output_csv=output_csv,
        started_at=started_at,
        finished_at=finished_at,
        products_scanned=products_scanned,
        adapter=adapter,
        product_summaries=product_summaries,
        errors=errors,
    )
    payload = json.loads(summary_json.read_text(encoding="utf-8"))
    if args.mode == "live":
        payload.update(live_summary)
    else:
        payload.update(
            {
                "source_workbook": str(args.workbook),
                "source_sheet": SHEET_NAME,
                "source_rows_read": len(raw_rows),
                "rows_from_workbook_with_image_and_product_url": len(rows),
                "scrape_scope_status": "legacy_workbook_conversion_only",
                "full_catalog_scrape_complete": False,
                "customer_review_feed_used": True,
                "aggregate_feed_used": False,
                "access_policy": "local_legacy_workbook_only; no_live_site_requests",
            }
        )
    payload["rows_supabase_qualified"] = payload.get("rows_with_image_product_size_and_measurement", 0)
    summary_json.write_text(json.dumps(payload, indent=2), encoding="utf-8")
    print(
        json.dumps(
            {
                RETAILER: {
                    "rows": len(rows),
                    "qualified_rows": payload.get("rows_with_image_product_size_and_measurement", 0),
                    "distinct_products": payload.get("distinct_products", 0),
                    "output_csv": str(output_csv),
                    "summary_json": str(summary_json),
                }
            },
            indent=2,
        )
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
