#!/usr/bin/env python3
"""Prefill Supabase review workbooks from prior human ground truth labels."""

from __future__ import annotations
import sys

import csv
import os
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation


REPO_ROOT = Path(__file__).resolve().parents[4]
PIPELINE_SCRIPTS_DIR = REPO_ROOT / "data-pipelines" / "scripts"
if str(PIPELINE_SCRIPTS_DIR) not in sys.path:
    sys.path.insert(0, str(PIPELINE_SCRIPTS_DIR))

from pipeline_paths import archive_root, cv_annotated_pending_human_review_root  # noqa: E402

LEGACY_OUTPUTS_ARCHIVE = archive_root() / "old_outputs" / "repo_outputs_archive" / "supabase_output_cleanup_2026_05_29"
CV_EXPERIMENTS_DIR = LEGACY_OUTPUTS_ARCHIVE / "cv_experiments"

PACKAGE_DIR = Path(
    os.environ.get(
        "SUPABASE_REVIEW_PACKAGE_DIR",
        cv_annotated_pending_human_review_root() / "supabase_production_image_review_2026_05_28_s3_refresh_sovrn_prioritized",
    )
)
SKIP_FILES: set[str] = set()

REJECTION_OPTIONS = [
    "LOW_RESOLUTION_AFTER_URL_REPAIR",
    "LOW_RESOLUTION",
    "TOO_DARK",
    "TOO_BRIGHT_OR_WASHED_OUT",
    "BLURRY_OR_MOTION_BLUR",
    "GRAINY_OR_NOISY",
    "GARMENT_CUT_OFF",
    "GARMENT_TOP_COVERED",
    "GARMENT_BOTTOM_CUT_OFF",
    "GARMENT_TOO_PARTIAL",
    "GARMENT_OBSCURED",
    "PERSON_TOO_FAR",
    "TARGET_WEARER_AMBIGUOUS",
    "BAD_ANGLE_TOP_DOWN",
    "BAD_ANGLE_SIDE_OR_TWISTED",
    "BACKGROUND_TOO_CLUTTERED",
    "BACKGROUND_VISUALLY_OFFPUTTING",
    "BATHROOM_OR_PUBLIC_RESTROOM_DISTRACTION",
    "MESSY_MIRROR_OR_DIRTY_SURFACE",
    "DISTRACTING_OBJECTS",
    "NOT_WORN_BY_PERSON",
    "NO_PERSON_VISIBLE",
    "NEEDS_CROP",
    "WRONG_PRODUCT_CONTEXT",
    "POSSIBLE_PRODUCT_CONTEXT_MISMATCH",
    "CATALOG_OR_STOCK_IMAGE",
    "DUPLICATE_OR_NEAR_DUPLICATE",
    "IMAGE_FETCH_FAILED",
    "INVALID_OR_DEAD_IMAGE_URL",
    "OTHER",
]

GROUND_TRUTH_FILES = [
    CV_EXPERIMENTS_DIR / "ground_truth_labeling/labeled_image_rejection_reason_queue.csv",
    CV_EXPERIMENTS_DIR / "ground_truth_labeling_broad/labeled_2026_05_25/combined_amazon_nonamazon_llm_seeded_ground_truth_queue_labeled.csv",
    CV_EXPERIMENTS_DIR / "ground_truth_labeling_broad/labeled_2026_05_25/usable_labeled_ground_truth_normalized.csv",
    CV_EXPERIMENTS_DIR / "combined_reason_ground_truth_2026_05_25/labeled_2026_05_27/combined_rejection_reason_yes_no_review_queue_labeled.csv",
    CV_EXPERIMENTS_DIR / "yolo_segmentation_crop_reasons_broad_2026_05_25/needs_crop_calibration_labeled_2026_05_25/needs_crop_yes_no_review_queue_labeled.csv",
    CV_EXPERIMENTS_DIR / "yolo_segmentation_crop_reasons_broad_2026_05_25/needs_crop_control_labeled_2026_05_25/needs_crop_yes_no_control_review_queue_labeled.csv",
    CV_EXPERIMENTS_DIR / "yolo_segmentation_crop_reasons_broad_2026_05_25/openai_not_worn_calibration_2026_05_25/human_review_labeled_2026_05_25/openai_not_worn_human_review_latest_labeled.csv",
    LEGACY_OUTPUTS_ARCHIVE / "supabase_approved_upload_staging/2026_05_28_batch_001/labeled_source/supabase_image_review_approve_candidates_part_001_labeled.csv",
]

REASON_NORMALIZATION = {
    "LOW_RESOLUTION": "LOW_RESOLUTION_AFTER_URL_REPAIR",
    "figure is too far away from the camera": "PERSON_TOO_FAR",
    "figure is too far from the camear": "PERSON_TOO_FAR",
    "figure is too far from camera": "PERSON_TOO_FAR",
    "figure is too small": "PERSON_TOO_FAR",
    "too dark": "TOO_DARK",
    "image too dark": "TOO_DARK",
    "too dark to see the garment": "TOO_DARK",
    "bad angle": "BAD_ANGLE_TOP_DOWN",
    "bad camera angle": "BAD_ANGLE_TOP_DOWN",
    "top of the jeans is covered": "GARMENT_TOP_COVERED",
    "jeans are cutoff": "GARMENT_CUT_OFF",
    "can't see full garment": "GARMENT_CUT_OFF",
    "no human": "NO_PERSON_VISIBLE",
    "rotation is incorrect": "BAD_ANGLE_SIDE_OR_TWISTED",
    "image URL is not valid": "INVALID_OR_DEAD_IMAGE_URL",
    "ugly photo": "BACKGROUND_VISUALLY_OFFPUTTING",
}


def read_csv(path: Path) -> list[dict[str, str]]:
    with path.open(newline="", encoding="utf-8-sig") as handle:
        return list(csv.DictReader(handle))


def key_url(value: str) -> str:
    return str(value or "").strip()


def product_url(row: dict[str, str]) -> str:
    return str(row.get("product_page_url_display") or row.get("monetized_product_url_display") or "").strip()


def label_keys(url: str, product: str = "") -> list[str]:
    url = key_url(url)
    product = key_url(product)
    if not url:
        return []
    keys = []
    if product:
        keys.append(f"{url}||{product}")
    keys.append(url)
    return keys


def normalize_decision(value: str) -> str:
    text = str(value or "").strip().upper()
    if text == "APPROVED":
        return "APPROVE"
    if text == "REJECTED":
        return "REJECT"
    return ""


def normalize_reason(value: str) -> str:
    text = str(value or "").strip()
    if not text:
        return ""
    if text in REJECTION_OPTIONS:
        return text
    upper = text.upper()
    if upper in REJECTION_OPTIONS:
        return upper
    lowered = text.lower()
    if lowered in REASON_NORMALIZATION:
        return REASON_NORMALIZATION[lowered]
    for fragment, reason in REASON_NORMALIZATION.items():
        if fragment in lowered:
            return reason
    return "OTHER"


def yes(value: str) -> bool:
    return str(value or "").strip().lower() in {"yes", "y", "true", "1", "not worn by person"}


def add_label(labels: dict[str, dict[str, str]], url: str, decision: str, reason: str, source: str, product: str = "") -> None:
    keys = label_keys(url, product)
    if not keys:
        return
    for key in keys:
        current = labels.get(key, {})
        if current.get("production_decision") == "REJECT" and decision != "REJECT":
            continue
        if decision == "REJECT" or not current.get("production_decision"):
            labels[key] = {
                "production_decision": decision,
                "rejection_reason": reason,
                "review_notes": f"Prefilled from prior ground truth: {source}",
            }


def build_label_lookup() -> dict[str, dict[str, str]]:
    labels: dict[str, dict[str, str]] = {}
    for path in GROUND_TRUTH_FILES:
        if not path.exists():
            continue
        for row in read_csv(path):
            url = row.get("original_url_display", "")
            product = product_url(row)
            decision = normalize_decision(row.get("production_decision", "")) or normalize_decision(row.get("final_human_decision", ""))
            reason = normalize_reason(row.get("rejection_reason", "")) or normalize_reason(row.get("primary_reason_code", ""))
            if decision:
                add_label(labels, url, decision, reason if decision == "REJECT" else "", path.name, product)
                continue

            if yes(row.get("answer_yes_no", "")):
                add_label(labels, url, "REJECT", normalize_reason(row.get("rejection_reason", "")), path.name, product)
                continue

            if yes(row.get("human_not_worn_by_person_yes_no", "")):
                add_label(labels, url, "REJECT", "NOT_WORN_BY_PERSON", path.name, product)
                continue

            if yes(row.get("needs_crop_yes_no", "")):
                add_label(labels, url, "NEEDS_MORE_REVIEW", "NEEDS_CROP", path.name, product)
    return labels


def headers_for(sheet) -> dict[str, int]:
    return {str(sheet.cell(1, column).value): column for column in range(1, sheet.max_column + 1)}


def refresh_rejection_options(workbook) -> None:
    name = "Rejection Reason Options"
    if name in workbook.sheetnames:
        sheet = workbook[name]
        sheet.delete_rows(1, sheet.max_row)
    else:
        sheet = workbook.create_sheet(name)
    sheet.sheet_state = "hidden"
    sheet.cell(1, 1).value = "rejection_reason"
    for index, reason in enumerate(REJECTION_OPTIONS, start=2):
        sheet.cell(index, 1).value = reason


def update_validation(sheet, reason_col: int) -> None:
    kept = []
    reason_letter = sheet.cell(1, reason_col).column_letter
    for validation in sheet.data_validations.dataValidation:
        if str(validation.sqref).startswith(f"{reason_letter}2:"):
            continue
        kept.append(validation)
    sheet.data_validations.dataValidation = kept
    validation = DataValidation(
        type="list",
        formula1=f"'Rejection Reason Options'!$A$2:$A${len(REJECTION_OPTIONS) + 1}",
        allow_blank=True,
    )
    sheet.add_data_validation(validation)
    validation.add(f"{reason_letter}2:{reason_letter}{max(2, sheet.max_row)}")


def prefill_workbook(path: Path, labels: dict[str, dict[str, str]]) -> Counter:
    workbook = load_workbook(path)
    sheet = workbook["Review"] if "Review" in workbook.sheetnames else workbook.active
    header = headers_for(sheet)
    required = ["production_decision", "rejection_reason", "review_notes", "image_url_to_use", "raw_scraped_image_url"]
    missing = [column for column in required if column not in header]
    if missing:
        raise SystemExit(f"{path} missing columns: {missing}")

    refresh_rejection_options(workbook)
    update_validation(sheet, header["rejection_reason"])

    counts = Counter()
    for row_index in range(2, sheet.max_row + 1):
        raw_url = key_url(sheet.cell(row_index, header["raw_scraped_image_url"]).value)
        image_url = key_url(sheet.cell(row_index, header["image_url_to_use"]).value)
        product = key_url(sheet.cell(row_index, header["product_page_url_display"]).value)
        label = (
            labels.get(f"{raw_url}||{product}")
            or labels.get(f"{image_url}||{product}")
            or labels.get(raw_url)
            or labels.get(image_url)
        )
        if not label:
            continue
        current_decision = str(sheet.cell(row_index, header["production_decision"]).value or "").strip()
        if current_decision:
            counts["already_had_decision"] += 1
            continue
        sheet.cell(row_index, header["production_decision"]).value = label["production_decision"]
        sheet.cell(row_index, header["rejection_reason"]).value = label["rejection_reason"]
        sheet.cell(row_index, header["review_notes"]).value = label["review_notes"]
        counts["prefilled"] += 1
        counts[f"prefilled_{label['production_decision']}"] += 1
    workbook.save(path)
    return counts


def main() -> None:
    labels = build_label_lookup()
    total = Counter()
    files = [path for path in sorted(PACKAGE_DIR.glob("*.xlsx")) if path.name not in SKIP_FILES]
    for index, path in enumerate(files, start=1):
        total.update(prefill_workbook(path, labels))
        if index % 25 == 0:
            print(f"processed {index}/{len(files)}", flush=True)
    print(f"prior labeled image lookup: {len(labels)}")
    print(f"workbooks updated: {len(files)}")
    for key, value in total.most_common():
        print(f"{key}: {value}")


if __name__ == "__main__":
    main()
