#!/usr/bin/env python3
"""Add Google Sheets IMAGE formulas to Supabase image review workbooks."""

from __future__ import annotations
import sys

from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


REPO_ROOT = Path(__file__).resolve().parents[4]
PIPELINE_SCRIPTS_DIR = REPO_ROOT / "data-pipelines" / "scripts"
if str(PIPELINE_SCRIPTS_DIR) not in sys.path:
    sys.path.insert(0, str(PIPELINE_SCRIPTS_DIR))

from pipeline_paths import archive_root, cv_annotated_pending_human_review_root  # noqa: E402

LEGACY_OUTPUTS_ARCHIVE = archive_root() / "old_outputs" / "repo_outputs_archive" / "supabase_output_cleanup_2026_05_29"
CV_EXPERIMENTS_DIR = LEGACY_OUTPUTS_ARCHIVE / "cv_experiments"

PACKAGE_DIR = cv_annotated_pending_human_review_root() / "supabase_production_image_review_2026_05_28_s3_refresh_sovrn_prioritized"
PREVIEW_COLUMN_NAME = "image_preview"
PREVIEW_COLUMN_INDEX = 4


def find_header(sheet, header: str) -> int | None:
    for column in range(1, sheet.max_column + 1):
        if sheet.cell(1, column).value == header:
            return column
    return None


def ensure_preview_column(path: Path) -> None:
    workbook = load_workbook(path)
    sheet = workbook["Review"] if "Review" in workbook.sheetnames else workbook.active

    existing_preview_col = find_header(sheet, PREVIEW_COLUMN_NAME)
    if existing_preview_col is None:
        sheet.insert_cols(PREVIEW_COLUMN_INDEX)
        preview_col = PREVIEW_COLUMN_INDEX
        sheet.cell(1, preview_col).value = PREVIEW_COLUMN_NAME
    else:
        preview_col = existing_preview_col
        if preview_col != PREVIEW_COLUMN_INDEX:
            sheet.move_range(
                f"{get_column_letter(preview_col)}1:{get_column_letter(preview_col)}{sheet.max_row}",
                rows=0,
                cols=PREVIEW_COLUMN_INDEX - preview_col,
            )
            preview_col = PREVIEW_COLUMN_INDEX

    image_url_col = find_header(sheet, "image_url_to_use")
    if image_url_col is None:
        raise SystemExit(f"Missing image_url_to_use column in {path}")

    preview_letter = get_column_letter(preview_col)
    image_url_letter = get_column_letter(image_url_col)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    sheet.column_dimensions[preview_letter].width = 22
    for row in range(2, sheet.max_row + 1):
        sheet.cell(row, preview_col).value = f'=IF({image_url_letter}{row}<>"",IMAGE({image_url_letter}{row}),"")'
        sheet.cell(row, preview_col).alignment = Alignment(horizontal="center", vertical="center")
        sheet.row_dimensions[row].height = 92

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(sheet.max_column)}{sheet.max_row}"
    workbook.save(path)


def update_readme() -> None:
    readme = PACKAGE_DIR / "README.md"
    if not readme.exists():
        return
    text = readme.read_text(encoding="utf-8")
    note = (
        "\n## Image Preview\n\n"
        "The fourth column, `image_preview`, uses a Google Sheets `=IMAGE(...)` formula based on "
        "`image_url_to_use`. If previews do not appear immediately after upload/import, wait for "
        "Google Sheets to finish loading external images.\n"
    )
    if "## Image Preview" not in text:
        readme.write_text(text.rstrip() + "\n" + note, encoding="utf-8")


def main() -> None:
    files = sorted(PACKAGE_DIR.glob("*.xlsx"))
    for index, path in enumerate(files, start=1):
        ensure_preview_column(path)
        if index % 25 == 0:
            print(f"updated {index}/{len(files)}", flush=True)
    update_readme()
    print(f"updated {len(files)} workbooks in {PACKAGE_DIR}")


if __name__ == "__main__":
    main()
