#!/usr/bin/env python3
"""Use OpenAI vision to label NOT_WORN_BY_PERSON calibration rows."""

from __future__ import annotations
import sys

import argparse
import csv
import json
import os
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime, timezone
from pathlib import Path
from typing import Dict, List
from urllib.request import Request, urlopen

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation


REPO_ROOT = Path(__file__).resolve().parents[4]
PIPELINE_SCRIPTS_DIR = REPO_ROOT / "data-pipelines" / "scripts"
if str(PIPELINE_SCRIPTS_DIR) not in sys.path:
    sys.path.insert(0, str(PIPELINE_SCRIPTS_DIR))

from pipeline_paths import archive_root, cv_annotated_pending_human_review_root  # noqa: E402

LEGACY_OUTPUTS_ARCHIVE = archive_root() / "old_outputs" / "repo_outputs_archive" / "supabase_output_cleanup_2026_05_29"
CV_EXPERIMENTS_DIR = LEGACY_OUTPUTS_ARCHIVE / "cv_experiments"

DEFAULT_ENV_PATH = REPO_ROOT / ".env"
EXP_DIR = CV_EXPERIMENTS_DIR / "yolo_segmentation_crop_reasons_broad_2026_05_25"
SOURCE_CSV = EXP_DIR / "not_worn_by_person_yes_no_review_queue.csv"
SOURCE_XLSX = EXP_DIR / "not_worn_by_person_yes_no_review_queue.xlsx"
OUT_DIR = EXP_DIR / "openai_not_worn_calibration_2026_05_25"

DEFAULT_MODEL = "gpt-4.1-mini"
PROMPT_VERSION = "not_worn_by_person_yes_no_2026_05_25"
ANSWER_CHOICES = ["YES", "NO", "UNSURE"]

JSON_SCHEMA = {
    "name": "not_worn_by_person_result",
    "schema": {
        "type": "object",
        "additionalProperties": False,
        "properties": {
            "not_worn_by_person": {"type": "string", "enum": ANSWER_CHOICES},
            "confidence": {"type": "number"},
            "visible_person": {"type": "string", "enum": ANSWER_CHOICES},
            "clothing_is_being_worn": {"type": "string", "enum": ANSWER_CHOICES},
            "evidence_summary": {"type": "string"},
        },
        "required": [
            "not_worn_by_person",
            "confidence",
            "visible_person",
            "clothing_is_being_worn",
            "evidence_summary",
        ],
    },
    "strict": True,
}

SYSTEM_PROMPT = """You are labeling clothing review images for a fashion fit-shopping website.

Answer one binary visual question:
Is the clothing NOT being worn by a visible person?

Use YES when the image is product-only, flat-lay, on a hanger, packaging/tag only, on a bed/floor/chair, no visible wearer, or otherwise not shown being worn by a visible person.

Use NO when a real visible person, mannequin, or model is wearing the clothing in the image. A cropped head, mirror selfie, partial body, or obscured face can still be NO if the clothing is visibly being worn.

Use UNSURE only when the image is too ambiguous to tell whether the garment is worn by a visible person.

Ignore whether the image needs crop, is grainy, dark, cluttered, pretty, or otherwise publishable. This task is only about whether the clothing is being worn by a visible person.

Return structured JSON only."""


def load_env_file(path: Path) -> None:
    if not path.exists():
        return
    for raw_line in path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        key = key.strip()
        value = value.strip().strip('"').strip("'")
        if key and key not in os.environ:
            os.environ[key] = value


def read_csv_rows(path: Path) -> List[Dict[str, str]]:
    with path.open(newline="", encoding="utf-8-sig") as handle:
        return list(csv.DictReader(handle))


def xlsx_to_csv(xlsx_path: Path, csv_path: Path) -> None:
    from openpyxl import load_workbook

    workbook = load_workbook(xlsx_path, data_only=False)
    sheet = workbook["Not Worn Review"]
    rows = [[("" if value is None else value) for value in row] for row in sheet.iter_rows(values_only=True)]
    with csv_path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.writer(handle)
        writer.writerows(rows)


def to_float(value: object, default: float = 0.0) -> float:
    try:
        if value is None or str(value).strip() == "":
            return default
        return float(value)
    except ValueError:
        return default


def bool_text(value: object) -> str:
    return "YES" if str(value).strip().lower() == "true" else "NO"


def build_payload(row: Dict[str, str], model: str, image_detail: str, max_output_tokens: int) -> Dict[str, object]:
    prompt = """Label this image for one question only:

Is the clothing NOT being worn by a visible person?

Answer YES, NO, or UNSURE.

Context from the row:
- current primary reason code: {primary}
- current secondary reason code: {secondary}
- product title: {title}
- user/comment context: {comment}

Remember: ignore crop, darkness, graininess, clutter, and publishability unless they make it impossible to tell whether clothing is being worn.
""".format(
        primary=row.get("current_primary_reason_code") or row.get("primary_reason_code") or "(blank)",
        secondary=row.get("current_secondary_reason_code") or row.get("secondary_reason_code") or "(blank)",
        title=row.get("product_title_raw") or "(blank)",
        comment=row.get("user_comment") or row.get("current_labeler_notes") or "(blank)",
    )
    return {
        "model": model,
        "input": [
            {"role": "system", "content": [{"type": "input_text", "text": SYSTEM_PROMPT}]},
            {
                "role": "user",
                "content": [
                    {"type": "input_text", "text": prompt},
                    {"type": "input_image", "image_url": row["original_url_display"], "detail": image_detail},
                ],
            },
        ],
        "text": {
            "format": {
                "type": "json_schema",
                "name": JSON_SCHEMA["name"],
                "schema": JSON_SCHEMA["schema"],
                "strict": JSON_SCHEMA["strict"],
            }
        },
        "max_output_tokens": max_output_tokens,
    }


def do_json_post(url: str, api_key: str, payload: Dict[str, object], timeout_seconds: int) -> Dict[str, object]:
    request = Request(
        url,
        data=json.dumps(payload).encode("utf-8"),
        headers={"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"},
        method="POST",
    )
    with urlopen(request, timeout=timeout_seconds) as response:
        return json.loads(response.read().decode("utf-8"))


def extract_response_text(response_json: Dict[str, object]) -> str:
    output_text = response_json.get("output_text")
    if isinstance(output_text, str) and output_text.strip():
        return output_text.strip()
    outputs = response_json.get("output")
    if isinstance(outputs, list):
        parts: List[str] = []
        for output in outputs:
            if not isinstance(output, dict):
                continue
            content = output.get("content")
            if not isinstance(content, list):
                continue
            for item in content:
                if isinstance(item, dict) and item.get("type") in {"output_text", "text"}:
                    text = item.get("text")
                    if isinstance(text, str) and text.strip():
                        parts.append(text.strip())
        if parts:
            return "\n".join(parts)
    raise ValueError("Could not extract response text")


def classify_row(row: Dict[str, str], args: argparse.Namespace, api_key: str) -> Dict[str, object]:
    base = dict(row)
    base.update(
        {
            "openai_model": args.model,
            "openai_prompt_version": PROMPT_VERSION,
            "openai_request_status": "",
            "openai_not_worn_by_person": "",
            "openai_confidence": "",
            "openai_visible_person": "",
            "openai_clothing_is_being_worn": "",
            "openai_evidence_summary": "",
            "openai_error_message": "",
            "openai_input_tokens": "",
            "openai_output_tokens": "",
            "openai_total_tokens": "",
        }
    )
    try:
        payload = build_payload(row, args.model, args.image_detail, args.max_output_tokens)
        response = do_json_post("https://api.openai.com/v1/responses", api_key, payload, args.timeout_seconds)
        result = json.loads(extract_response_text(response))
        usage = response.get("usage") if isinstance(response.get("usage"), dict) else {}
        base.update(
            {
                "openai_request_status": "ok",
                "openai_not_worn_by_person": result["not_worn_by_person"],
                "openai_confidence": round(float(result["confidence"]), 4),
                "openai_visible_person": result["visible_person"],
                "openai_clothing_is_being_worn": result["clothing_is_being_worn"],
                "openai_evidence_summary": result["evidence_summary"],
                "openai_input_tokens": usage.get("input_tokens", ""),
                "openai_output_tokens": usage.get("output_tokens", ""),
                "openai_total_tokens": usage.get("total_tokens", ""),
            }
        )
    except Exception as exc:
        base.update({"openai_request_status": "error", "openai_error_message": repr(exc)})
    return base


def write_csv(path: Path, rows: List[Dict[str, object]]) -> None:
    if not rows:
        return
    fieldnames = sorted({key for row in rows for key in row})
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def style_header(sheet) -> None:
    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)
    for cell in sheet[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(wrap_text=True, vertical="top")


def create_human_review_workbook(rows: List[Dict[str, object]], path: Path) -> None:
    headers = [
        "human_not_worn_by_person_yes_no",
        "image_preview",
        "review_reason",
        "openai_not_worn_by_person",
        "openai_confidence",
        "openai_evidence_summary",
        "model_predicted_not_worn",
        "current_primary_reason_code",
        "current_secondary_reason_code",
        "original_url_display",
        "product_page_url_display",
        "review_row_key",
    ]
    review_rows = []
    for row in rows:
        openai_answer = str(row.get("openai_not_worn_by_person") or "")
        confidence = to_float(row.get("openai_confidence"), default=0.0)
        yolo_answer = bool_text(row.get("model_predicted_not_worn"))
        reasons = []
        if openai_answer == "UNSURE" or confidence < 0.85:
            reasons.append("low_confidence_or_unsure")
        if openai_answer in {"YES", "NO"} and openai_answer != yolo_answer:
            reasons.append("openai_yolo_disagree")
        if str(row.get("current_primary_reason_code") or "") == "NOT_WORN_BY_PERSON" and openai_answer == "NO":
            reasons.append("openai_disagrees_with_existing_label")
        if reasons:
            item = dict(row)
            item["review_reason"] = ";".join(reasons)
            review_rows.append(item)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Human Review"
    sheet.append(headers)
    for row in review_rows:
        sheet.append([row.get(header, "") for header in headers])
    style_header(sheet)
    sheet.freeze_panes = "A2"
    if sheet.max_row > 1:
        sheet.auto_filter.ref = f"A1:L{sheet.max_row}"
        validation = DataValidation(type="list", formula1='"YES,NO,UNSURE"', allow_blank=True)
        sheet.add_data_validation(validation)
        validation.add(f"A2:A{sheet.max_row}")
    widths = {"A": 30, "B": 26, "C": 34, "D": 22, "E": 18, "F": 68, "G": 22, "H": 26, "I": 26, "J": 42, "K": 42, "L": 28}
    for col, width in widths.items():
        sheet.column_dimensions[col].width = width
    for row in sheet.iter_rows(min_row=2):
        sheet.row_dimensions[row[0].row].height = 72
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    workbook.save(path)


def write_report(rows: List[Dict[str, object]], review_path: Path, report_path: Path) -> None:
    ok_rows = [row for row in rows if row.get("openai_request_status") == "ok"]
    errors = [row for row in rows if row.get("openai_request_status") != "ok"]
    yes = sum(1 for row in ok_rows if row.get("openai_not_worn_by_person") == "YES")
    no = sum(1 for row in ok_rows if row.get("openai_not_worn_by_person") == "NO")
    unsure = sum(1 for row in ok_rows if row.get("openai_not_worn_by_person") == "UNSURE")
    low_conf = sum(1 for row in ok_rows if to_float(row.get("openai_confidence"), default=0.0) < 0.85)
    disagreements = sum(
        1
        for row in ok_rows
        if row.get("openai_not_worn_by_person") in {"YES", "NO"}
        and row.get("openai_not_worn_by_person") != bool_text(row.get("model_predicted_not_worn"))
    )
    total_tokens = sum(int(row.get("openai_total_tokens") or 0) for row in ok_rows)
    lines = [
        "# OpenAI NOT_WORN_BY_PERSON Calibration",
        "",
        f"- rows sent: `{len(rows)}`",
        f"- successful responses: `{len(ok_rows)}`",
        f"- errors: `{len(errors)}`",
        f"- OpenAI answers: `{{'YES': {yes}, 'NO': {no}, 'UNSURE': {unsure}}}`",
        f"- low-confidence rows (<0.85): `{low_conf}`",
        f"- OpenAI vs YOLO disagreements: `{disagreements}`",
        f"- total tokens: `{total_tokens}`",
        f"- human review workbook: `{review_path}`",
        "",
        "## Recommendation",
        "",
        "Review only the generated disagreement/low-confidence workbook. If the review set is small and OpenAI is mostly correct, use OpenAI vision as the primary labeler for `NOT_WORN_BY_PERSON`, with YOLO `seg_person_count == 0` as a cheap prefilter and audit trigger.",
        "",
    ]
    report_path.write_text("\n".join(lines), encoding="utf-8")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument("--env-file", type=Path, default=DEFAULT_ENV_PATH)
    parser.add_argument("--model", default=DEFAULT_MODEL)
    parser.add_argument("--image-detail", default="low", choices=["low", "high", "auto"])
    parser.add_argument("--max-output-tokens", type=int, default=180)
    parser.add_argument("--timeout-seconds", type=int, default=60)
    parser.add_argument("--max-workers", type=int, default=4)
    parser.add_argument("--limit", type=int, default=0)
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    load_env_file(args.env_file)
    api_key = os.environ.get("OPENAI_API_KEY", "").strip()
    if not api_key:
        raise SystemExit("OPENAI_API_KEY is not set")

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    if not SOURCE_CSV.exists():
        xlsx_to_csv(SOURCE_XLSX, SOURCE_CSV)
    rows = [
        row for row in read_csv_rows(SOURCE_CSV)
        if row.get("original_url_display", "").strip()
    ]
    if args.limit:
        rows = rows[: args.limit]

    results: List[Dict[str, object]] = []
    with ThreadPoolExecutor(max_workers=args.max_workers) as executor:
        futures = {executor.submit(classify_row, row, args, api_key): idx for idx, row in enumerate(rows, start=1)}
        for future in as_completed(futures):
            results.append(future.result())
            if len(results) % 10 == 0:
                print(f"processed {len(results)}/{len(rows)}", flush=True)
            time.sleep(0.02)
    results.sort(key=lambda row: str(row.get("review_row_key", "")))

    run_stamp = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
    row_results = OUT_DIR / f"openai_not_worn_results_{run_stamp}.csv"
    review_workbook = OUT_DIR / f"openai_not_worn_human_review_{run_stamp}.xlsx"
    report = OUT_DIR / f"openai_not_worn_report_{run_stamp}.md"
    latest_results = OUT_DIR / "openai_not_worn_results_latest.csv"
    latest_review = OUT_DIR / "openai_not_worn_human_review_latest.xlsx"
    latest_report = OUT_DIR / "openai_not_worn_report_latest.md"

    write_csv(row_results, results)
    write_csv(latest_results, results)
    create_human_review_workbook(results, review_workbook)
    create_human_review_workbook(results, latest_review)
    write_report(results, review_workbook, report)
    write_report(results, latest_review, latest_report)
    print(report)
    print(review_workbook)


if __name__ == "__main__":
    main()
