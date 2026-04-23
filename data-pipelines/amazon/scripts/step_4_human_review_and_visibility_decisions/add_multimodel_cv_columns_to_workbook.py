#!/usr/bin/env python3
import argparse
import os
import shutil
import sys
from copy import copy
from io import BytesIO
from pathlib import Path
from typing import Optional


PIPELINE_ROOT = Path(__file__).resolve().parents[2]
MODEL_DIR = PIPELINE_ROOT / "models"
DEFAULT_SCRFD_MODEL_ROOT = PIPELINE_ROOT.parents[1] / ".codex_vendor" / "insightface_models"
DEFAULT_VENDOR_DIR = PIPELINE_ROOT.parents[1] / ".codex_vendor"

DESIRED_COLUMNS = [
    "has_face_scrfd",
    "has_face_blazeface",
    "person_count_yolo_detect",
    "main_person_height_pct_yolo_detect",
    "main_person_bbox_area_pct_yolo_detect",
    "person_count_yolo_pose",
    "main_person_height_pct_yolo_pose",
    "main_person_bbox_area_pct_yolo_pose",
    "body_coverage_score_yolo_pose",
    "body_coverage_score_mediapipe_pose",
    "head_visible_mediapipe_pose",
    "shoulders_visible_mediapipe_pose",
    "hips_visible_mediapipe_pose",
    "knees_visible_mediapipe_pose",
    "ankles_visible_mediapipe_pose",
    "feet_visible_mediapipe_pose",
]

YOLO_POSE_KEYPOINT_WEIGHTS = {
    0: 1.0,   # nose
    5: 1.5,   # left shoulder
    6: 1.5,   # right shoulder
    11: 1.5,  # left hip
    12: 1.5,  # right hip
    13: 1.0,  # left knee
    14: 1.0,  # right knee
    15: 1.5,  # left ankle
    16: 1.5,  # right ankle
}

MEDIAPIPE_POSE_SCORE_GROUPS = {
    "head_visible_mediapipe_pose": (0, 1, 2, 3, 4, 7, 8),
    "shoulders_visible_mediapipe_pose": (11, 12),
    "hips_visible_mediapipe_pose": (23, 24),
    "knees_visible_mediapipe_pose": (25, 26),
    "ankles_visible_mediapipe_pose": (27, 28),
    "feet_visible_mediapipe_pose": (29, 30, 31, 32),
}
MEDIAPIPE_POSE_SCORE_WEIGHTS = {
    "head_visible_mediapipe_pose": 1.0,
    "shoulders_visible_mediapipe_pose": 1.5,
    "hips_visible_mediapipe_pose": 1.5,
    "knees_visible_mediapipe_pose": 1.0,
    "ankles_visible_mediapipe_pose": 1.5,
    "feet_visible_mediapipe_pose": 1.0,
}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Add several computer-vision model outputs to a Step 4 review workbook "
            "in one pass using the image URL in original_url_display."
        )
    )
    parser.add_argument(
        "--workbook",
        type=Path,
        required=True,
        help="Path to the source .xlsx workbook.",
    )
    parser.add_argument(
        "--output",
        type=Path,
        help="Optional output workbook path. If omitted, updates the source workbook in place.",
    )
    parser.add_argument(
        "--sheet",
        default="",
        help="Optional worksheet name. Defaults to the workbook's first sheet.",
    )
    parser.add_argument(
        "--url-column",
        default="original_url_display",
        help="Header name containing the image URL.",
    )
    parser.add_argument(
        "--anchor-column",
        default="body_coverage_score_yolo_pose",
        help="Insert missing multi-model columns immediately after this header when possible.",
    )
    parser.add_argument(
        "--fallback-anchor-column",
        default="full_lower_body_visible",
        help="Secondary anchor used when the primary anchor is absent.",
    )
    parser.add_argument(
        "--timeout",
        type=float,
        default=20.0,
        help="Per-image download timeout in seconds.",
    )
    parser.add_argument(
        "--limit",
        type=int,
        default=0,
        help="Optional row limit for smoke tests. 0 means process every row.",
    )
    parser.add_argument(
        "--backup-suffix",
        default=".bak",
        help="Suffix for the workbook backup created before saving.",
    )
    parser.add_argument(
        "--det-size",
        type=int,
        default=640,
        help="Detection input size used by SCRFD.",
    )
    parser.add_argument(
        "--min-face-score",
        type=float,
        default=0.6,
        help="Minimum accepted face-detection score for SCRFD and BlazeFace.",
    )
    parser.add_argument(
        "--min-person-confidence",
        type=float,
        default=0.35,
        help="Minimum YOLO confidence used when counting person detections.",
    )
    parser.add_argument(
        "--min-pose-keypoint-confidence",
        type=float,
        default=0.35,
        help="Minimum YOLO pose keypoint confidence used in the YOLO coverage score.",
    )
    parser.add_argument(
        "--min-mp-visibility",
        type=float,
        default=0.5,
        help="Minimum MediaPipe landmark visibility considered visible.",
    )
    parser.add_argument(
        "--min-mp-presence",
        type=float,
        default=0.5,
        help="Minimum MediaPipe landmark presence considered visible when available.",
    )
    parser.add_argument(
        "--vendor-dir",
        type=Path,
        default=DEFAULT_VENDOR_DIR,
        help="Directory containing vendored model runtimes such as yolo_test and mediapipe_test.",
    )
    parser.add_argument(
        "--scrfd-model-root",
        type=Path,
        default=DEFAULT_SCRFD_MODEL_ROOT,
        help="Directory used to store InsightFace model files.",
    )
    parser.add_argument(
        "--yolo-detect-model",
        default="yolov8n.pt",
        help="YOLO detect model path or model name.",
    )
    parser.add_argument(
        "--yolo-pose-model",
        default="yolov8n-pose.pt",
        help="YOLO pose model path or model name.",
    )
    parser.add_argument(
        "--verbose",
        action="store_true",
        help="Print progress and per-row failures to stderr.",
    )
    return parser.parse_args()


def bootstrap_vendor_paths(vendor_dir: Path) -> None:
    vendor_subdirs = [
        vendor_dir / "yolo_test",
        vendor_dir / "mediapipe_test",
        vendor_dir / "scrfd_test",
    ]
    for path in reversed(vendor_subdirs):
        if path.exists():
            sys.path.insert(0, str(path))


def import_dependencies():
    os.environ.setdefault("YOLO_CONFIG_DIR", "/tmp/Ultralytics")
    os.environ.setdefault("MPLCONFIGDIR", "/tmp/matplotlib-codex")
    try:
        import cv2  # type: ignore
        import mediapipe as mp  # type: ignore
        import numpy as np  # type: ignore
        import requests  # type: ignore
        from insightface.app import FaceAnalysis  # type: ignore
        from mediapipe.tasks.python.core.base_options import BaseOptions as MPBaseOptions  # type: ignore
        from mediapipe.tasks.python.vision import face_detector as mp_face_detector  # type: ignore
        from mediapipe.tasks.python.vision import pose_landmarker as mp_pose_landmarker  # type: ignore
        from openpyxl import load_workbook  # type: ignore
        from PIL import Image  # type: ignore
        from ultralytics import YOLO  # type: ignore
    except ModuleNotFoundError as exc:
        missing = exc.name or "required package"
        raise SystemExit(
            "Missing dependency: "
            f"{missing}. This script expects vendored runtimes under .codex_vendor "
            "for yolo_test, mediapipe_test, and scrfd_test."
        ) from exc
    return cv2, mp, np, requests, FaceAnalysis, MPBaseOptions, mp_face_detector, mp_pose_landmarker, load_workbook, Image, YOLO


def bool_to_str(value: bool) -> str:
    return "true" if value else "false"


def round_or_blank(value):
    if value is None:
        return ""
    return round(float(value), 3)


def find_header_map(worksheet) -> dict[str, int]:
    header_map: dict[str, int] = {}
    for index in range(1, worksheet.max_column + 1):
        value = worksheet.cell(row=1, column=index).value
        if isinstance(value, str) and value.strip():
            header_map[value.strip()] = index
    return header_map


def copy_column_style(worksheet, source_col: int, target_col: int, max_row: int) -> None:
    for row in range(1, max_row + 1):
        source = worksheet.cell(row=row, column=source_col)
        target = worksheet.cell(row=row, column=target_col)
        if source.has_style:
            target._style = copy(source._style)
            target.font = copy(source.font)
            target.fill = copy(source.fill)
            target.border = copy(source.border)
            target.alignment = copy(source.alignment)
            target.protection = copy(source.protection)
            target.number_format = source.number_format


def insert_missing_columns(
    worksheet,
    headers: dict[str, int],
    anchor_column: str,
    fallback_anchor_column: str,
) -> dict[str, int]:
    if all(column in headers for column in DESIRED_COLUMNS):
        return headers

    insert_after = headers.get(anchor_column) or headers.get(fallback_anchor_column) or worksheet.max_column
    reference_col = min(insert_after, worksheet.max_column)

    for column_name in DESIRED_COLUMNS:
        if column_name in headers:
            insert_after = headers[column_name]
            continue
        worksheet.insert_cols(insert_after + 1)
        target_col = insert_after + 1
        worksheet.cell(row=1, column=target_col, value=column_name)
        copy_column_style(worksheet, reference_col, target_col, worksheet.max_row)
        headers = find_header_map(worksheet)
        insert_after = headers[column_name]

    return headers


def fetch_rgb_image(url: str, timeout: float, session, Image):
    response = session.get(
        url,
        timeout=timeout,
        headers={
            "User-Agent": (
                "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/123.0 Safari/537.36"
            )
        },
    )
    response.raise_for_status()
    return Image.open(BytesIO(response.content)).convert("RGB")


def compute_yolo_coverage_score(keypoint_confidences, min_confidence: float) -> float:
    total_weight = float(sum(YOLO_POSE_KEYPOINT_WEIGHTS.values()))
    visible_weight = 0.0
    for index, weight in YOLO_POSE_KEYPOINT_WEIGHTS.items():
        confidence = float(keypoint_confidences[index])
        if confidence >= min_confidence:
            visible_weight += weight
    return round((visible_weight / total_weight) * 100.0, 1)


def extract_person_metrics(result, image_width: int, image_height: int, min_person_confidence: float) -> tuple[list[int], list[list[float]]]:
    valid_indexes: list[int] = []
    boxes = []
    if result.boxes is None or not len(result.boxes):
        return valid_indexes, boxes

    boxes = result.boxes.xyxy.tolist()
    for index, (cls_value, confidence) in enumerate(zip(result.boxes.cls.tolist(), result.boxes.conf.tolist())):
        if int(cls_value) == 0 and float(confidence) >= min_person_confidence:
            valid_indexes.append(index)
    return valid_indexes, boxes


def summarize_main_person(
    valid_indexes: list[int],
    boxes: list[list[float]],
    image_width: int,
    image_height: int,
) -> tuple[int, float, float, Optional[int]]:
    if not valid_indexes:
        return 0, 0.0, 0.0, None

    best_index = max(
        valid_indexes,
        key=lambda idx: max(0.0, float(boxes[idx][2] - boxes[idx][0])) * max(0.0, float(boxes[idx][3] - boxes[idx][1])),
    )
    x1, y1, x2, y2 = [float(value) for value in boxes[best_index]]
    bbox_width = max(0.0, x2 - x1)
    bbox_height = max(0.0, y2 - y1)
    height_pct = bbox_height / float(image_height) if image_height else 0.0
    area_pct = (bbox_width * bbox_height) / float(image_width * image_height) if image_width and image_height else 0.0
    return len(valid_indexes), height_pct, area_pct, best_index


def analyze_yolo_detect(image_rgb, model, np, min_person_confidence: float) -> dict[str, object]:
    result = model.predict(np.array(image_rgb), verbose=False, device="cpu")[0]
    image_height, image_width = image_rgb.size[1], image_rgb.size[0]
    valid_indexes, boxes = extract_person_metrics(result, image_width, image_height, min_person_confidence)
    person_count, height_pct, area_pct, _best_index = summarize_main_person(
        valid_indexes,
        boxes,
        image_width,
        image_height,
    )
    return {
        "person_count_yolo_detect": person_count,
        "main_person_height_pct_yolo_detect": round_or_blank(height_pct),
        "main_person_bbox_area_pct_yolo_detect": round_or_blank(area_pct),
    }


def analyze_yolo_pose(image_rgb, model, np, min_person_confidence: float, min_keypoint_confidence: float) -> dict[str, object]:
    result = model.predict(np.array(image_rgb), verbose=False, device="cpu")[0]
    image_height, image_width = image_rgb.size[1], image_rgb.size[0]
    valid_indexes, boxes = extract_person_metrics(result, image_width, image_height, min_person_confidence)
    person_count, height_pct, area_pct, best_index = summarize_main_person(
        valid_indexes,
        boxes,
        image_width,
        image_height,
    )

    body_coverage_score = 0.0
    if best_index is not None and getattr(result, "keypoints", None) is not None and len(result.keypoints) > best_index:
        confidences = getattr(result.keypoints, "conf", None)
        if confidences is not None:
            body_coverage_score = compute_yolo_coverage_score(
                confidences[best_index].tolist(),
                min_keypoint_confidence,
            )

    return {
        "person_count_yolo_pose": person_count,
        "main_person_height_pct_yolo_pose": round_or_blank(height_pct),
        "main_person_bbox_area_pct_yolo_pose": round_or_blank(area_pct),
        "body_coverage_score_yolo_pose": round_or_blank(body_coverage_score),
    }


def analyze_scrfd(image_bgr, detector, min_face_score: float) -> dict[str, str]:
    faces = detector.get(image_bgr)
    has_face = any(float(getattr(face, "det_score", 0.0)) >= min_face_score for face in faces or [])
    return {"has_face_scrfd": bool_to_str(has_face)}


def analyze_blazeface(mp_image, blaze_detector, min_face_score: float) -> dict[str, str]:
    if blaze_detector is None:
        return {"has_face_blazeface": ""}
    result = blaze_detector.detect(mp_image)
    detections = getattr(result, "detections", None) or []
    has_face = False
    for detection in detections:
        categories = getattr(detection, "categories", None) or []
        if not categories:
            has_face = True
            break
        if any(float(getattr(category, "score", 0.0)) >= min_face_score for category in categories):
            has_face = True
            break
    return {"has_face_blazeface": bool_to_str(has_face)}


def landmark_visible(landmark, min_visibility: float, min_presence: float) -> bool:
    visibility = float(getattr(landmark, "visibility", 0.0))
    presence = float(getattr(landmark, "presence", 1.0))
    return visibility >= min_visibility and presence >= min_presence


def analyze_mediapipe_pose(mp_image, mp_pose_detector, min_visibility: float, min_presence: float) -> dict[str, object]:
    if mp_pose_detector is None:
        return {
            "body_coverage_score_mediapipe_pose": "",
            "head_visible_mediapipe_pose": "",
            "shoulders_visible_mediapipe_pose": "",
            "hips_visible_mediapipe_pose": "",
            "knees_visible_mediapipe_pose": "",
            "ankles_visible_mediapipe_pose": "",
            "feet_visible_mediapipe_pose": "",
        }
    result = mp_pose_detector.detect(mp_image)
    pose_landmarks = getattr(result, "pose_landmarks", None) or []
    if not pose_landmarks:
        return {
            "body_coverage_score_mediapipe_pose": 0.0,
            "head_visible_mediapipe_pose": "false",
            "shoulders_visible_mediapipe_pose": "false",
            "hips_visible_mediapipe_pose": "false",
            "knees_visible_mediapipe_pose": "false",
            "ankles_visible_mediapipe_pose": "false",
            "feet_visible_mediapipe_pose": "false",
        }

    landmark_list = pose_landmarks[0]
    outputs: dict[str, object] = {}
    visible_weight = 0.0
    total_weight = float(sum(MEDIAPIPE_POSE_SCORE_WEIGHTS.values()))

    for column_name, indices in MEDIAPIPE_POSE_SCORE_GROUPS.items():
        group_visible = all(
            landmark_visible(landmark_list[index], min_visibility, min_presence)
            for index in indices
        )
        outputs[column_name] = bool_to_str(group_visible)
        if group_visible:
            visible_weight += MEDIAPIPE_POSE_SCORE_WEIGHTS[column_name]

    outputs["body_coverage_score_mediapipe_pose"] = round_or_blank((visible_weight / total_weight) * 100.0)
    return outputs


def build_model_bundle(args, cv2, mp, np, FaceAnalysis, MPBaseOptions, mp_face_detector, mp_pose_landmarker, YOLO):
    scrfd_model_root = args.scrfd_model_root.resolve()
    scrfd_model_root.mkdir(parents=True, exist_ok=True)

    scrfd = FaceAnalysis(
        name="buffalo_l",
        root=str(scrfd_model_root),
        providers=["CPUExecutionProvider"],
    )
    scrfd.prepare(ctx_id=-1, det_size=(args.det_size, args.det_size))

    warnings: list[str] = []

    blaze_detector = None
    blaze_model_path = MODEL_DIR / "blaze_face_short_range.tflite"
    if not blaze_model_path.exists():
        warnings.append(f"BlazeFace model not found: {blaze_model_path}")
    else:
        try:
            blaze_options = mp_face_detector.FaceDetectorOptions(
                base_options=MPBaseOptions(
                    model_asset_path=str(blaze_model_path),
                    delegate=MPBaseOptions.Delegate.CPU,
                )
            )
            blaze_detector = mp_face_detector.FaceDetector.create_from_options(blaze_options)
        except Exception as exc:  # noqa: BLE001
            warnings.append(f"BlazeFace unavailable in this environment: {exc}")

    mp_pose_detector = None
    pose_model_path = MODEL_DIR / "pose_landmarker_full.task"
    if not pose_model_path.exists():
        warnings.append(f"MediaPipe pose model not found: {pose_model_path}")
    else:
        try:
            pose_options = mp_pose_landmarker.PoseLandmarkerOptions(
                base_options=MPBaseOptions(
                    model_asset_path=str(pose_model_path),
                    delegate=MPBaseOptions.Delegate.CPU,
                )
            )
            mp_pose_detector = mp_pose_landmarker.PoseLandmarker.create_from_options(pose_options)
        except Exception as exc:  # noqa: BLE001
            warnings.append(f"MediaPipe pose unavailable in this environment: {exc}")

    return {
        "scrfd": scrfd,
        "blazeface": blaze_detector,
        "yolo_detect": YOLO(args.yolo_detect_model),
        "yolo_pose": YOLO(args.yolo_pose_model),
        "mediapipe_pose": mp_pose_detector,
        "np": np,
        "cv2": cv2,
        "warnings": warnings,
    }


def blank_outputs() -> dict[str, str]:
    return {column_name: "" for column_name in DESIRED_COLUMNS}


def analyze_image(image_rgb, model_bundle, args, np):
    image_bgr = np.array(image_rgb)[:, :, ::-1]
    mp_image = model_bundle["mp_image_cls"](
        image_format=model_bundle["mp_image_format"].SRGB,
        data=np.array(image_rgb),
    )
    outputs: dict[str, object] = {}
    outputs.update(analyze_scrfd(image_bgr, model_bundle["scrfd"], args.min_face_score))
    outputs.update(analyze_blazeface(mp_image, model_bundle["blazeface"], args.min_face_score))
    outputs.update(analyze_yolo_detect(image_rgb, model_bundle["yolo_detect"], np, args.min_person_confidence))
    outputs.update(
        analyze_yolo_pose(
            image_rgb,
            model_bundle["yolo_pose"],
            np,
            args.min_person_confidence,
            args.min_pose_keypoint_confidence,
        )
    )
    outputs.update(
        analyze_mediapipe_pose(
            mp_image,
            model_bundle["mediapipe_pose"],
            args.min_mp_visibility,
            args.min_mp_presence,
        )
    )
    return outputs


def main() -> None:
    args = parse_args()
    bootstrap_vendor_paths(args.vendor_dir.resolve())
    cv2, mp, np, requests, FaceAnalysis, MPBaseOptions, mp_face_detector, mp_pose_landmarker, load_workbook, Image, YOLO = import_dependencies()

    workbook_path = args.workbook.resolve()
    if not workbook_path.exists():
        raise SystemExit(f"Workbook not found: {workbook_path}")

    output_path = args.output.resolve() if args.output else workbook_path
    if output_path != workbook_path:
        output_path.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(workbook_path, output_path)

    backup_path = output_path.with_name(output_path.name + args.backup_suffix)
    shutil.copy2(output_path, backup_path)
    if args.verbose:
        print(f"Backup created: {backup_path}", file=sys.stderr)

    workbook = load_workbook(output_path)
    worksheet = workbook[args.sheet] if args.sheet else workbook[workbook.sheetnames[0]]
    headers = find_header_map(worksheet)

    if args.url_column not in headers:
        raise SystemExit(f"Required URL column not found: {args.url_column}")

    headers = insert_missing_columns(
        worksheet,
        headers,
        args.anchor_column,
        args.fallback_anchor_column,
    )

    model_bundle = build_model_bundle(args, cv2, mp, np, FaceAnalysis, MPBaseOptions, mp_face_detector, mp_pose_landmarker, YOLO)
    model_bundle["mp_image_cls"] = mp.Image
    model_bundle["mp_image_format"] = mp.ImageFormat
    if args.verbose:
        for warning in model_bundle.get("warnings", []):
            print(f"[warning] {warning}", file=sys.stderr)
    session = requests.Session()
    url_col = headers[args.url_column]
    processed = 0

    for row in range(2, worksheet.max_row + 1):
        url_value = worksheet.cell(row=row, column=url_col).value
        url = str(url_value or "").strip()
        if not url:
            for column_name in DESIRED_COLUMNS:
                worksheet.cell(row=row, column=headers[column_name], value="")
            continue

        if args.limit > 0 and processed >= args.limit:
            break

        try:
            image_rgb = fetch_rgb_image(url, args.timeout, session, Image)
            outputs = analyze_image(image_rgb, model_bundle, args, np)
            for column_name in DESIRED_COLUMNS:
                worksheet.cell(row=row, column=headers[column_name], value=outputs.get(column_name, ""))
        except Exception as exc:  # noqa: BLE001
            for column_name, value in blank_outputs().items():
                worksheet.cell(row=row, column=headers[column_name], value=value)
            if args.verbose:
                print(f"[row {row}] failed for {url}: {exc}", file=sys.stderr)

        processed += 1
        if args.verbose and processed % 25 == 0:
            print(f"Processed {processed} images", file=sys.stderr)

    workbook.save(output_path)
    if args.verbose:
        print(f"Saved workbook: {output_path}", file=sys.stderr)


if __name__ == "__main__":
    main()
