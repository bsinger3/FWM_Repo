#!/usr/bin/env python3
import argparse
import os
import shutil
import sys
import urllib.request
from copy import copy
from io import BytesIO
from pathlib import Path


PIPELINE_ROOT = Path(__file__).resolve().parents[2]
DEFAULT_VENDOR_DIR = PIPELINE_ROOT.parents[1] / ".codex_vendor"
DEFAULT_MODEL_DIR = PIPELINE_ROOT / "models"
DEFAULT_PROTO_URL = "https://raw.githubusercontent.com/opencv/opencv/master/samples/dnn/face_detector/deploy.prototxt"
DEFAULT_MODEL_URL = "https://raw.githubusercontent.com/opencv/opencv_3rdparty/dnn_samples_face_detector_20170830/res10_300x300_ssd_iter_140000.caffemodel"


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Add an OpenCV DNN SSD face-presence column to a review workbook by "
            "running the detector against original_url_display."
        )
    )
    parser.add_argument("--workbook", type=Path, required=True, help="Path to the .xlsx workbook to update in place.")
    parser.add_argument("--sheet", default="", help="Optional worksheet name. Defaults to the workbook's first sheet.")
    parser.add_argument("--url-column", default="original_url_display", help="Header name containing the image URL.")
    parser.add_argument("--anchor-column", default="has_face_scrfd", help="Insert the new column immediately after this header when possible.")
    parser.add_argument("--fallback-anchor-column", default="has_face_yunet", help="Secondary anchor used when the primary anchor is absent.")
    parser.add_argument("--output-column", default="has_face_opencv_dnn_ssd", help="Header name for the new face-detection column.")
    parser.add_argument("--timeout", type=float, default=20.0, help="Per-image download timeout in seconds.")
    parser.add_argument("--backup-suffix", default=".bak", help="Suffix for the workbook backup created before saving.")
    parser.add_argument("--min-score", type=float, default=0.6, help="Minimum accepted DNN face score.")
    parser.add_argument("--input-width", type=int, default=300, help="Detector input width.")
    parser.add_argument("--input-height", type=int, default=300, help="Detector input height.")
    parser.add_argument("--proto-path", type=Path, help="Optional existing deploy.prototxt path to use directly.")
    parser.add_argument("--model-path", type=Path, help="Optional existing .caffemodel path to use directly.")
    parser.add_argument("--proto-url", default=DEFAULT_PROTO_URL, help="Direct URL used when the prototxt is not already cached.")
    parser.add_argument("--model-url", default=DEFAULT_MODEL_URL, help="Direct URL used when the Caffe model is not already cached.")
    parser.add_argument("--vendor-dir", type=Path, default=DEFAULT_VENDOR_DIR, help="Directory containing the vendored cv2 runtime.")
    parser.add_argument("--model-dir", type=Path, default=DEFAULT_MODEL_DIR, help="Directory used to store downloaded detector files.")
    parser.add_argument("--limit", type=int, default=0, help="Optional row limit for smoke tests. 0 means process every row.")
    parser.add_argument("--verbose", action="store_true", help="Print progress and failures to stderr.")
    return parser.parse_args()


def bootstrap_vendor_paths(vendor_dir: Path) -> None:
    runtime_dir = vendor_dir / "scrfd_test"
    if runtime_dir.exists():
        sys.path.insert(0, str(runtime_dir))


def import_dependencies():
    try:
        import cv2  # type: ignore
        import numpy as np  # type: ignore
        import requests  # type: ignore
        from openpyxl import load_workbook  # type: ignore
        from PIL import Image  # type: ignore
    except ModuleNotFoundError as exc:
        missing = exc.name or "required package"
        raise SystemExit(
            "Missing dependency: "
            f"{missing}. This script expects the vendored scrfd_test runtime under .codex_vendor."
        ) from exc
    return cv2, np, requests, load_workbook, Image


def bool_to_str(value: bool) -> str:
    return "true" if value else "false"


def find_header_map(worksheet) -> dict[str, int]:
    header_map: dict[str, int] = {}
    for index in range(1, worksheet.max_column + 1):
        value = worksheet.cell(row=1, column=index).value
        if isinstance(value, str) and value.strip():
            header_map[value.strip()] = index
    return header_map


def copy_column_style(worksheet, source_col: int, target_col: int, max_row: int) -> None:
    for row in range(1, max_row + 1):
        source = worksheet.cell(row=row, column=source_col)
        target = worksheet.cell(row=row, column=target_col)
        if source.has_style:
            target._style = copy(source._style)
            target.font = copy(source.font)
            target.fill = copy(source.fill)
            target.border = copy(source.border)
            target.alignment = copy(source.alignment)
            target.protection = copy(source.protection)
            target.number_format = source.number_format


def ensure_output_column(worksheet, headers: dict[str, int], output_column: str, anchor_column: str, fallback_anchor_column: str) -> tuple[dict[str, int], int]:
    if output_column in headers:
        return headers, headers[output_column]

    insert_after = headers.get(anchor_column) or headers.get(fallback_anchor_column) or worksheet.max_column
    worksheet.insert_cols(insert_after + 1)
    output_col = insert_after + 1
    worksheet.cell(row=1, column=output_col, value=output_column)
    copy_column_style(worksheet, insert_after, output_col, worksheet.max_row)
    headers = find_header_map(worksheet)
    return headers, headers[output_column]


def ensure_file(path: Path, url: str) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    if path.exists():
        return path
    urllib.request.urlretrieve(url, str(path))
    return path


def fetch_bgr_image(url: str, timeout: float, session, Image, np):
    response = session.get(
        url,
        timeout=timeout,
        headers={
            "User-Agent": (
                "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/123.0 Safari/537.36"
            )
        },
    )
    response.raise_for_status()
    image = Image.open(BytesIO(response.content)).convert("RGB")
    rgb = np.array(image)
    return rgb[:, :, ::-1]


def build_detector(cv2, proto_path: Path, model_path: Path):
    return cv2.dnn.readNetFromCaffe(str(proto_path), str(model_path))


def detect_face_opencv_dnn_ssd(cv2, image_bgr, net, input_width: int, input_height: int, min_score: float) -> bool:
    blob = cv2.dnn.blobFromImage(
        image_bgr,
        scalefactor=1.0,
        size=(input_width, input_height),
        mean=(104.0, 177.0, 123.0),
        swapRB=False,
        crop=False,
    )
    net.setInput(blob)
    detections = net.forward()
    if detections is None:
        return False
    for index in range(detections.shape[2]):
        confidence = float(detections[0, 0, index, 2])
        if confidence >= min_score:
            return True
    return False


def main() -> None:
    args = parse_args()
    bootstrap_vendor_paths(args.vendor_dir.resolve())
    cv2, np, requests, load_workbook, Image = import_dependencies()

    workbook_path = args.workbook.resolve()
    if not workbook_path.exists():
        raise SystemExit(f"Workbook not found: {workbook_path}")

    proto_path = args.proto_path.resolve() if args.proto_path else (args.model_dir.resolve() / "opencv_dnn_face_detector_deploy.prototxt")
    model_path = args.model_path.resolve() if args.model_path else (args.model_dir.resolve() / "opencv_dnn_face_detector_res10_300x300_ssd_iter_140000.caffemodel")
    proto_path = ensure_file(proto_path, args.proto_url)
    model_path = ensure_file(model_path, args.model_url)

    backup_path = workbook_path.with_name(workbook_path.name + args.backup_suffix)
    shutil.copy2(workbook_path, backup_path)
    if args.verbose:
        print(f"Backup created: {backup_path}", file=sys.stderr)

    workbook = load_workbook(workbook_path)
    worksheet = workbook[args.sheet] if args.sheet else workbook[workbook.sheetnames[0]]
    headers = find_header_map(worksheet)

    if args.url_column not in headers:
        raise SystemExit(f"Required URL column not found: {args.url_column}")

    headers, output_col = ensure_output_column(
        worksheet,
        headers,
        args.output_column,
        args.anchor_column,
        args.fallback_anchor_column,
    )
    url_col = headers[args.url_column]

    net = build_detector(cv2, proto_path, model_path)
    session = requests.Session()
    processed = 0

    for row in range(2, worksheet.max_row + 1):
        url_value = worksheet.cell(row=row, column=url_col).value
        url = str(url_value or "").strip()
        if not url:
            worksheet.cell(row=row, column=output_col, value="")
            continue

        if args.limit > 0 and processed >= args.limit:
            break

        try:
            image_bgr = fetch_bgr_image(url, args.timeout, session, Image, np)
            detected = detect_face_opencv_dnn_ssd(
                cv2,
                image_bgr,
                net,
                args.input_width,
                args.input_height,
                args.min_score,
            )
            worksheet.cell(row=row, column=output_col, value=bool_to_str(detected))
        except Exception as exc:  # noqa: BLE001
            worksheet.cell(row=row, column=output_col, value="")
            if args.verbose:
                print(f"[row {row}] failed for {url}: {exc}", file=sys.stderr)

        processed += 1
        if args.verbose and processed % 25 == 0:
            print(f"Processed {processed} images", file=sys.stderr)

    workbook.save(workbook_path)
    if args.verbose:
        print(f"Saved workbook with column '{args.output_column}': {workbook_path}", file=sys.stderr)


if __name__ == "__main__":
    main()
