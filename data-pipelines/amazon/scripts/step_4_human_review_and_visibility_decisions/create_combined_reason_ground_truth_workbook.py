#!/usr/bin/env python3
"""Create one combined yes/no ground-truth workbook for sparse rejection reasons."""

from __future__ import annotations

import csv
import re
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation


REPO_ROOT = Path(__file__).resolve().parents[4]
BROAD_DIR = REPO_ROOT / "outputs/cv_experiments/ground_truth_labeling_broad"
LABELED_CSV = BROAD_DIR / "labeled_2026_05_25/usable_labeled_ground_truth_normalized.csv"
BROAD_SOURCE_CSV = BROAD_DIR / "combined_amazon_nonamazon_llm_seeded_ground_truth_queue.csv"
YOLO_ROWS_CSV = REPO_ROOT / "outputs/cv_experiments/yolo_segmentation_crop_reasons_broad_2026_05_25/yolo_segmentation_crop_reason_rows.csv"
OUT_DIR = REPO_ROOT / "outputs/cv_experiments/combined_reason_ground_truth_2026_05_25"
OUTPUT_XLSX = OUT_DIR / "combined_rejection_reason_yes_no_review_queue.xlsx"
OUTPUT_CSV = OUT_DIR / "combined_rejection_reason_yes_no_review_queue.csv"
REPORT_MD = OUT_DIR / "combined_rejection_reason_yes_no_review_queue_report.md"

ANSWER_CHOICES = ["YES", "NO", "UNSURE"]
TARGET_POSITIVE_GOAL = 30
MAX_ROWS_PER_REASON = 28

REASONS = {
    "TOO_DARK": "Is the image too dark to judge clothing fit?",
    "BAD_ANGLE_TOP_DOWN": "Is this a top-down or strange angle that makes fit hard to judge?",
    "GARMENT_TOP_COVERED": "Is the top/waist/important upper edge of the garment covered or hidden?",
    "GARMENT_BOTTOM_CUT_OFF": "Is the bottom/hem/ankle edge of the garment cut off or missing?",
    "GARMENT_OBSCURED": "Is the garment materially blocked by another object, body part, phone, mirror, bag, hair, furniture, or overlay?",
    "GARMENT_CUT_OFF": "Is an important part of the worn garment outside the image boundary?",
    "DISTRACTING_OBJECTS": "Are there distracting objects or clutter that make the image less useful/shoppable?",
    "BACKGROUND_TOO_CLUTTERED": "Is the background too cluttered or visually noisy for a shopping context?",
    "BAD_ANGLE_SIDE_OR_TWISTED": "Is the body pose/side angle/twist/seated angle bad enough that fit is hard to judge?",
    "BLURRY_OR_MOTION_BLUR": "Is the image blurry enough that fit details are hard to judge?",
    "GRAINY_OR_NOISY": "Is the image grainy/noisy enough that fit details are hard to judge?",
    "TOO_BRIGHT_OR_WASHED_OUT": "Is the image overexposed or washed out enough that fit details are hard to judge?",
    "LOW_RESOLUTION": "Is the usable image too small, pixelated, or compressed after URL upgrade?",
    "NO_PERSON_VISIBLE": "Is there no visible person wearing clothing in the image?",
}

KEYWORDS = {
    "TOO_DARK": ["too dark", "dark", "low light", "shadow"],
    "BAD_ANGLE_TOP_DOWN": ["top down", "top-down", "from above", "feet", "strange angle", "weird angle"],
    "GARMENT_TOP_COVERED": ["top covered", "waistband", "waist", "covered by shirt", "covered by top", "top part"],
    "GARMENT_BOTTOM_CUT_OFF": ["bottom cut", "hem", "ankle", "feet cut", "bottom of", "lower part"],
    "GARMENT_OBSCURED": ["obscured", "blocked", "covered", "phone", "bag", "mirror frame", "hair"],
    "GARMENT_CUT_OFF": ["cut off", "cropped", "out of frame", "outside the frame", "image boundary"],
    "DISTRACTING_OBJECTS": ["distracting", "object", "clutter", "messy", "background"],
    "BACKGROUND_TOO_CLUTTERED": ["clutter", "messy", "background", "room", "objects"],
    "BAD_ANGLE_SIDE_OR_TWISTED": ["side view", "twisted", "seated", "sitting", "pose", "angle"],
    "BLURRY_OR_MOTION_BLUR": ["blur", "blurry", "motion"],
    "GRAINY_OR_NOISY": ["grainy", "noise", "noisy"],
    "TOO_BRIGHT_OR_WASHED_OUT": ["washed out", "overexposed", "too bright", "bright"],
    "LOW_RESOLUTION": ["low resolution", "pixelated", "compressed", "small"],
    "NO_PERSON_VISIBLE": ["no person", "no human", "no visible person"],
}

HEADERS = [
    "answer_yes_no",
    "rejection_reason",
    "question",
    "image_preview",
    "original_url_display",
    "product_page_url_display",
    "source_site_display",
    "why_included",
    "current_primary_reason_code",
    "current_secondary_reason_code",
    "current_labeler_notes",
    "current_final_human_decision",
    "llm_suggested_labels",
    "llm_summary",
    "luminance_mean",
    "bright_pixel_pct",
    "laplacian_variance",
    "seg_mask_area_pct",
    "seg_bbox_height_pct",
    "review_row_key",
]


def read_csv(path: Path) -> list[dict[str, str]]:
    with path.open(newline="", encoding="utf-8-sig") as handle:
        return list(csv.DictReader(handle))


def truthy(value: object) -> bool:
    return str(value or "").strip().lower() in {"true", "yes", "1", "x", "checked", "true"}


def norm(value: object) -> str:
    return str(value or "").strip().upper()


def to_float(value: object, default: float = 0.0) -> float:
    try:
        if value is None or str(value).strip() == "":
            return default
        return float(value)
    except ValueError:
        return default


def row_key(row: dict[str, str]) -> str:
    return row.get("review_row_key") or row.get("original_url_display") or ""


def indexed_rows() -> dict[str, dict[str, str]]:
    combined: dict[str, dict[str, str]] = {}
    for source_name, rows in [
        ("broad", read_csv(BROAD_SOURCE_CSV)),
        ("labeled", read_csv(LABELED_CSV)),
        ("yolo", read_csv(YOLO_ROWS_CSV)),
    ]:
        for row in rows:
            key = row_key(row)
            if not key or not row.get("original_url_display"):
                continue
            existing = combined.setdefault(key, {})
            existing.update({k: v for k, v in row.items() if v not in (None, "")})
            existing["_sources"] = (existing.get("_sources", "") + ";" + source_name).strip(";")
    return combined


def current_positive_counts(rows: list[dict[str, str]]) -> Counter:
    counts: Counter = Counter()
    reason_cols = [reason for reason in REASONS]
    for row in rows:
        labels = {norm(row.get("primary_reason_code")), norm(row.get("secondary_reason_code")), norm(row.get("third_reason_code"))}
        labels.update(reason for reason in reason_cols if truthy(row.get(reason)))
        for label in labels:
            if label in REASONS:
                counts[label] += 1
    return counts


def text_blob(row: dict[str, str]) -> str:
    parts = [
        row.get("primary_reason_code", ""),
        row.get("secondary_reason_code", ""),
        row.get("labeler_notes", ""),
        row.get("llm_suggested_labels", ""),
        row.get("llm_summary", ""),
        row.get("candidate_heuristic_tags", ""),
        row.get("user_comment", ""),
        row.get("product_title_raw", ""),
    ]
    return " ".join(str(part or "").lower() for part in parts)


def reason_score(row: dict[str, str], reason: str) -> tuple[int, list[str]]:
    score = 0
    why: list[str] = []
    labels = {
        norm(row.get("primary_reason_code")),
        norm(row.get("secondary_reason_code")),
        norm(row.get("third_reason_code")),
        *[norm(part) for part in re.split(r"[;,]", str(row.get("llm_suggested_labels") or ""))],
    }
    if reason in labels or truthy(row.get(reason)):
        score += 100
        why.append("existing_or_llm_label")

    blob = text_blob(row)
    hits = [keyword for keyword in KEYWORDS.get(reason, []) if keyword in blob]
    if hits:
        score += 20 + 4 * len(hits)
        why.append("text_hint:" + ",".join(hits[:3]))

    luminance = to_float(row.get("luminance_mean"), default=999.0)
    bright = to_float(row.get("bright_pixel_pct"), default=0.0)
    blur = to_float(row.get("laplacian_variance"), default=999999.0)
    height = to_float(row.get("height"), default=99999.0)
    width = to_float(row.get("width"), default=99999.0)

    if reason == "TOO_DARK" and luminance < 75:
        score += int(max(0, 75 - luminance))
        why.append("low_luminance")
    if reason == "TOO_BRIGHT_OR_WASHED_OUT" and bright > 0.10:
        score += int(bright * 100)
        why.append("high_bright_pixel_pct")
    if reason == "BLURRY_OR_MOTION_BLUR" and blur < 120:
        score += int(max(0, 120 - blur) / 3)
        why.append("low_laplacian_variance")
    if reason == "LOW_RESOLUTION" and min(width, height) < 700:
        score += 25
        why.append("small_loaded_dimension")
    if reason == "GARMENT_CUT_OFF" and to_float(row.get("seg_mask_area_pct"), default=1.0) < 0.25:
        score += 8
        why.append("small_seg_mask")

    return score, why


def build_queue() -> tuple[list[dict[str, object]], Counter]:
    all_rows_by_key = indexed_rows()
    labeled_rows = read_csv(LABELED_CSV)
    counts = current_positive_counts(labeled_rows)
    needed_reasons = [reason for reason in REASONS if counts.get(reason, 0) < TARGET_POSITIVE_GOAL]

    queue: list[dict[str, object]] = []
    seen: set[tuple[str, str]] = set()
    for reason in needed_reasons:
        scored = []
        for row in all_rows_by_key.values():
            if not row.get("original_url_display"):
                continue
            score, why = reason_score(row, reason)
            if score <= 0:
                continue
            scored.append((score, why, row))
        scored.sort(key=lambda item: (-item[0], row_key(item[2])))
        for score, why, row in scored[:MAX_ROWS_PER_REASON]:
            key = (reason, row_key(row))
            if key in seen:
                continue
            seen.add(key)
            queue.append(
                {
                    "answer_yes_no": "",
                    "rejection_reason": reason,
                    "question": REASONS[reason],
                    "image_preview": row.get("image_preview") or f'=IMAGE("{row.get("original_url_display", "")}")',
                    "original_url_display": row.get("original_url_display", ""),
                    "product_page_url_display": row.get("product_page_url_display", ""),
                    "source_site_display": row.get("source_site_display", ""),
                    "why_included": ";".join(why) or f"score={score}",
                    "current_primary_reason_code": row.get("primary_reason_code", ""),
                    "current_secondary_reason_code": row.get("secondary_reason_code", ""),
                    "current_labeler_notes": row.get("labeler_notes", ""),
                    "current_final_human_decision": row.get("final_human_decision_norm") or row.get("final_human_decision", ""),
                    "llm_suggested_labels": row.get("llm_suggested_labels", ""),
                    "llm_summary": row.get("llm_summary", ""),
                    "luminance_mean": row.get("luminance_mean", ""),
                    "bright_pixel_pct": row.get("bright_pixel_pct", ""),
                    "laplacian_variance": row.get("laplacian_variance", ""),
                    "seg_mask_area_pct": row.get("seg_mask_area_pct", ""),
                    "seg_bbox_height_pct": row.get("seg_bbox_height_pct", ""),
                    "review_row_key": row_key(row),
                    "_score": score,
                }
            )

    queue.sort(key=lambda item: (str(item["rejection_reason"]), -int(item["_score"]), str(item["review_row_key"])))
    for item in queue:
        item.pop("_score", None)
    return queue, counts


def style_header(sheet) -> None:
    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)
    for cell in sheet[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(wrap_text=True, vertical="top")


def write_workbook(rows: list[dict[str, object]]) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Combined Review"
    sheet.append(HEADERS)
    for row in rows:
        sheet.append([row.get(header, "") for header in HEADERS])
    style_header(sheet)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:T{sheet.max_row}"
    validation = DataValidation(type="list", formula1=f'"{",".join(ANSWER_CHOICES)}"', allow_blank=True)
    sheet.add_data_validation(validation)
    validation.add(f"A2:A{sheet.max_row}")
    widths = {
        "A": 18,
        "B": 28,
        "C": 72,
        "D": 26,
        "E": 42,
        "F": 42,
        "G": 22,
        "H": 36,
        "I": 26,
        "J": 26,
        "K": 38,
        "L": 20,
        "M": 28,
        "N": 64,
        "T": 34,
    }
    for col, width in widths.items():
        sheet.column_dimensions[col].width = width
    for col in ["O", "P", "Q", "R", "S"]:
        sheet.column_dimensions[col].width = 16
    for row in sheet.iter_rows(min_row=2):
        sheet.row_dimensions[row[0].row].height = 72
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    workbook.save(OUTPUT_XLSX)


def write_csv_output(rows: list[dict[str, object]]) -> None:
    with OUTPUT_CSV.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=HEADERS)
        writer.writeheader()
        writer.writerows([{header: row.get(header, "") for header in HEADERS} for row in rows])


def write_report(rows: list[dict[str, object]], counts: Counter) -> None:
    queue_counts = Counter(str(row["rejection_reason"]) for row in rows)
    lines = [
        "# Combined Rejection Reason Ground Truth Queue",
        "",
        f"- workbook: `{OUTPUT_XLSX}`",
        f"- rows to label: `{len(rows)}`",
        "- answer column: `answer_yes_no`",
        "- answer choices: `YES`, `NO`, `UNSURE`",
        "",
        "## Current Positive Counts Before This Queue",
        "",
        "| reason | current positives | target | rows in this queue |",
        "| --- | ---: | ---: | ---: |",
    ]
    for reason in REASONS:
        if counts.get(reason, 0) < TARGET_POSITIVE_GOAL:
            lines.append(f"| `{reason}` | {counts.get(reason, 0)} | {TARGET_POSITIVE_GOAL} | {queue_counts.get(reason, 0)} |")
    lines.extend(
        [
            "",
            "## Labeling Instruction",
            "",
            "For each row, answer the question in column C for the reason in column B. Answer `YES` if that reason applies to the image, even if another rejection reason is more important. Answer `NO` if it does not apply. Use `UNSURE` only when the image is ambiguous.",
            "",
        ]
    )
    REPORT_MD.write_text("\n".join(lines), encoding="utf-8")


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    rows, counts = build_queue()
    write_workbook(rows)
    write_csv_output(rows)
    write_report(rows, counts)
    print(OUTPUT_XLSX)
    print(f"rows: {len(rows)}")


if __name__ == "__main__":
    main()
