#!/usr/bin/env python3
"""Ingest a human-labeled Supabase image review workbook.

Splits rows by `production_decision` into approved upload rows and rejected rows
for sorter training, while preserving the labeled source workbook and CSV.
"""

from __future__ import annotations

import argparse
import csv
import re
import shutil
from pathlib import Path

from openpyxl import load_workbook


VALID_DECISIONS = {"APPROVE", "REJECT"}
APPROVE_NOTES = {"APPROVE", "APPROVED"}
REJECT_NOTES = {"REJECT", "REJECTED"}


def normalize_decision(value: object) -> str:
    return str(value or "").strip().upper()


def final_human_decision(row: dict[str, object], prefer_review_notes: bool) -> str:
    note = normalize_decision(row.get("review_notes"))
    if prefer_review_notes and note in APPROVE_NOTES:
        return "APPROVE"
    if prefer_review_notes and note in REJECT_NOTES:
        return "REJECT"
    return normalize_decision(row.get("production_decision"))


def safe_stem(path: Path) -> str:
    stem = path.stem
    stem = re.sub(r"_labelled$", "_labeled", stem, flags=re.IGNORECASE)
    return re.sub(r"[^A-Za-z0-9_.-]+", "_", stem).strip("_")


def part_id_from_stem(stem: str) -> str:
    match = re.search(r"part_(\d+)", stem)
    if match:
        return match.group(1)
    return "unknown"


def read_rows(workbook_path: Path) -> tuple[list[str], list[dict[str, object]]]:
    workbook = load_workbook(workbook_path, data_only=False, read_only=True)
    sheet = workbook.active
    rows_iter = sheet.iter_rows(values_only=True)
    headers = [str(value or "").strip() for value in next(rows_iter)]
    if "production_decision" not in headers:
        raise ValueError("Workbook is missing required column: production_decision")

    rows: list[dict[str, object]] = []
    for values in rows_iter:
        row = {header: values[index] if index < len(values) else "" for index, header in enumerate(headers)}
        if any(value not in (None, "") for value in values):
            rows.append(row)
    return headers, rows


def write_csv(path: Path, headers: list[str], rows: list[dict[str, object]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as file:
        writer = csv.DictWriter(file, fieldnames=headers, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("workbook", type=Path)
    parser.add_argument("--output-root", type=Path, default=Path("outputs/01_supabase_ready_human_approved"))
    parser.add_argument("--batch-name", required=True)
    parser.add_argument(
        "--prefer-review-notes",
        action="store_true",
        help="Treat review_notes values APPROVE/APPROVED/REJECT/REJECTED as the final human decision.",
    )
    args = parser.parse_args()

    workbook_path = args.workbook
    headers, rows = read_rows(workbook_path)
    stem = safe_stem(workbook_path)
    part_id = part_id_from_stem(stem)

    unknown = sorted({final_human_decision(row, args.prefer_review_notes) for row in rows} - VALID_DECISIONS)
    unknown = [value for value in unknown if value]
    if unknown:
        raise ValueError(f"Unexpected final human decision values: {unknown}")

    for row in rows:
        row["final_human_decision"] = final_human_decision(row, args.prefer_review_notes)

    headers_with_final = list(headers)
    if "final_human_decision" not in headers_with_final:
        headers_with_final.append("final_human_decision")

    approved = [row for row in rows if row["final_human_decision"] == "APPROVE"]
    rejected = [row for row in rows if row["final_human_decision"] == "REJECT"]

    batch_dir = args.output_root / args.batch_name
    approved_path = batch_dir / "approved_rows" / f"supabase_approved_rows_part_{part_id}.csv"
    rejected_path = batch_dir / "sorter_training_labels" / f"supabase_rejected_rows_part_{part_id}_for_sorter_training.csv"
    labeled_csv_path = batch_dir / "labeled_source" / f"{stem}.csv"
    labeled_workbook_path = batch_dir / "labeled_source" / workbook_path.name
    readme_path = batch_dir / "README.md"

    write_csv(approved_path, headers_with_final, approved)
    write_csv(rejected_path, headers_with_final, rejected)
    write_csv(labeled_csv_path, headers_with_final, rows)
    labeled_workbook_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(workbook_path, labeled_workbook_path)

    readme_path.write_text(
        "\n".join(
            [
                f"# {args.batch_name}",
                "",
                "Human-labeled Supabase image review ingest.",
                "",
                f"- source workbook: `{workbook_path}`",
                f"- total labeled rows: `{len(rows)}`",
                f"- approved rows: `{len(approved)}`",
                f"- rejected rows: `{len(rejected)}`",
                f"- final human decision source: `{'review_notes overrides production_decision' if args.prefer_review_notes else 'production_decision'}`",
                "",
                f"Approved upload rows: `{approved_path}`",
                "",
                f"Rejected rows for sorter training: `{rejected_path}`",
                "",
            ]
        ),
        encoding="utf-8",
    )

    print(f"batch_dir={batch_dir}")
    print(f"approved_rows={len(approved)}")
    print(f"rejected_rows={len(rejected)}")
    print(f"labeled_rows={len(rows)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
