#!/usr/bin/env python3
import argparse
from dataclasses import dataclass
from pathlib import Path
from typing import Optional


DEFAULT_GT_COLUMN = "has_face_GroundTruth (1=face,2=noFace)"
DEFAULT_DETECTOR_COLUMNS = [
    "has_face_yunet",
    "has_face_scrfd",
    "has_face_blazeface",
    "has_face_opencv_dnn_ssd",
    "has_face_retinaface",
]


@dataclass
class DetectorSummary:
    column_name: str
    rows_scored: int
    true_positive: int
    false_positive: int
    true_negative: int
    false_negative: int
    accuracy: float
    precision: float
    recall: float
    specificity: float
    f1: float


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Compare face detector columns against the workbook face ground truth column."
    )
    parser.add_argument("--workbook", type=Path, required=True, help="Path to the workbook containing detector columns and face GT.")
    parser.add_argument("--sheet", default="", help="Optional worksheet name. Defaults to the workbook's first sheet.")
    parser.add_argument("--ground-truth-column", default=DEFAULT_GT_COLUMN, help="Header name for the face ground truth column.")
    parser.add_argument("--detector-column", action="append", dest="detector_columns", help="Optional detector column(s) to analyze.")
    parser.add_argument("--output-report", type=Path, help="Optional Markdown report path.")
    return parser.parse_args()


def import_dependencies():
    try:
        from openpyxl import load_workbook  # type: ignore
    except ModuleNotFoundError as exc:
        raise SystemExit("Missing dependency: openpyxl") from exc
    return load_workbook


def normalize_cell(value) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def parse_ground_truth(value) -> Optional[int]:
    normalized = normalize_cell(value)
    if normalized == "1":
        return 1
    if normalized == "2":
        return 0
    return None


def parse_detector_value(value) -> Optional[int]:
    normalized = normalize_cell(value).lower()
    if normalized == "true":
        return 1
    if normalized == "false":
        return 0
    return None


def find_header_map(worksheet) -> dict[str, int]:
    headers: dict[str, int] = {}
    for column_index in range(1, worksheet.max_column + 1):
        value = worksheet.cell(row=1, column=column_index).value
        if isinstance(value, str) and value.strip():
            headers[value.strip()] = column_index
    return headers


def safe_divide(numerator: int, denominator: int) -> float:
    return numerator / denominator if denominator else 0.0


def summarize_detector(column_name: str, pairs: list[tuple[int, int]]) -> DetectorSummary:
    tp = sum(1 for gt, pred in pairs if gt == 1 and pred == 1)
    fp = sum(1 for gt, pred in pairs if gt == 0 and pred == 1)
    tn = sum(1 for gt, pred in pairs if gt == 0 and pred == 0)
    fn = sum(1 for gt, pred in pairs if gt == 1 and pred == 0)

    accuracy = safe_divide(tp + tn, len(pairs))
    precision = safe_divide(tp, tp + fp)
    recall = safe_divide(tp, tp + fn)
    specificity = safe_divide(tn, tn + fp)
    f1 = safe_divide(2 * precision * recall, precision + recall)

    return DetectorSummary(
        column_name=column_name,
        rows_scored=len(pairs),
        true_positive=tp,
        false_positive=fp,
        true_negative=tn,
        false_negative=fn,
        accuracy=accuracy,
        precision=precision,
        recall=recall,
        specificity=specificity,
        f1=f1,
    )


def format_float(value: float) -> str:
    return f"{value:.3f}"


def build_report(workbook_path: Path, sheet_name: str, ground_truth_column: str, gt_rows: int, summaries: list[DetectorSummary], missing_columns: list[str], blank_columns: list[str]) -> str:
    lines = [
        f"# Face Detector vs Ground Truth Report: {workbook_path.name}",
        "",
        "## Scope",
        "",
        f"- workbook: `{workbook_path}`",
        f"- sheet: `{sheet_name}`",
        f"- ground truth column: `{ground_truth_column}`",
        f"- GT-labeled rows analyzed: `{gt_rows}`",
        "",
        "## Detector Ranking",
        "",
    ]

    for summary in sorted(summaries, key=lambda item: (item.f1, item.accuracy), reverse=True):
        lines.append(
            "- {}: accuracy `{}`, precision `{}`, recall `{}`, specificity `{}`, F1 `{}`".format(
                f"`{summary.column_name}`",
                format_float(summary.accuracy),
                format_float(summary.precision),
                format_float(summary.recall),
                format_float(summary.specificity),
                format_float(summary.f1),
            )
        )

    lines.extend(
        [
            "",
            "## Confusion Summary",
            "",
            "| detector | rows | accuracy | precision | recall | specificity | F1 | TP | FP | TN | FN |",
            "| --- | ---: | ---: | ---: | ---: | ---: | ---: | ---: | ---: | ---: | ---: |",
        ]
    )

    for summary in sorted(summaries, key=lambda item: (item.f1, item.accuracy), reverse=True):
        lines.append(
            "| {detector} | {rows} | {accuracy} | {precision} | {recall} | {specificity} | {f1} | {tp} | {fp} | {tn} | {fn} |".format(
                detector=f"`{summary.column_name}`",
                rows=summary.rows_scored,
                accuracy=format_float(summary.accuracy),
                precision=format_float(summary.precision),
                recall=format_float(summary.recall),
                specificity=format_float(summary.specificity),
                f1=format_float(summary.f1),
                tp=summary.true_positive,
                fp=summary.false_positive,
                tn=summary.true_negative,
                fn=summary.false_negative,
            )
        )

    if missing_columns:
        lines.extend(
            [
                "",
                "## Missing Columns",
                "",
                "- detector columns not present in the workbook: {}".format(", ".join(f"`{name}`" for name in missing_columns)),
            ]
        )

    if blank_columns:
        lines.extend(
            [
                "",
                "## Blank Columns",
                "",
                "- present but blank for all GT-labeled rows: {}".format(", ".join(f"`{name}`" for name in blank_columns)),
            ]
        )

    return "\n".join(lines) + "\n"


def main() -> None:
    args = parse_args()
    load_workbook = import_dependencies()

    workbook_path = args.workbook.resolve()
    if not workbook_path.exists():
        raise SystemExit(f"Workbook not found: {workbook_path}")

    workbook = load_workbook(workbook_path, read_only=False, data_only=True)
    worksheet = workbook[args.sheet] if args.sheet else workbook[workbook.sheetnames[0]]
    headers = find_header_map(worksheet)

    if args.ground_truth_column not in headers:
        raise SystemExit(f"Ground truth column not found: {args.ground_truth_column}")

    gt_col = headers[args.ground_truth_column]
    gt_rows: list[tuple[int, int]] = []
    for row_index in range(2, worksheet.max_row + 1):
        gt_value = parse_ground_truth(worksheet.cell(row=row_index, column=gt_col).value)
        if gt_value is not None:
            gt_rows.append((row_index, gt_value))

    detector_columns = args.detector_columns or DEFAULT_DETECTOR_COLUMNS
    summaries: list[DetectorSummary] = []
    missing_columns: list[str] = []
    blank_columns: list[str] = []

    for detector_column in detector_columns:
        if detector_column not in headers:
            missing_columns.append(detector_column)
            continue

        detector_col = headers[detector_column]
        pairs: list[tuple[int, int]] = []
        for row_index, gt_value in gt_rows:
            pred_value = parse_detector_value(worksheet.cell(row=row_index, column=detector_col).value)
            if pred_value is not None:
                pairs.append((gt_value, pred_value))

        if not pairs:
            blank_columns.append(detector_column)
            continue

        summaries.append(summarize_detector(detector_column, pairs))

    report_text = build_report(
        workbook_path=workbook_path,
        sheet_name=worksheet.title,
        ground_truth_column=args.ground_truth_column,
        gt_rows=len(gt_rows),
        summaries=summaries,
        missing_columns=missing_columns,
        blank_columns=blank_columns,
    )

    if args.output_report:
        output_path = args.output_report.resolve()
        output_path.parent.mkdir(parents=True, exist_ok=True)
        output_path.write_text(report_text, encoding="utf-8")
        print(f"Wrote report: {output_path}")
    else:
        print(report_text, end="")


if __name__ == "__main__":
    main()
