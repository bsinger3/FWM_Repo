#!/usr/bin/env python3
import argparse
from dataclasses import dataclass
from pathlib import Path
from statistics import mean
from typing import Iterable, Optional


DEFAULT_LABEL_COLUMN = "image_approved? (1=Approved,2=NotApproved)"
DEFAULT_APPROVED_VALUE = "1"
DEFAULT_REJECTED_VALUE = "2"
DEFAULT_METRIC_COLUMNS = [
    "has_face_scrfd",
    "has_face_retinaface",
    "has_face_blazeface",
    "person_count_yolo_detect",
    "main_person_height_pct_yolo_detect",
    "main_person_bbox_area_pct_yolo_detect",
    "person_count_yolo_pose",
    "main_person_height_pct_yolo_pose",
    "main_person_bbox_area_pct_yolo_pose",
    "body_coverage_score_yolo_pose",
    "body_coverage_score_mediapipe_pose",
    "head_visible_mediapipe_pose",
    "shoulders_visible_mediapipe_pose",
    "hips_visible_mediapipe_pose",
    "knees_visible_mediapipe_pose",
    "ankles_visible_mediapipe_pose",
    "feet_visible_mediapipe_pose",
]


@dataclass
class MetricSummary:
    column_name: str
    nonblank_rows: int
    approved_mean: float
    rejected_mean: float
    mean_gap: float
    preferred_direction: str
    best_rule: str
    best_accuracy: float
    best_balanced_accuracy: float


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Compare workbook CV columns against a manual approval label and "
            "write a compact Markdown report."
        )
    )
    parser.add_argument(
        "--workbook",
        type=Path,
        required=True,
        help="Path to the workbook containing manual labels and CV columns.",
    )
    parser.add_argument(
        "--sheet",
        default="",
        help="Optional worksheet name. Defaults to the workbook's first sheet.",
    )
    parser.add_argument(
        "--label-column",
        default=DEFAULT_LABEL_COLUMN,
        help="Header name for the manual approval label.",
    )
    parser.add_argument(
        "--approved-value",
        default=DEFAULT_APPROVED_VALUE,
        help="Cell value representing an approved row.",
    )
    parser.add_argument(
        "--rejected-value",
        default=DEFAULT_REJECTED_VALUE,
        help="Cell value representing a rejected row.",
    )
    parser.add_argument(
        "--output-report",
        type=Path,
        help="Optional Markdown report path.",
    )
    parser.add_argument(
        "--metric-column",
        action="append",
        dest="metric_columns",
        help="Optional metric column(s) to analyze. Defaults to the known CV columns.",
    )
    return parser.parse_args()


def import_dependencies():
    try:
        from openpyxl import load_workbook  # type: ignore
    except ModuleNotFoundError as exc:
        raise SystemExit("Missing dependency: openpyxl") from exc
    return load_workbook


def normalize_cell(value) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def parse_metric_value(value) -> Optional[float]:
    normalized = normalize_cell(value).lower()
    if not normalized:
        return None
    if normalized == "true":
        return 1.0
    if normalized == "false":
        return 0.0
    try:
        return float(normalized)
    except ValueError:
        return None


def find_header_map(worksheet) -> dict[str, int]:
    headers: dict[str, int] = {}
    for column_index in range(1, worksheet.max_column + 1):
        value = worksheet.cell(row=1, column=column_index).value
        if isinstance(value, str) and value.strip():
            headers[value.strip()] = column_index
    return headers


def iter_labeled_rows(
    worksheet,
    label_column_index: int,
    approved_value: str,
    rejected_value: str,
) -> Iterable[tuple[int, int]]:
    for row_index in range(2, worksheet.max_row + 1):
        label = normalize_cell(worksheet.cell(row=row_index, column=label_column_index).value)
        if label == approved_value:
            yield row_index, 1
        elif label == rejected_value:
            yield row_index, 0


def summarize_metric(metric_name: str, pairs: list[tuple[int, float]]) -> MetricSummary:
    approved_values = [value for label, value in pairs if label == 1]
    rejected_values = [value for label, value in pairs if label == 0]

    approved_mean = mean(approved_values)
    rejected_mean = mean(rejected_values)
    mean_gap = approved_mean - rejected_mean
    preferred_direction = ">=" if mean_gap >= 0 else "<="

    best_accuracy = -1.0
    best_balanced_accuracy = -1.0
    best_rule = "n/a"

    unique_values = sorted(set(value for _label, value in pairs))
    for threshold in unique_values:
        for direction in (">=", "<="):
            true_positive = 0
            true_negative = 0
            false_positive = 0
            false_negative = 0

            for label, value in pairs:
                predicted_approved = value >= threshold if direction == ">=" else value <= threshold
                if predicted_approved and label == 1:
                    true_positive += 1
                elif predicted_approved and label == 0:
                    false_positive += 1
                elif not predicted_approved and label == 0:
                    true_negative += 1
                else:
                    false_negative += 1

            total = true_positive + true_negative + false_positive + false_negative
            accuracy = (true_positive + true_negative) / total if total else 0.0
            sensitivity = true_positive / (true_positive + false_negative) if (true_positive + false_negative) else 0.0
            specificity = true_negative / (true_negative + false_positive) if (true_negative + false_positive) else 0.0
            balanced_accuracy = (sensitivity + specificity) / 2.0

            if (
                balanced_accuracy > best_balanced_accuracy
                or (
                    balanced_accuracy == best_balanced_accuracy
                    and accuracy > best_accuracy
                )
            ):
                best_balanced_accuracy = balanced_accuracy
                best_accuracy = accuracy
                best_rule = f"`{metric_name}` {direction} `{format_threshold(threshold)}`"

    return MetricSummary(
        column_name=metric_name,
        nonblank_rows=len(pairs),
        approved_mean=approved_mean,
        rejected_mean=rejected_mean,
        mean_gap=mean_gap,
        preferred_direction=preferred_direction,
        best_rule=best_rule,
        best_accuracy=best_accuracy,
        best_balanced_accuracy=best_balanced_accuracy,
    )


def format_threshold(value: float) -> str:
    if float(value).is_integer():
        return str(int(value))
    return f"{value:.3f}".rstrip("0").rstrip(".")


def format_float(value: float) -> str:
    return f"{value:.3f}"


def build_report(
    workbook_path: Path,
    sheet_name: str,
    label_column: str,
    labeled_rows: int,
    approved_rows: int,
    rejected_rows: int,
    missing_metrics: list[str],
    blank_metrics: list[str],
    summaries: list[MetricSummary],
) -> str:
    lines = [
        f"# CV vs Manual Approval Report: {workbook_path.name}",
        "",
        "## Scope",
        "",
        f"- workbook: `{workbook_path}`",
        f"- sheet: `{sheet_name}`",
        f"- manual label column: `{label_column}`",
        f"- labeled rows analyzed: `{labeled_rows}`",
        f"- approved rows: `{approved_rows}`",
        f"- rejected rows: `{rejected_rows}`",
        "",
        "## Strongest Single-Column Signals",
        "",
    ]

    top_summaries = sorted(summaries, key=lambda item: item.best_balanced_accuracy, reverse=True)[:5]
    for summary in top_summaries:
        lines.append(
            "- {}: balanced accuracy `{}`, raw accuracy `{}`, best rule {}".format(
                f"`{summary.column_name}`",
                format_float(summary.best_balanced_accuracy),
                format_float(summary.best_accuracy),
                summary.best_rule,
            )
        )

    lines.extend(
        [
            "",
            "## Metric Summary",
            "",
            "| metric | rows | approved mean | rejected mean | gap | best balanced acc. | best rule |",
            "| --- | ---: | ---: | ---: | ---: | ---: | --- |",
        ]
    )

    for summary in sorted(summaries, key=lambda item: abs(item.mean_gap), reverse=True):
        lines.append(
            "| {metric} | {rows} | {approved_mean} | {rejected_mean} | {gap} | {balanced_accuracy} | {rule} |".format(
                metric=f"`{summary.column_name}`",
                rows=summary.nonblank_rows,
                approved_mean=format_float(summary.approved_mean),
                rejected_mean=format_float(summary.rejected_mean),
                gap=format_float(summary.mean_gap),
                balanced_accuracy=format_float(summary.best_balanced_accuracy),
                rule=summary.best_rule,
            )
        )

    if missing_metrics:
        lines.extend(
            [
                "",
                "## Missing Metrics",
                "",
                "- columns not present in the workbook: {}".format(
                    ", ".join(f"`{name}`" for name in missing_metrics)
                ),
            ]
        )

    if blank_metrics:
        lines.extend(
            [
                "",
                "## Blank Metrics",
                "",
                "- present but blank for all labeled rows: {}".format(
                    ", ".join(f"`{name}`" for name in blank_metrics)
                ),
            ]
        )

    return "\n".join(lines) + "\n"


def main() -> None:
    args = parse_args()
    load_workbook = import_dependencies()

    workbook_path = args.workbook.resolve()
    if not workbook_path.exists():
        raise SystemExit(f"Workbook not found: {workbook_path}")

    workbook = load_workbook(workbook_path, read_only=False, data_only=True)
    worksheet = workbook[args.sheet] if args.sheet else workbook[workbook.sheetnames[0]]
    headers = find_header_map(worksheet)

    if args.label_column not in headers:
        raise SystemExit(f"Label column not found: {args.label_column}")

    metric_columns = args.metric_columns or DEFAULT_METRIC_COLUMNS
    label_column_index = headers[args.label_column]
    labeled_rows = list(
        iter_labeled_rows(
            worksheet,
            label_column_index,
            approved_value=args.approved_value,
            rejected_value=args.rejected_value,
        )
    )
    approved_rows = sum(label for _row_index, label in labeled_rows)
    rejected_rows = len(labeled_rows) - approved_rows

    missing_metrics: list[str] = []
    blank_metrics: list[str] = []
    summaries: list[MetricSummary] = []

    for metric_name in metric_columns:
        if metric_name not in headers:
            missing_metrics.append(metric_name)
            continue

        metric_column_index = headers[metric_name]
        pairs: list[tuple[int, float]] = []
        for row_index, label in labeled_rows:
            metric_value = parse_metric_value(worksheet.cell(row=row_index, column=metric_column_index).value)
            if metric_value is not None:
                pairs.append((label, metric_value))

        if not pairs:
            blank_metrics.append(metric_name)
            continue

        summaries.append(summarize_metric(metric_name, pairs))

    report_text = build_report(
        workbook_path=workbook_path,
        sheet_name=worksheet.title,
        label_column=args.label_column,
        labeled_rows=len(labeled_rows),
        approved_rows=approved_rows,
        rejected_rows=rejected_rows,
        missing_metrics=missing_metrics,
        blank_metrics=blank_metrics,
        summaries=summaries,
    )

    if args.output_report:
        output_path = args.output_report.resolve()
        output_path.parent.mkdir(parents=True, exist_ok=True)
        output_path.write_text(report_text, encoding="utf-8")
        print(f"Wrote report: {output_path}")
    else:
        print(report_text, end="")


if __name__ == "__main__":
    main()
