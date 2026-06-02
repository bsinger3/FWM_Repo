#!/usr/bin/env python3
"""Create a Google-Sheets-friendly workbook for YOLO segmentation error analysis."""

from __future__ import annotations

import csv
from pathlib import Path
from typing import Callable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


REPO_ROOT = Path(__file__).resolve().parents[4]
EXP_DIR = REPO_ROOT / "outputs/cv_experiments/yolo_segmentation_crop_reasons_broad_2026_05_25"
SOURCE_CSV = EXP_DIR / "yolo_segmentation_crop_reason_rows.csv"
OUTPUT_XLSX = EXP_DIR / "yolo_error_analysis_workbook.xlsx"
OUTPUT_SUMMARY = EXP_DIR / "yolo_error_analysis_summary.csv"

NEEDS_CROP_RULE = "seg_mask_area_pct < 0.20"
NOT_WORN_RULE = "seg_person_count == 0"

REVIEW_HEADERS = [
    "bucket",
    "reviewer_takeaway",
    "reviewer_notes",
    "image_preview",
    "original_url_display",
    "product_page_url_display",
    "source_site_display",
    "final_human_decision",
    "primary_reason_code",
    "secondary_reason_code",
    "labeler_notes",
    "truth_NEEDS_CROP",
    "pred_NEEDS_CROP_mask_area_lt_0_20",
    "truth_NOT_WORN_BY_PERSON",
    "pred_NOT_WORN_no_person_segmented",
    "seg_person_count",
    "seg_person_confidence",
    "seg_mask_area_pct",
    "seg_bbox_area_pct",
    "seg_bbox_height_pct",
    "seg_mask_touches_top",
    "seg_mask_touches_bottom",
    "seg_mask_top_gap_pct",
    "seg_mask_bottom_gap_pct",
    "llm_suggested_labels",
    "llm_summary",
    "user_comment",
    "product_title_raw",
    "review_row_key",
]


def read_rows() -> list[dict[str, str]]:
    with SOURCE_CSV.open(newline="", encoding="utf-8-sig") as handle:
        return list(csv.DictReader(handle))


def truthy(value: object) -> bool:
    return str(value or "").strip().lower() in {"true", "yes", "1", "x", "checked"}


def to_float(value: object, default: float = 0.0) -> float:
    try:
        if value is None or str(value).strip() == "":
            return default
        return float(value)
    except ValueError:
        return default


def pred_needs_crop(row: dict[str, str]) -> bool:
    return to_float(row.get("seg_mask_area_pct"), default=1.0) < 0.20


def pred_not_worn(row: dict[str, str]) -> bool:
    return to_float(row.get("seg_person_count"), default=999.0) == 0


def style_header(sheet) -> None:
    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)
    for cell in sheet[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(wrap_text=True, vertical="top")


def review_value(row: dict[str, str], header: str, bucket: str) -> object:
    if header == "bucket":
        return bucket
    if header in {"reviewer_takeaway", "reviewer_notes"}:
        return ""
    if header == "pred_NEEDS_CROP_mask_area_lt_0_20":
        return pred_needs_crop(row)
    if header == "pred_NOT_WORN_no_person_segmented":
        return pred_not_worn(row)
    return row.get(header, "")


def add_review_sheet(workbook: Workbook, title: str, rows: list[dict[str, str]], bucket: str) -> None:
    sheet = workbook.create_sheet(title)
    sheet.append(REVIEW_HEADERS)
    for row in rows:
        sheet.append([review_value(row, header, bucket) for header in REVIEW_HEADERS])
    style_header(sheet)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

    widths = {
        "A": 28,
        "B": 24,
        "C": 42,
        "D": 26,
        "E": 42,
        "F": 42,
        "G": 22,
        "H": 18,
        "I": 24,
        "J": 24,
        "K": 38,
        "Y": 28,
        "Z": 58,
        "AA": 46,
        "AB": 46,
        "AC": 28,
    }
    for column_letter, width in widths.items():
        sheet.column_dimensions[column_letter].width = width
    for idx in range(12, 25):
        sheet.column_dimensions[sheet.cell(1, idx).column_letter].width = 16

    for row in sheet.iter_rows(min_row=2):
        sheet.row_dimensions[row[0].row].height = 68
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")


def add_summary(workbook: Workbook, bucketed: dict[str, list[dict[str, str]]], rows: list[dict[str, str]]) -> None:
    sheet = workbook.active
    sheet.title = "Summary"
    needs_positive = [row for row in rows if truthy(row.get("truth_NEEDS_CROP"))]
    not_worn_positive = [row for row in rows if truthy(row.get("truth_NOT_WORN_BY_PERSON"))]
    segmented = [row for row in rows if to_float(row.get("seg_person_count"), default=0.0) > 0]

    summary_rows = [
        ("Source rows", len(rows)),
        ("Rows with person segmentation", len(segmented)),
        ("NEEDS_CROP rule", NEEDS_CROP_RULE),
        ("NEEDS_CROP positives", len(needs_positive)),
        ("NEEDS_CROP rule positives", sum(1 for row in rows if pred_needs_crop(row))),
        ("NOT_WORN_BY_PERSON rule", NOT_WORN_RULE),
        ("NOT_WORN_BY_PERSON positives", len(not_worn_positive)),
        ("NOT_WORN_BY_PERSON rule positives", sum(1 for row in rows if pred_not_worn(row))),
    ]
    sheet.append(["metric", "value"])
    for item in summary_rows:
        sheet.append(list(item))
    sheet.append([])
    sheet.append(["review_tab", "rows", "purpose"])
    purposes = {
        "NC True Positives": "True NEEDS_CROP rows caught by mask_area_lt_0.20.",
        "NC False Negatives": "True NEEDS_CROP rows missed by mask_area_lt_0.20.",
        "NC FP Approved": "Approved rows falsely flagged by mask_area_lt_0.20.",
        "NC FP Rejected Other": "Rejected non-NEEDS_CROP rows flagged by mask_area_lt_0.20.",
        "NW True Positives": "True NOT_WORN_BY_PERSON rows caught by no_person_segmented.",
        "NW False Negatives": "True NOT_WORN_BY_PERSON rows where YOLO still found a person.",
        "NW False Positives": "Non-NOT_WORN_BY_PERSON rows where YOLO found no person.",
    }
    for tab, tab_rows in bucketed.items():
        sheet.append([tab, len(tab_rows), purposes.get(tab, "")])
    style_header(sheet)
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    sheet.column_dimensions["A"].width = 30
    sheet.column_dimensions["B"].width = 18
    sheet.column_dimensions["C"].width = 90


def write_summary_csv(bucketed: dict[str, list[dict[str, str]]]) -> None:
    with OUTPUT_SUMMARY.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=["review_tab", "rows"])
        writer.writeheader()
        for tab, rows in bucketed.items():
            writer.writerow({"review_tab": tab, "rows": len(rows)})


def main() -> None:
    rows = read_rows()

    def needs_truth(row: dict[str, str]) -> bool:
        return truthy(row.get("truth_NEEDS_CROP"))

    def not_worn_truth(row: dict[str, str]) -> bool:
        return truthy(row.get("truth_NOT_WORN_BY_PERSON"))

    def approved(row: dict[str, str]) -> bool:
        return str(row.get("final_human_decision_norm") or row.get("final_human_decision") or "").upper() == "APPROVED"

    def rejected(row: dict[str, str]) -> bool:
        return str(row.get("final_human_decision_norm") or row.get("final_human_decision") or "").upper() == "REJECTED"

    def select(predicate: Callable[[dict[str, str]], bool]) -> list[dict[str, str]]:
        return sorted(
            [row for row in rows if predicate(row)],
            key=lambda row: (
                to_float(row.get("seg_mask_area_pct"), default=999.0),
                row.get("review_row_key", ""),
            ),
        )

    bucketed = {
        "NC True Positives": select(lambda row: needs_truth(row) and pred_needs_crop(row)),
        "NC False Negatives": select(lambda row: needs_truth(row) and not pred_needs_crop(row)),
        "NC FP Approved": select(lambda row: not needs_truth(row) and pred_needs_crop(row) and approved(row)),
        "NC FP Rejected Other": select(lambda row: not needs_truth(row) and pred_needs_crop(row) and rejected(row)),
        "NW True Positives": select(lambda row: not_worn_truth(row) and pred_not_worn(row)),
        "NW False Negatives": select(lambda row: not_worn_truth(row) and not pred_not_worn(row)),
        "NW False Positives": select(lambda row: not not_worn_truth(row) and pred_not_worn(row)),
    }

    workbook = Workbook()
    add_summary(workbook, bucketed, rows)
    for tab, tab_rows in bucketed.items():
        add_review_sheet(workbook, tab, tab_rows, tab)

    workbook.save(OUTPUT_XLSX)
    write_summary_csv(bucketed)
    print(OUTPUT_XLSX)


if __name__ == "__main__":
    main()
