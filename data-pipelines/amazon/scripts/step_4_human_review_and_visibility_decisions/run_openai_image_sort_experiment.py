#!/usr/bin/env python3
import argparse
import csv
import json
import os
import sys
import time
from collections import Counter, defaultdict
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Sequence, Tuple
from urllib.parse import parse_qs, unquote, urlparse
from urllib.request import Request, urlopen


SCRIPT_DIR = Path(__file__).resolve().parent
PROJECTS_ROOT = SCRIPT_DIR.parents[4]
DEFAULT_ENV_PATH = PROJECTS_ROOT / "FWM_Repo" / ".env"
DEFAULT_DATA_ROOT = (
    PROJECTS_ROOT
    / "FWM_Data"
    / "amazon"
    / "data"
    / "step_4_human_review_and_visibility_decisions"
)
DEFAULT_INPUT_DIR = DEFAULT_DATA_ROOT / "manual_chunks"
DEFAULT_OUTPUT_DIR = DEFAULT_DATA_ROOT / "openai_experiments"
PROMPT_VERSION = "2026-04-23_v1"
DEFAULT_MODEL = "gpt-4.1-mini"
DEFAULT_IMAGE_DETAIL = "low"
DEFAULT_MAX_OUTPUT_TOKENS = 200
DEFAULT_TIMEOUT_SECONDS = 60
DEFAULT_MAX_WORKERS = 4

RESULT_COLUMNS = [
    "review_row_key",
    "source_file",
    "source_row_number",
    "original_url_display",
    "product_page_url_display",
    "monetized_product_url_display",
    "clothing_type_id",
    "product_title_hint",
    "openai_decision",
    "openai_reason_code",
    "openai_reviewed_garment_present",
    "openai_full_reviewed_garment_visible",
    "openai_reviewed_garment_obscured_or_cut_off",
    "openai_reason_summary",
    "openai_confidence",
    "openai_model",
    "openai_image_detail",
    "openai_prompt_version",
    "openai_run_id",
    "openai_request_status",
    "openai_input_tokens",
    "openai_output_tokens",
    "openai_total_tokens",
    "openai_estimated_cost_usd",
    "openai_error_message",
]

ALLOWED_DECISIONS = {"APPROVED", "REJECTED", "STILL_NEEDS_HUMAN_REVIEW"}
ALLOWED_REASON_CODES = {
    "CLEAR_FIT_PHOTO",
    "IMAGE_NOT_RENDERABLE",
    "NO_PERSON",
    "MULTIPLE_PEOPLE",
    "TOO_FAR_AWAY",
    "GARMENT_NOT_VISIBLE",
    "GARMENT_CATEGORY_MISMATCH",
    "GARMENT_OBSCURED",
    "TOO_CROPPED",
    "GARMENT_CUT_OFF",
    "TOO_DARK",
    "BAD_ANGLE",
    "TOO_CLUTTERED",
    "CATALOG_OR_PRODUCT_IMAGE",
    "BACKGROUND_REMOVED_OR_STYLIZED",
    "BORDERLINE_VISIBILITY",
    "BORDERLINE_COMPOSITION",
    "INSUFFICIENT_VISUAL_EVIDENCE",
}
TRI_STATE = {"YES", "NO", "UNCERTAIN"}

MODEL_PRICING = {
    "gpt-5.4": {"input_per_million": 2.50, "output_per_million": 15.00},
    "gpt-5.4-mini": {"input_per_million": 0.75, "output_per_million": 4.50},
    "gpt-5.4-nano": {"input_per_million": 0.20, "output_per_million": 1.25},
    "gpt-4.1-mini": {"input_per_million": 0.40, "output_per_million": 1.60},
}

SYSTEM_PROMPT = """You are reviewing product-review photos for a clothing fit website.

Your task is to decide whether each image is useful for showing how the reviewed garment fits on a real person.

You will be given the reviewed garment category as metadata. Use that category as the target garment.

Do not guess the target garment from the image alone when category metadata is provided.

Return exactly one decision:
- APPROVED
- REJECTED
- STILL_NEEDS_HUMAN_REVIEW

You must be conservative. If the image is borderline or uncertain, use STILL_NEEDS_HUMAN_REVIEW.

Approve only when the image is clearly useful for fit evaluation.

To approve, the reviewed garment should usually be fully visible for its category and not materially obscured or cut off.

You must explicitly judge:
- whether the reviewed garment appears to be the garment shown in the image
- whether the reviewed garment is fully visible enough for that category
- whether the reviewed garment is obscured or cut off
- whether the image successfully renders and is visually inspectable

Check the image in this order:
1. Does the image render successfully and show actual photo content?
2. Is there a person visible?
3. Is there one clear primary subject?
4. Does the visible garment match the reviewed garment category?
5. Is the reviewed garment fully visible enough for its category?
6. Is the reviewed garment obscured, covered, cropped, or cut off?
7. Is the image useful for understanding fit, not just color or fabric?
8. Is the image a real shopper/review photo rather than a catalog or product-only image?

Reject when the image is clearly not useful, such as:
- image does not render or cannot be visually inspected
- no visible person
- too many people
- person too far away
- reviewed garment not visible enough
- reviewed garment category does not match the garment clearly shown
- reviewed garment is materially obscured
- reviewed garment is cut off
- crop too tight
- too dark
- bad angle
- cluttered
- catalog/product-only image
- obvious non-review image

We care about whether a shopper can understand garment fit from the image.

Return structured JSON only."""

USER_PROMPT_TEMPLATE = """Review this clothing review image.

The reviewed garment category for this row is: {clothing_type_id}

Optional product title hint from URL slug: {product_title_hint}

Decide whether it should be:
- APPROVED
- REJECTED
- STILL_NEEDS_HUMAN_REVIEW

Also provide:
- a short reason_code from the allowed list
- a short explanation
- a confidence score from 0.00 to 1.00
- whether the reviewed garment is clearly present in the image
- whether the full reviewed garment is visible enough for its category
- whether the reviewed garment is materially obscured or cut off

Evaluate the image using this exact checklist:
1. Does the image render successfully?
2. Is there a visible person?
3. Is there one clear primary subject?
4. Using the provided clothing_type_id as the target garment category, does the visible garment match that category?
5. Is the reviewed garment fully visible enough for that category?
6. Is the reviewed garment obscured by pose, hands, hair, outerwear, furniture, mirror framing, or cropping?
7. Is the reviewed garment cut off at a critical boundary for that category?
8. Is the image useful for judging fit, rather than only showing fabric, color, or a small detail?
9. Is the image clearly a shopper/review image rather than a catalog or product-only image?

Allowed reason codes:
CLEAR_FIT_PHOTO
IMAGE_NOT_RENDERABLE
NO_PERSON
MULTIPLE_PEOPLE
TOO_FAR_AWAY
GARMENT_NOT_VISIBLE
GARMENT_CATEGORY_MISMATCH
GARMENT_OBSCURED
TOO_CROPPED
GARMENT_CUT_OFF
TOO_DARK
BAD_ANGLE
TOO_CLUTTERED
CATALOG_OR_PRODUCT_IMAGE
BACKGROUND_REMOVED_OR_STYLIZED
BORDERLINE_VISIBILITY
BORDERLINE_COMPOSITION
INSUFFICIENT_VISUAL_EVIDENCE

Use these decision rules:
- If the image does not render, use REJECTED with IMAGE_NOT_RENDERABLE.
- If the image clearly fails, use REJECTED with the best matching specific reason code.
- If the image clearly passes, use APPROVED with CLEAR_FIT_PHOTO.
- If the image is borderline, mixed, or lacks enough visual evidence for a confident approve/reject call, use STILL_NEEDS_HUMAN_REVIEW with BORDERLINE_VISIBILITY, BORDERLINE_COMPOSITION, or INSUFFICIENT_VISUAL_EVIDENCE.

Optional metadata:
- clothing_type_id: {clothing_type_id}
- product_title_hint: {product_title_hint}
- product_page_url_display: {product_page_url_display}
- user_comment: {user_comment}
- height_raw: {height_raw}
- weight_raw: {weight_raw}
"""

JSON_SCHEMA = {
    "name": "image_sort_result",
    "schema": {
        "type": "object",
        "additionalProperties": False,
        "properties": {
            "decision": {
                "type": "string",
                "enum": sorted(ALLOWED_DECISIONS),
            },
            "reason_code": {
                "type": "string",
                "enum": sorted(ALLOWED_REASON_CODES),
            },
            "reviewed_garment_present": {
                "type": "string",
                "enum": sorted(TRI_STATE),
            },
            "full_reviewed_garment_visible": {
                "type": "string",
                "enum": sorted(TRI_STATE),
            },
            "reviewed_garment_obscured_or_cut_off": {
                "type": "string",
                "enum": sorted(TRI_STATE),
            },
            "reason_summary": {"type": "string"},
            "confidence": {"type": "number"},
        },
        "required": [
            "decision",
            "reason_code",
            "reviewed_garment_present",
            "full_reviewed_garment_visible",
            "reviewed_garment_obscured_or_cut_off",
            "reason_summary",
            "confidence",
        ],
    },
    "strict": True,
}


@dataclass
class RunPaths:
    run_dir: Path
    row_results_csv: Path
    workbook_path: Path
    report_path: Path


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Run the OpenAI image-sorting experiment for Amazon Step 4 chunks, "
            "then write a consolidated Excel workbook and markdown run report."
        )
    )
    parser.add_argument(
        "--input-files",
        nargs="+",
        type=Path,
        default=[
            DEFAULT_INPUT_DIR / "images_to_approve_part_001.csv",
            DEFAULT_INPUT_DIR / "images_to_approve_part_002.csv",
        ],
        help="Manual chunk CSV files to process.",
    )
    parser.add_argument("--data-root", type=Path, default=DEFAULT_DATA_ROOT, help="Step 4 data root.")
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT_DIR, help="Directory for experiment outputs.")
    parser.add_argument("--run-name", default="", help="Optional run name. Default uses the current UTC timestamp.")
    parser.add_argument("--model", default=DEFAULT_MODEL, help="OpenAI model name.")
    parser.add_argument("--env-file", type=Path, default=DEFAULT_ENV_PATH, help="Optional .env file to load before reading OPENAI_API_KEY.")
    parser.add_argument(
        "--image-detail",
        default=DEFAULT_IMAGE_DETAIL,
        choices=["low", "high", "auto"],
        help="Image detail setting for OpenAI image input.",
    )
    parser.add_argument("--max-output-tokens", type=int, default=DEFAULT_MAX_OUTPUT_TOKENS)
    parser.add_argument("--timeout-seconds", type=int, default=DEFAULT_TIMEOUT_SECONDS)
    parser.add_argument("--max-workers", type=int, default=DEFAULT_MAX_WORKERS)
    parser.add_argument("--limit", type=int, default=0, help="Optional row limit for smoke tests.")
    parser.add_argument("--resume", action="store_true", help="Skip rows already present in the run CSV.")
    parser.add_argument("--dry-run", action="store_true", help="Prepare outputs without calling OpenAI.")
    return parser.parse_args()


def load_env_file(path: Path) -> None:
    if not path.exists():
        return
    for raw_line in path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        key = key.strip()
        value = value.strip().strip('"').strip("'")
        if key and key not in os.environ:
            os.environ[key] = value


def utc_now_iso() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat()


def read_csv_rows(path: Path) -> Tuple[List[str], List[Dict[str, str]]]:
    with path.open("r", newline="", encoding="utf-8-sig") as infile:
        reader = csv.DictReader(infile)
        if not reader.fieldnames:
            raise SystemExit("No CSV header found in {}".format(path))
        return list(reader.fieldnames), list(reader)


def ensure_run_paths(output_root: Path, run_name: str) -> RunPaths:
    run_dir = output_root / run_name
    run_dir.mkdir(parents=True, exist_ok=True)
    return RunPaths(
        run_dir=run_dir,
        row_results_csv=run_dir / "openai_row_results.csv",
        workbook_path=run_dir / "openai_chunks_001_002_experiment_results.xlsx",
        report_path=run_dir / "openai_chunks_001_002_experiment_run_report.md",
    )


def build_review_row_key(source_file: Path, source_row_number: int) -> str:
    return "{}::{}".format(source_file.name, source_row_number)


def decode_redirect_url(url: str) -> str:
    parsed = urlparse(url)
    query = parse_qs(parsed.query)
    for key in ("u", "url", "target"):
        values = query.get(key)
        if values:
            return unquote(values[0])
    return url


def choose_source_url(row: Dict[str, str]) -> str:
    product_url = str(row.get("product_page_url_display", "") or "").strip()
    if product_url:
        return product_url
    monetized = str(row.get("monetized_product_url_display", "") or "").strip()
    if monetized:
        return decode_redirect_url(monetized)
    return ""


def extract_product_title_hint(row: Dict[str, str]) -> str:
    source_url = choose_source_url(row)
    if not source_url:
        return ""
    parsed = urlparse(source_url)
    path_parts = [part for part in parsed.path.split("/") if part]
    if "dp" in path_parts:
        dp_index = path_parts.index("dp")
        if dp_index > 0:
            slug = path_parts[dp_index - 1]
            return normalize_slug_text(slug)
    if path_parts:
        return normalize_slug_text(path_parts[-1])
    return ""


def normalize_slug_text(value: str) -> str:
    text = unquote(value).replace("-", " ").replace("_", " ").replace("+", " ").strip()
    return " ".join(part for part in text.split() if part)


def append_csv_row(path: Path, fieldnames: Sequence[str], row: Dict[str, object]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    write_header = not path.exists()
    with path.open("a", newline="", encoding="utf-8") as outfile:
        writer = csv.DictWriter(outfile, fieldnames=list(fieldnames))
        if write_header:
            writer.writeheader()
        writer.writerow({field: row.get(field, "") for field in fieldnames})


def load_existing_results(path: Path) -> Dict[str, Dict[str, str]]:
    if not path.exists():
        return {}
    with path.open("r", newline="", encoding="utf-8-sig") as infile:
        reader = csv.DictReader(infile)
        return {str(row.get("review_row_key", "")): row for row in reader}


def format_user_prompt(row: Dict[str, str], product_title_hint: str) -> str:
    return USER_PROMPT_TEMPLATE.format(
        clothing_type_id=(row.get("clothing_type_id", "") or "").strip() or "unknown",
        product_title_hint=product_title_hint or "(none)",
        product_page_url_display=(row.get("product_page_url_display", "") or "").strip() or "(blank)",
        user_comment=(row.get("user_comment", "") or "").strip() or "(blank)",
        height_raw=(row.get("height_raw", "") or "").strip() or "(blank)",
        weight_raw=(row.get("weight_raw", "") or "").strip() or "(blank)",
    )


def build_openai_request_payload(
    model: str,
    image_detail: str,
    row: Dict[str, str],
    product_title_hint: str,
    max_output_tokens: int,
) -> Dict[str, object]:
    image_url = str(row.get("original_url_display", "") or "").strip()
    if not image_url:
        raise ValueError("Missing original_url_display")
    return {
        "model": model,
        "input": [
            {
                "role": "system",
                "content": [{"type": "input_text", "text": SYSTEM_PROMPT}],
            },
            {
                "role": "user",
                "content": [
                    {"type": "input_text", "text": format_user_prompt(row, product_title_hint)},
                    {"type": "input_image", "image_url": image_url, "detail": image_detail},
                ],
            },
        ],
        "text": {
            "format": {
                "type": "json_schema",
                "name": JSON_SCHEMA["name"],
                "schema": JSON_SCHEMA["schema"],
                "strict": JSON_SCHEMA["strict"],
            }
        },
        "max_output_tokens": max_output_tokens,
    }


def do_json_post(url: str, api_key: str, payload: Dict[str, object], timeout_seconds: int) -> Dict[str, object]:
    body = json.dumps(payload).encode("utf-8")
    request = Request(
        url,
        data=body,
        headers={
            "Authorization": "Bearer {}".format(api_key),
            "Content-Type": "application/json",
        },
        method="POST",
    )
    with urlopen(request, timeout=timeout_seconds) as response:
        return json.loads(response.read().decode("utf-8"))


def extract_response_text(response_json: Dict[str, object]) -> str:
    output_text = response_json.get("output_text")
    if isinstance(output_text, str) and output_text.strip():
        return output_text.strip()

    outputs = response_json.get("output")
    if isinstance(outputs, list):
        text_parts: List[str] = []
        for output in outputs:
            if not isinstance(output, dict):
                continue
            content = output.get("content")
            if not isinstance(content, list):
                continue
            for part in content:
                if not isinstance(part, dict):
                    continue
                if part.get("type") in {"output_text", "text"}:
                    text = part.get("text")
                    if isinstance(text, str) and text.strip():
                        text_parts.append(text.strip())
        if text_parts:
            return "\n".join(text_parts)

    raise ValueError("Could not extract text output from response JSON")


def estimate_cost_usd(model: str, input_tokens: int, output_tokens: int) -> float:
    pricing = MODEL_PRICING.get(model)
    if not pricing:
        return 0.0
    cost = (
        (float(input_tokens) / 1_000_000.0) * pricing["input_per_million"]
        + (float(output_tokens) / 1_000_000.0) * pricing["output_per_million"]
    )
    return round(cost, 6)


def validate_model_result(result: Dict[str, object]) -> Dict[str, object]:
    decision = str(result.get("decision", "") or "").strip()
    reason_code = str(result.get("reason_code", "") or "").strip()
    garment_present = str(result.get("reviewed_garment_present", "") or "").strip()
    garment_visible = str(result.get("full_reviewed_garment_visible", "") or "").strip()
    garment_obscured = str(result.get("reviewed_garment_obscured_or_cut_off", "") or "").strip()
    reason_summary = str(result.get("reason_summary", "") or "").strip()
    confidence = result.get("confidence")

    if decision not in ALLOWED_DECISIONS:
        raise ValueError("Invalid decision: {}".format(decision))
    if reason_code not in ALLOWED_REASON_CODES:
        raise ValueError("Invalid reason_code: {}".format(reason_code))
    if garment_present not in TRI_STATE:
        raise ValueError("Invalid reviewed_garment_present: {}".format(garment_present))
    if garment_visible not in TRI_STATE:
        raise ValueError("Invalid full_reviewed_garment_visible: {}".format(garment_visible))
    if garment_obscured not in TRI_STATE:
        raise ValueError("Invalid reviewed_garment_obscured_or_cut_off: {}".format(garment_obscured))
    if not reason_summary:
        raise ValueError("Missing reason_summary")
    if not isinstance(confidence, (int, float)):
        raise ValueError("Confidence must be numeric")
    validated = dict(result)
    validated["confidence"] = round(float(confidence), 4)
    return validated


def classify_one_row(
    row: Dict[str, str],
    source_file: Path,
    source_row_number: int,
    run_id: str,
    model: str,
    image_detail: str,
    max_output_tokens: int,
    timeout_seconds: int,
    api_key: str,
    dry_run: bool,
) -> Dict[str, object]:
    review_row_key = build_review_row_key(source_file, source_row_number)
    product_title_hint = extract_product_title_hint(row)
    base_result: Dict[str, object] = {
        "review_row_key": review_row_key,
        "source_file": source_file.name,
        "source_row_number": source_row_number,
        "original_url_display": row.get("original_url_display", ""),
        "product_page_url_display": row.get("product_page_url_display", ""),
        "monetized_product_url_display": row.get("monetized_product_url_display", ""),
        "clothing_type_id": row.get("clothing_type_id", ""),
        "product_title_hint": product_title_hint,
        "openai_model": model,
        "openai_image_detail": image_detail,
        "openai_prompt_version": PROMPT_VERSION,
        "openai_run_id": run_id,
    }

    if dry_run:
        base_result.update(
            {
                "openai_decision": "",
                "openai_reason_code": "",
                "openai_reviewed_garment_present": "",
                "openai_full_reviewed_garment_visible": "",
                "openai_reviewed_garment_obscured_or_cut_off": "",
                "openai_reason_summary": "",
                "openai_confidence": "",
                "openai_request_status": "DRY_RUN",
                "openai_input_tokens": 0,
                "openai_output_tokens": 0,
                "openai_total_tokens": 0,
                "openai_estimated_cost_usd": 0.0,
                "openai_error_message": "",
            }
        )
        return base_result

    payload = build_openai_request_payload(model, image_detail, row, product_title_hint, max_output_tokens)
    response_json = do_json_post(
        "https://api.openai.com/v1/responses",
        api_key=api_key,
        payload=payload,
        timeout_seconds=timeout_seconds,
    )
    response_text = extract_response_text(response_json)
    parsed = validate_model_result(json.loads(response_text))
    usage = response_json.get("usage") if isinstance(response_json.get("usage"), dict) else {}
    input_tokens = int(usage.get("input_tokens", 0) or 0)
    output_tokens = int(usage.get("output_tokens", 0) or 0)
    total_tokens = int(usage.get("total_tokens", input_tokens + output_tokens) or 0)
    base_result.update(
        {
            "openai_decision": parsed["decision"],
            "openai_reason_code": parsed["reason_code"],
            "openai_reviewed_garment_present": parsed["reviewed_garment_present"],
            "openai_full_reviewed_garment_visible": parsed["full_reviewed_garment_visible"],
            "openai_reviewed_garment_obscured_or_cut_off": parsed["reviewed_garment_obscured_or_cut_off"],
            "openai_reason_summary": parsed["reason_summary"],
            "openai_confidence": parsed["confidence"],
            "openai_request_status": "OK",
            "openai_input_tokens": input_tokens,
            "openai_output_tokens": output_tokens,
            "openai_total_tokens": total_tokens,
            "openai_estimated_cost_usd": estimate_cost_usd(model, input_tokens, output_tokens),
            "openai_error_message": "",
        }
    )
    return base_result


def build_summary_sheet_rows(all_rows: Sequence[Dict[str, object]]) -> List[Tuple[str, int]]:
    decision_counts = Counter(str(row.get("openai_decision", "") or "").strip() for row in all_rows)
    reason_counts = Counter(str(row.get("openai_reason_code", "") or "").strip() for row in all_rows)
    rows: List[Tuple[str, int]] = []
    for key, value in sorted(decision_counts.items()):
        rows.append(("decision:{}".format(key or "<blank>"), value))
    for key, value in sorted(reason_counts.items()):
        rows.append(("reason:{}".format(key or "<blank>"), value))
    return rows


def build_category_rows(all_rows: Sequence[Dict[str, object]]) -> List[List[object]]:
    grouped: Dict[str, Counter] = defaultdict(Counter)
    for row in all_rows:
        category = str(row.get("clothing_type_id", "") or "").strip() or "blank"
        decision = str(row.get("openai_decision", "") or "").strip() or "<blank>"
        grouped[category][decision] += 1

    rows: List[List[object]] = []
    for category in sorted(grouped):
        counter = grouped[category]
        rows.append(
            [
                category,
                sum(counter.values()),
                counter.get("APPROVED", 0),
                counter.get("REJECTED", 0),
                counter.get("STILL_NEEDS_HUMAN_REVIEW", 0),
                counter.get("<blank>", 0),
            ]
        )
    return rows


def build_category_reason_rows(all_rows: Sequence[Dict[str, object]]) -> List[List[object]]:
    grouped: Dict[Tuple[str, str], int] = defaultdict(int)
    for row in all_rows:
        category = str(row.get("clothing_type_id", "") or "").strip() or "blank"
        reason_code = str(row.get("openai_reason_code", "") or "").strip() or "<blank>"
        grouped[(category, reason_code)] += 1

    rows: List[List[object]] = []
    for (category, reason_code), count in sorted(grouped.items()):
        rows.append([category, reason_code, count])
    return rows


def write_excel_workbook(path: Path, fieldnames: Sequence[str], all_rows: Sequence[Dict[str, object]]) -> None:
    try:
        from openpyxl import Workbook  # type: ignore
        from openpyxl.styles import PatternFill  # type: ignore
    except ModuleNotFoundError as exc:
        raise SystemExit("Missing dependency: openpyxl") from exc

    workbook = Workbook()
    all_results_sheet = workbook.active
    all_results_sheet.title = "all_results"
    all_results_sheet.append(list(fieldnames))
    for row in all_rows:
        all_results_sheet.append([row.get(field, "") for field in fieldnames])
    all_results_sheet.freeze_panes = "A2"
    all_results_sheet.auto_filter.ref = all_results_sheet.dimensions

    fills = {
        "APPROVED": PatternFill(fill_type="solid", start_color="C6EFCE", end_color="C6EFCE"),
        "REJECTED": PatternFill(fill_type="solid", start_color="FFC7CE", end_color="FFC7CE"),
        "STILL_NEEDS_HUMAN_REVIEW": PatternFill(fill_type="solid", start_color="FFEB9C", end_color="FFEB9C"),
    }
    decision_col = fieldnames.index("openai_decision") + 1
    for row_index in range(2, all_results_sheet.max_row + 1):
        decision = str(all_results_sheet.cell(row=row_index, column=decision_col).value or "").strip()
        fill = fills.get(decision)
        if fill:
            for column_index in range(1, all_results_sheet.max_column + 1):
                all_results_sheet.cell(row=row_index, column=column_index).fill = fill

    summary_sheet = workbook.create_sheet("summary")
    summary_sheet.append(["metric", "count"])
    for metric, count in build_summary_sheet_rows(all_rows):
        summary_sheet.append([metric, count])
    summary_sheet.freeze_panes = "A2"
    summary_sheet.auto_filter.ref = summary_sheet.dimensions

    category_sheet = workbook.create_sheet("category_summary")
    category_sheet.append(["clothing_type_id", "total_rows", "approved", "rejected", "still_needs_human_review", "blank"])
    for row in build_category_rows(all_rows):
        category_sheet.append(row)
    category_sheet.freeze_panes = "A2"
    category_sheet.auto_filter.ref = category_sheet.dimensions

    category_reason_sheet = workbook.create_sheet("category_reasons")
    category_reason_sheet.append(["clothing_type_id", "reason_code", "count"])
    for row in build_category_reason_rows(all_rows):
        category_reason_sheet.append(row)
    category_reason_sheet.freeze_panes = "A2"
    category_reason_sheet.auto_filter.ref = category_reason_sheet.dimensions

    approved_sheet = workbook.create_sheet("approved")
    rejected_sheet = workbook.create_sheet("rejected")
    review_sheet = workbook.create_sheet("needs_review")
    for sheet, decision in [
        (approved_sheet, "APPROVED"),
        (rejected_sheet, "REJECTED"),
        (review_sheet, "STILL_NEEDS_HUMAN_REVIEW"),
    ]:
        sheet.append(list(fieldnames))
        for row in all_rows:
            if str(row.get("openai_decision", "") or "").strip() == decision:
                sheet.append([row.get(field, "") for field in fieldnames])
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def markdown_table(headers: Sequence[str], rows: Sequence[Sequence[object]]) -> List[str]:
    lines = [
        "| " + " | ".join(headers) + " |",
        "| " + " | ".join(["---"] * len(headers)) + " |",
    ]
    for row in rows:
        lines.append("| " + " | ".join(str(cell) for cell in row) + " |")
    return lines


def build_run_report(
    run_id: str,
    model: str,
    image_detail: str,
    prompt_version: str,
    started_at: str,
    finished_at: str,
    elapsed_seconds: float,
    all_rows: Sequence[Dict[str, object]],
) -> str:
    total_rows = len(all_rows)
    status_counts = Counter(str(row.get("openai_request_status", "") or "").strip() for row in all_rows)
    decision_counts = Counter(str(row.get("openai_decision", "") or "").strip() for row in all_rows)
    reason_counts = Counter(str(row.get("openai_reason_code", "") or "").strip() for row in all_rows)
    input_tokens = sum(int(float(row.get("openai_input_tokens", 0) or 0)) for row in all_rows)
    output_tokens = sum(int(float(row.get("openai_output_tokens", 0) or 0)) for row in all_rows)
    total_tokens = sum(int(float(row.get("openai_total_tokens", 0) or 0)) for row in all_rows)
    estimated_cost = sum(float(row.get("openai_estimated_cost_usd", 0) or 0) for row in all_rows)
    rows_per_minute = (total_rows / elapsed_seconds) * 60.0 if elapsed_seconds > 0 else 0.0

    category_decision_rows = build_category_rows(all_rows)
    category_reason_counter: Dict[str, Counter] = defaultdict(Counter)
    for row in all_rows:
        category = str(row.get("clothing_type_id", "") or "").strip() or "blank"
        reason_code = str(row.get("openai_reason_code", "") or "").strip()
        if reason_code:
            category_reason_counter[category][reason_code] += 1
    top_category_reason_rows: List[List[object]] = []
    for category in sorted(category_reason_counter):
        for reason_code, count in category_reason_counter[category].most_common(5):
            top_category_reason_rows.append([category, reason_code, count])

    lines = [
        "# OpenAI Chunks 001 002 Experiment Run Report",
        "",
        "## Run Metadata",
        "",
        "- run id: `{}`".format(run_id),
        "- model: `{}`".format(model),
        "- image detail: `{}`".format(image_detail),
        "- prompt version: `{}`".format(prompt_version),
        "- run started at: `{}`".format(started_at),
        "- run finished at: `{}`".format(finished_at),
        "- elapsed seconds: `{:.2f}`".format(elapsed_seconds),
        "- elapsed minutes: `{:.2f}`".format(elapsed_seconds / 60.0 if elapsed_seconds else 0.0),
        "- images per minute: `{:.2f}`".format(rows_per_minute),
        "",
        "## Completion Summary",
        "",
        "- total rows attempted: `{}`".format(total_rows),
        "- total rows completed: `{}`".format(status_counts.get("OK", 0)),
        "- total rows failed: `{}`".format(status_counts.get("ERROR", 0)),
        "- total rows dry run: `{}`".format(status_counts.get("DRY_RUN", 0)),
        "",
        "## Decision Histogram",
        "",
    ]
    lines.extend(markdown_table(["decision", "count"], [[k or "<blank>", v] for k, v in sorted(decision_counts.items())]))
    lines.extend(["", "## Reason Histogram", ""])
    lines.extend(markdown_table(["reason_code", "count"], [[k or "<blank>", v] for k, v in reason_counts.most_common()]))
    lines.extend(["", "## Category Decision Overview", ""])
    lines.extend(markdown_table(
        ["clothing_type_id", "total_rows", "approved", "rejected", "still_needs_human_review", "blank"],
        category_decision_rows,
    ))
    lines.extend(["", "## Category Reason Label Overview", ""])
    lines.extend(markdown_table(["clothing_type_id", "reason_code", "count"], top_category_reason_rows))
    lines.extend(
        [
            "",
            "## Token And Cost Summary",
            "",
            "- total input tokens: `{}`".format(input_tokens),
            "- total output tokens: `{}`".format(output_tokens),
            "- total tokens: `{}`".format(total_tokens),
            "- estimated total cost usd: `{:.4f}`".format(estimated_cost),
            "",
        ]
    )
    return "\n".join(lines) + "\n"


def build_experiment_rows(
    input_files: Sequence[Path],
    limit: int,
) -> List[Tuple[Path, int, Dict[str, str]]]:
    rows: List[Tuple[Path, int, Dict[str, str]]] = []
    for input_file in input_files:
        _fieldnames, input_rows = read_csv_rows(input_file)
        for source_row_number, row in enumerate(input_rows, start=2):
            rows.append((input_file, source_row_number, row))
            if limit > 0 and len(rows) >= limit:
                return rows
    return rows


def main() -> None:
    args = parse_args()
    load_env_file(args.env_file.resolve())
    input_files = [path.resolve() for path in args.input_files]
    for path in input_files:
        if not path.exists():
            raise SystemExit("Input CSV not found: {}".format(path))

    api_key = os.environ.get("OPENAI_API_KEY", "").strip()
    if not api_key and not args.dry_run:
        raise SystemExit("OPENAI_API_KEY is not set. Set it or use --dry-run.")

    run_id = args.run_name.strip() or "openai_chunks_001_002_{}".format(
        datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
    )
    paths = ensure_run_paths(args.output_dir.resolve(), run_id)
    existing_results = load_existing_results(paths.row_results_csv) if args.resume else {}
    experiment_rows = build_experiment_rows(input_files, args.limit)

    started_at = utc_now_iso()
    start_time = time.perf_counter()
    completed_results: Dict[str, Dict[str, object]] = {key: dict(value) for key, value in existing_results.items()}

    pending_rows = []
    for input_file, source_row_number, row in experiment_rows:
        review_row_key = build_review_row_key(input_file, source_row_number)
        if args.resume and review_row_key in existing_results:
            continue
        pending_rows.append((input_file, source_row_number, row))

    def worker(item: Tuple[Path, int, Dict[str, str]]) -> Dict[str, object]:
        input_file, source_row_number, row = item
        try:
            return classify_one_row(
                row=row,
                source_file=input_file,
                source_row_number=source_row_number,
                run_id=run_id,
                model=args.model,
                image_detail=args.image_detail,
                max_output_tokens=args.max_output_tokens,
                timeout_seconds=args.timeout_seconds,
                api_key=api_key,
                dry_run=args.dry_run,
            )
        except Exception as exc:  # noqa: BLE001
            review_row_key = build_review_row_key(input_file, source_row_number)
            return {
                "review_row_key": review_row_key,
                "source_file": input_file.name,
                "source_row_number": source_row_number,
                "original_url_display": row.get("original_url_display", ""),
                "product_page_url_display": row.get("product_page_url_display", ""),
                "monetized_product_url_display": row.get("monetized_product_url_display", ""),
                "clothing_type_id": row.get("clothing_type_id", ""),
                "product_title_hint": extract_product_title_hint(row),
                "openai_decision": "",
                "openai_reason_code": "",
                "openai_reviewed_garment_present": "",
                "openai_full_reviewed_garment_visible": "",
                "openai_reviewed_garment_obscured_or_cut_off": "",
                "openai_reason_summary": "",
                "openai_confidence": "",
                "openai_model": args.model,
                "openai_image_detail": args.image_detail,
                "openai_prompt_version": PROMPT_VERSION,
                "openai_run_id": run_id,
                "openai_request_status": "ERROR",
                "openai_input_tokens": 0,
                "openai_output_tokens": 0,
                "openai_total_tokens": 0,
                "openai_estimated_cost_usd": 0.0,
                "openai_error_message": str(exc),
            }

    if pending_rows:
        with ThreadPoolExecutor(max_workers=max(1, args.max_workers)) as executor:
            futures = [executor.submit(worker, item) for item in pending_rows]
            for future in as_completed(futures):
                result = future.result()
                completed_results[str(result["review_row_key"])] = result
                append_csv_row(paths.row_results_csv, RESULT_COLUMNS, result)

    ordered_results: List[Dict[str, object]] = []
    for input_file, source_row_number, _row in experiment_rows:
        review_row_key = build_review_row_key(input_file, source_row_number)
        result = completed_results.get(review_row_key)
        if result:
            ordered_results.append(result)

    finished_at = utc_now_iso()
    elapsed_seconds = time.perf_counter() - start_time

    write_excel_workbook(paths.workbook_path, RESULT_COLUMNS, ordered_results)
    report_text = build_run_report(
        run_id=run_id,
        model=args.model,
        image_detail=args.image_detail,
        prompt_version=PROMPT_VERSION,
        started_at=started_at,
        finished_at=finished_at,
        elapsed_seconds=elapsed_seconds,
        all_rows=ordered_results,
    )
    paths.report_path.write_text(report_text, encoding="utf-8")

    print("Wrote row results to {}".format(paths.row_results_csv))
    print("Wrote workbook to {}".format(paths.workbook_path))
    print("Wrote report to {}".format(paths.report_path))
    print("Rows in workbook: {}".format(len(ordered_results)))


if __name__ == "__main__":
    main()
