#!/usr/bin/env python3
"""Add a rejection-reason dropdown to generated Supabase image review workbooks."""

from __future__ import annotations
import sys

from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


REPO_ROOT = Path(__file__).resolve().parents[4]
PIPELINE_SCRIPTS_DIR = REPO_ROOT / "data-pipelines" / "scripts"
if str(PIPELINE_SCRIPTS_DIR) not in sys.path:
    sys.path.insert(0, str(PIPELINE_SCRIPTS_DIR))

from pipeline_paths import archive_root, cv_annotated_pending_human_review_root  # noqa: E402

LEGACY_OUTPUTS_ARCHIVE = archive_root() / "old_outputs" / "repo_outputs_archive" / "supabase_output_cleanup_2026_05_29"
CV_EXPERIMENTS_DIR = LEGACY_OUTPUTS_ARCHIVE / "cv_experiments"

PACKAGE_DIR = cv_annotated_pending_human_review_root() / "supabase_production_image_review_2026_05_28_s3_refresh_sovrn_prioritized"

REJECTION_REASONS = [
    "LOW_RESOLUTION_AFTER_URL_REPAIR",
    "TOO_DARK",
    "TOO_BRIGHT_OR_WASHED_OUT",
    "BLURRY_OR_MOTION_BLUR",
    "GRAINY_OR_NOISY",
    "GARMENT_CUT_OFF",
    "GARMENT_TOP_COVERED",
    "GARMENT_BOTTOM_CUT_OFF",
    "GARMENT_OBSCURED",
    "PERSON_TOO_FAR",
    "TARGET_WEARER_AMBIGUOUS",
    "BAD_ANGLE_TOP_DOWN",
    "BAD_ANGLE_SIDE_OR_TWISTED",
    "BACKGROUND_TOO_CLUTTERED",
    "DISTRACTING_OBJECTS",
    "NOT_WORN_BY_PERSON",
    "NO_PERSON_VISIBLE",
    "WRONG_PRODUCT_CONTEXT",
    "DUPLICATE_OR_NEAR_DUPLICATE",
    "IMAGE_FETCH_FAILED",
    "OTHER",
]


def remove_named_validations(sheet, column_letter: str) -> None:
    kept = []
    target_prefix = f"{column_letter}2:"
    for validation in sheet.data_validations.dataValidation:
        ranges = str(validation.sqref)
        if ranges.startswith(target_prefix) or ranges == column_letter:
            continue
        kept.append(validation)
    sheet.data_validations.dataValidation = kept


def ensure_rejection_reason_column(path: Path) -> None:
    workbook = load_workbook(path)
    sheet = workbook["Review"] if "Review" in workbook.sheetnames else workbook.active

    headers = [sheet.cell(1, column).value for column in range(1, sheet.max_column + 1)]
    if "rejection_reason" in headers:
        reason_col = headers.index("rejection_reason") + 1
    else:
        sheet.insert_cols(2)
        reason_col = 2
        sheet.cell(1, reason_col).value = "rejection_reason"

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    reason_letter = get_column_letter(reason_col)
    sheet.column_dimensions[reason_letter].width = 34
    for row in range(2, sheet.max_row + 1):
        sheet.cell(row, reason_col).alignment = Alignment(wrap_text=True, vertical="top")

    list_sheet_name = "Rejection Reason Options"
    if list_sheet_name in workbook.sheetnames:
        list_sheet = workbook[list_sheet_name]
        list_sheet.delete_rows(1, list_sheet.max_row)
    else:
        list_sheet = workbook.create_sheet(list_sheet_name)
    list_sheet.sheet_state = "hidden"
    list_sheet.cell(1, 1).value = "rejection_reason"
    for index, reason in enumerate(REJECTION_REASONS, start=2):
        list_sheet.cell(index, 1).value = reason

    remove_named_validations(sheet, reason_letter)
    validation = DataValidation(
        type="list",
        formula1=f"'{list_sheet_name}'!$A$2:$A${len(REJECTION_REASONS) + 1}",
        allow_blank=True,
    )
    sheet.add_data_validation(validation)
    validation.add(f"{reason_letter}2:{reason_letter}{max(2, sheet.max_row)}")

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(sheet.max_column)}{sheet.max_row}"
    workbook.save(path)


def update_readme() -> None:
    readme = PACKAGE_DIR / "README.md"
    if not readme.exists():
        return
    text = readme.read_text(encoding="utf-8")
    addition = (
        "\n## Rejection Reasons\n\n"
        "When `production_decision` is `REJECT`, choose the closest `rejection_reason`. "
        "Use `OTHER` only when none of the listed labels fit.\n\n"
        + "\n".join(f"- `{reason}`" for reason in REJECTION_REASONS)
        + "\n"
    )
    if "## Rejection Reasons" not in text:
        text = text.rstrip() + "\n" + addition
        readme.write_text(text, encoding="utf-8")


def main() -> None:
    files = sorted(PACKAGE_DIR.glob("*.xlsx"))
    for index, path in enumerate(files, start=1):
        ensure_rejection_reason_column(path)
        if index % 25 == 0:
            print(f"updated {index}/{len(files)}", flush=True)
    update_readme()
    print(f"updated {len(files)} workbooks in {PACKAGE_DIR}")


if __name__ == "__main__":
    main()
