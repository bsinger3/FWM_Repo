#!/usr/bin/env python3
"""Create a one-tab Google-Sheets-friendly YOLO error review queue."""

from __future__ import annotations
import sys

import csv
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation


REPO_ROOT = Path(__file__).resolve().parents[4]
PIPELINE_SCRIPTS_DIR = REPO_ROOT / "data-pipelines" / "scripts"
if str(PIPELINE_SCRIPTS_DIR) not in sys.path:
    sys.path.insert(0, str(PIPELINE_SCRIPTS_DIR))

from pipeline_paths import archive_root, cv_annotated_pending_human_review_root  # noqa: E402

LEGACY_OUTPUTS_ARCHIVE = archive_root() / "old_outputs" / "repo_outputs_archive" / "supabase_output_cleanup_2026_05_29"
CV_EXPERIMENTS_DIR = LEGACY_OUTPUTS_ARCHIVE / "cv_experiments"

EXP_DIR = CV_EXPERIMENTS_DIR / "yolo_segmentation_crop_reasons_broad_2026_05_25"
SOURCE_CSV = EXP_DIR / "yolo_segmentation_crop_reason_rows.csv"
OUTPUT_XLSX = EXP_DIR / "yolo_error_review_queue_single_tab.xlsx"

ANSWER_CHOICES = [
    "MODEL_RIGHT",
    "MODEL_WRONG",
    "GROUND_TRUTH_WRONG",
    "UNSURE",
]

HEADERS = [
    "your_answer",
    "image_preview",
    "review_prompt",
    "model_bucket",
    "model_rule",
    "current_ground_truth",
    "original_url_display",
    "product_page_url_display",
    "source_site_display",
    "primary_reason_code",
    "secondary_reason_code",
    "labeler_notes",
    "seg_person_count",
    "seg_mask_area_pct",
    "seg_bbox_area_pct",
    "seg_bbox_height_pct",
    "llm_summary",
    "review_row_key",
]


def read_rows() -> list[dict[str, str]]:
    with SOURCE_CSV.open(newline="", encoding="utf-8-sig") as handle:
        return list(csv.DictReader(handle))


def truthy(value: object) -> bool:
    return str(value or "").strip().lower() in {"true", "yes", "1", "x", "checked"}


def to_float(value: object, default: float = 0.0) -> float:
    try:
        if value is None or str(value).strip() == "":
            return default
        return float(value)
    except ValueError:
        return default


def pred_needs_crop(row: dict[str, str]) -> bool:
    return to_float(row.get("seg_mask_area_pct"), default=1.0) < 0.20


def pred_not_worn(row: dict[str, str]) -> bool:
    return to_float(row.get("seg_person_count"), default=999.0) == 0


def decision(row: dict[str, str]) -> str:
    return str(row.get("final_human_decision_norm") or row.get("final_human_decision") or "").upper()


def add_candidate(candidates: list[dict[str, str]], row: dict[str, str], bucket: str, prompt: str, rule: str, truth: str) -> None:
    item = dict(row)
    item["_model_bucket"] = bucket
    item["_review_prompt"] = prompt
    item["_model_rule"] = rule
    item["_current_ground_truth"] = truth
    candidates.append(item)


def build_candidates(rows: list[dict[str, str]]) -> list[dict[str, str]]:
    candidates: list[dict[str, str]] = []
    seen: set[tuple[str, str]] = set()

    for row in rows:
        needs_truth = truthy(row.get("truth_NEEDS_CROP"))
        needs_pred = pred_needs_crop(row)
        not_worn_truth = truthy(row.get("truth_NOT_WORN_BY_PERSON"))
        not_worn_pred = pred_not_worn(row)

        if needs_truth and not needs_pred:
            add_candidate(
                candidates,
                row,
                "NEEDS_CROP_MISSED",
                "Should this image be treated as NEEDS_CROP even though YOLO did not flag it?",
                "NEEDS_CROP if seg_mask_area_pct < 0.20",
                "NEEDS_CROP",
            )
        if not needs_truth and needs_pred and decision(row) == "APPROVED":
            add_candidate(
                candidates,
                row,
                "NEEDS_CROP_FALSE_FLAG_APPROVED",
                "Did YOLO incorrectly flag this approved image as needing crop?",
                "NEEDS_CROP if seg_mask_area_pct < 0.20",
                "APPROVED / not NEEDS_CROP",
            )
        if not needs_truth and needs_pred and decision(row) == "REJECTED":
            add_candidate(
                candidates,
                row,
                "NEEDS_CROP_FLAGGED_OTHER_REJECT",
                "Is YOLO finding a real crop issue here, or is the existing non-crop label enough?",
                "NEEDS_CROP if seg_mask_area_pct < 0.20",
                "REJECTED for another reason",
            )
        if not_worn_truth and not not_worn_pred:
            add_candidate(
                candidates,
                row,
                "NOT_WORN_MISSED",
                "Should this remain NOT_WORN_BY_PERSON even though YOLO found a person?",
                "NOT_WORN_BY_PERSON if seg_person_count == 0",
                "NOT_WORN_BY_PERSON",
            )
        if not not_worn_truth and not_worn_pred:
            add_candidate(
                candidates,
                row,
                "NOT_WORN_FALSE_FLAG",
                "Did YOLO incorrectly treat this image as having no wearable person?",
                "NOT_WORN_BY_PERSON if seg_person_count == 0",
                "not NOT_WORN_BY_PERSON",
            )

    deduped: list[dict[str, str]] = []
    for item in candidates:
        key = (item.get("review_row_key", ""), item.get("_model_bucket", ""))
        if key in seen:
            continue
        seen.add(key)
        deduped.append(item)
    return deduped


def value_for(item: dict[str, str], header: str) -> object:
    if header == "your_answer":
        return ""
    if header == "model_bucket":
        return item.get("_model_bucket", "")
    if header == "review_prompt":
        return item.get("_review_prompt", "")
    if header == "model_rule":
        return item.get("_model_rule", "")
    if header == "current_ground_truth":
        return item.get("_current_ground_truth", "")
    return item.get(header, "")


def style_header(sheet) -> None:
    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)
    for cell in sheet[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(wrap_text=True, vertical="top")


def add_instructions(sheet, start_row: int) -> None:
    instructions = [
        "Fill only column A (your_answer).",
        "MODEL_RIGHT = the model rule/prediction is useful for this row.",
        "MODEL_WRONG = the current human label looks right and the model behavior is wrong.",
        "GROUND_TRUTH_WRONG = the current human label/reason should be changed.",
        "UNSURE = ambiguous or needs discussion.",
    ]
    row = start_row
    for text in instructions:
        sheet.cell(row=row, column=1, value=text)
        sheet.cell(row=row, column=1).alignment = Alignment(wrap_text=True, vertical="top")
        row += 1


def main() -> None:
    rows = read_rows()
    candidates = build_candidates(rows)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Review Queue"
    sheet.append(HEADERS)
    for item in candidates:
        sheet.append([value_for(item, header) for header in HEADERS])

    style_header(sheet)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:R{sheet.max_row}"

    answer_validation = DataValidation(type="list", formula1=f'"{",".join(ANSWER_CHOICES)}"', allow_blank=True)
    sheet.add_data_validation(answer_validation)
    answer_validation.add(f"A2:A{sheet.max_row}")

    widths = {
        "A": 24,
        "B": 26,
        "C": 54,
        "D": 30,
        "E": 32,
        "F": 26,
        "G": 42,
        "H": 42,
        "I": 22,
        "J": 24,
        "K": 24,
        "L": 38,
        "Q": 64,
        "R": 28,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width
    for column in ["M", "N", "O", "P"]:
        sheet.column_dimensions[column].width = 16

    for row in sheet.iter_rows(min_row=2):
        sheet.row_dimensions[row[0].row].height = 72
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    add_instructions(sheet, sheet.max_row + 3)
    workbook.save(OUTPUT_XLSX)
    print(OUTPUT_XLSX)
    print(f"review rows: {len(candidates)}")


if __name__ == "__main__":
    main()
