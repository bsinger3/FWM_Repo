#!/usr/bin/env python3
"""Run the known YOLO/pose CV gate over Supabase approval-candidate workbooks."""

from __future__ import annotations

import argparse
import csv
import os
import re
import time
from collections import Counter
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from cv_rules_workflow_lib import (
    DEFAULT_VENDOR_DIR,
    DEFAULT_YOLO_DETECT_MODEL,
    DEFAULT_YOLO_POSE_MODEL,
    compute_yolo_coverage_score,
    extract_person_metrics,
    fetch_rgb_image,
    round_or_blank,
    summarize_main_person,
    bootstrap_vendor_paths,
)
from build_supabase_image_review_package import gather_candidates


REPO_ROOT = Path(__file__).resolve().parents[4]
DEFAULT_INPUT_PACKAGE = REPO_ROOT / "outputs/supabase_production_image_review_2026_05_28_s3_refresh_sovrn_prioritized"
DEFAULT_OUTPUT_PACKAGE = REPO_ROOT / "outputs/supabase_production_image_review_2026_05_28_s3_refresh_cv_gated"
MAX_ROWS_PER_WORKBOOK = 1000

BASE_COLUMNS = [
    "production_decision",
    "rejection_reason",
    "review_notes",
    "image_preview",
    "sovrn_has_payout",
    "sovrn_priority",
    "sovrn_payout_priority_rank",
    "sovrn_estimated_commission_per_click",
    "sovrn_pricing",
    "sovrn_merchant_group",
    "sorter_recommendation",
    "sorter_reason_codes",
    "cv_decision",
    "cv_reason_code",
    "cv_reason_summary",
    "person_count_yolo_detect",
    "main_person_height_pct_yolo_detect",
    "main_person_bbox_area_pct_yolo_detect",
    "body_coverage_score_yolo_pose",
    "has_face_yunet",
    "source_family",
    "source_site_display",
    "review_row_key",
    "image_url_to_use",
    "raw_scraped_image_url",
    "needs_url_update",
    "product_page_url_display",
    "monetized_product_url_display",
    "brand",
    "product_title_raw",
    "product_category_raw",
    "product_variant_raw",
    "clothing_type_id",
    "size_display",
    "height_in_display",
    "weight_lbs_display",
    "waist_in",
    "hips_in_display",
    "bust_in_display",
    "bra_band_in_display",
    "cupsize_display",
    "inseam_inches_display",
    "user_comment",
    "source_file",
    "source_row_number",
]

PRODUCTION_DECISIONS = "APPROVE,REJECT,NEEDS_MORE_REVIEW,SKIP"
REJECTION_REASONS = [
    "LOW_RESOLUTION_AFTER_URL_REPAIR",
    "TOO_DARK",
    "TOO_BRIGHT_OR_WASHED_OUT",
    "BLURRY_OR_MOTION_BLUR",
    "GRAINY_OR_NOISY",
    "GARMENT_CUT_OFF",
    "GARMENT_TOP_COVERED",
    "GARMENT_BOTTOM_CUT_OFF",
    "GARMENT_OBSCURED",
    "PERSON_TOO_FAR",
    "TARGET_WEARER_AMBIGUOUS",
    "BAD_ANGLE_TOP_DOWN",
    "BAD_ANGLE_SIDE_OR_TWISTED",
    "BACKGROUND_TOO_CLUTTERED",
    "DISTRACTING_OBJECTS",
    "NOT_WORN_BY_PERSON",
    "NO_PERSON_VISIBLE",
    "WRONG_PRODUCT_CONTEXT",
    "DUPLICATE_OR_NEAR_DUPLICATE",
    "IMAGE_FETCH_FAILED",
    "OTHER",
]

CV_REJECTION_REASON = {
    "NO_PERSON": "NO_PERSON_VISIBLE",
    "MULTIPLE_PEOPLE": "TARGET_WEARER_AMBIGUOUS",
    "LOW_BODY_COVERAGE": "GARMENT_CUT_OFF",
    "SUBJECT_TOO_SMALL": "PERSON_TOO_FAR",
    "SMALL_SUBJECT_NO_FACE": "PERSON_TOO_FAR",
}

METRIC_WEIGHT_SITES = {
    "babyboofashion.com",
    "petalandpup.com",
    "popflexactive.com",
    "wearfigs.com",
}

WEIGHT_CHANGE_RE = re.compile(
    r"\b(?:lost|loss|weight\s+loss|gained|gain(?:ed|ing)?\s+(?:about\s+|around\s+|~\s*)?\d+(?:\.\d+)?)\b",
    re.IGNORECASE,
)
NON_BODY_WEIGHT_RE = re.compile(
    r"\b(?:toddler|daughter|son|child|kid|year old|year-old|items? weigh|total of|fabric weight)\b",
    re.IGNORECASE,
)


def parse_numeric(value: object) -> float | None:
    text = str(value or "").strip().replace(",", "")
    if not text:
        return None
    unit_match = re.search(r"(\d+(?:\.\d+)?)\s*(?:lbs?|pounds?|kgs?|kilograms?)\b", text, re.IGNORECASE)
    if unit_match:
        return float(unit_match.group(1))
    matches = [float(match.group(0)) for match in re.finditer(r"\d+(?:\.\d+)?", text)]
    for match in matches:
        if 70 <= match <= 350:
            return match
    if matches:
        return matches[0]
    try:
        return float(text)
    except ValueError:
        return None


def format_measurement(value: float) -> float | int:
    rounded = round(value, 1)
    return int(rounded) if rounded.is_integer() else rounded


def source_uses_metric_weight(row: dict[str, object]) -> bool:
    source_site = str(row.get("source_site_display") or "").lower()
    return any(site in source_site for site in METRIC_WEIGHT_SITES)


def number_has_unit_context(text: str, value: float, unit_pattern: str) -> bool:
    value_patterns = {str(int(value)) if float(value).is_integer() else str(value)}
    value_patterns.add(f"{value:.1f}".rstrip("0").rstrip("."))
    for value_text in value_patterns:
        escaped = re.escape(value_text)
        if re.search(rf"\b{escaped}\s*(?:{unit_pattern})\b", text, re.IGNORECASE):
            return True
        if re.search(rf"\b(?:{unit_pattern})\s*[:=]?\s*{escaped}\b", text, re.IGNORECASE):
            return True
    return False


def recover_current_weight_lbs(comment: str, excluded_value: float) -> float | None:
    patterns = [
        r"\b(?:currently|current weight|for reference|i am|i'm|am|weigh|weight)\D{0,40}(\d{2,3}(?:\.\d+)?)\s*(?:lbs?|pounds?)\b",
        r"\b(\d{2,3}(?:\.\d+)?)\s*(?:lbs?|pounds?)\b",
    ]
    for pattern in patterns:
        for match in re.finditer(pattern, comment, re.IGNORECASE):
            value = parse_numeric(match.group(1))
            if value is not None and 70 <= value <= 350 and abs(value - excluded_value) > 0.01:
                return value
    return None


def normalize_measurements(row: dict[str, object]) -> None:
    """Repair scraper values that landed in the wrong measurement unit."""
    weight_text = str(row.get("weight_lbs_display") or "")
    weight = parse_numeric(row.get("weight_lbs_display"))
    if weight is None:
        return

    comment = str(row.get("user_comment") or "")
    context = f"{weight_text} {comment}"
    has_kg_context = number_has_unit_context(context, weight, r"kgs?|kilograms?")
    has_lb_context = number_has_unit_context(context, weight, r"lbs?|pounds?")

    if has_kg_context:
        row["weight_lbs_display"] = format_measurement(weight * 2.2046226218)
    elif 30 <= weight < 70 and (WEIGHT_CHANGE_RE.search(comment) or NON_BODY_WEIGHT_RE.search(comment)):
        recovered = recover_current_weight_lbs(comment, weight)
        row["weight_lbs_display"] = format_measurement(recovered) if recovered is not None else ""
    elif source_uses_metric_weight(row) and not has_lb_context and 30 <= weight < 100:
        row["weight_lbs_display"] = format_measurement(weight * 2.2046226218)
    elif has_lb_context and 30 <= weight < 70:
        row["weight_lbs_display"] = ""
    elif weight < 70:
        row["weight_lbs_display"] = ""


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument("--input-package", type=Path, default=DEFAULT_INPUT_PACKAGE)
    parser.add_argument("--output-package", type=Path, default=DEFAULT_OUTPUT_PACKAGE)
    parser.add_argument("--read-xlsx-package", action="store_true", help="Read approval rows from existing xlsx files instead of rebuilding candidate rows.")
    parser.add_argument("--limit", type=int, default=0, help="Optional smoke-test row limit.")
    parser.add_argument("--timeout", type=float, default=20.0)
    parser.add_argument("--yolo-batch-size", type=int, default=8)
    parser.add_argument("--download-workers", type=int, default=8)
    parser.add_argument("--vendor-dir", type=Path, default=DEFAULT_VENDOR_DIR)
    parser.add_argument("--yolo-detect-model", type=Path, default=DEFAULT_YOLO_DETECT_MODEL)
    parser.add_argument("--yolo-pose-model", type=Path, default=DEFAULT_YOLO_POSE_MODEL)
    parser.add_argument("--verbose", action="store_true")
    return parser.parse_args()


def headers_for(sheet) -> dict[str, int]:
    return {str(sheet.cell(1, column).value): column for column in range(1, sheet.max_column + 1)}


def read_approval_rows(package_dir: Path, limit: int) -> list[dict[str, object]]:
    rows: list[dict[str, object]] = []
    files = sorted(package_dir.glob("supabase_image_review_approve_candidates_part_*.xlsx"))
    for path in files:
        workbook = load_workbook(path, read_only=True, data_only=False)
        sheet = workbook["Review"] if "Review" in workbook.sheetnames else workbook.active
        headers = headers_for(sheet)
        for row_index in range(2, sheet.max_row + 1):
            row = {column: sheet.cell(row_index, column_index).value or "" for column, column_index in headers.items()}
            row["source_review_workbook"] = path.name
            row["source_review_workbook_row"] = row_index
            row["original_url_display"] = row.get("image_url_to_use") or row.get("raw_scraped_image_url") or ""
            rows.append(row)
            if limit and len(rows) >= limit:
                return rows
    return rows


def load_approval_rows_from_builder(limit: int) -> list[dict[str, object]]:
    rows = [dict(row) for row in gather_candidates()["approve_candidates"]]
    if limit:
        rows = rows[:limit]
    for row in rows:
        row["original_url_display"] = row.get("image_url_to_use") or row.get("raw_scraped_image_url") or ""
    return rows


def route_row(row: dict[str, object]) -> str:
    decision = str(row.get("cv_decision") or "")
    reason = str(row.get("cv_reason_code") or "")
    if decision == "APPROVE":
        row["sorter_recommendation"] = "CV_APPROVE_CANDIDATE"
        row["sorter_reason_codes"] = reason
        return "approve_candidates"
    if decision == "REJECT":
        row["sorter_recommendation"] = "CV_REJECT"
        row["sorter_reason_codes"] = reason
        row["production_decision"] = row.get("production_decision") or "REJECT"
        row["rejection_reason"] = row.get("rejection_reason") or CV_REJECTION_REASON.get(reason, "OTHER")
        return "disapprove_candidates"
    row["sorter_recommendation"] = "CV_NEEDS_HUMAN_REVIEW"
    row["sorter_reason_codes"] = reason
    return "needs_human_review"


def import_yolo_dependencies():
    os.environ.setdefault("YOLO_CONFIG_DIR", "/tmp/Ultralytics")
    os.environ.setdefault("MPLCONFIGDIR", "/tmp/matplotlib-codex")
    import numpy as np  # type: ignore
    import requests  # type: ignore
    from PIL import Image  # type: ignore
    from ultralytics import YOLO  # type: ignore

    return np, requests, Image, YOLO


def evaluate_yolo_gate(row: dict[str, object]) -> None:
    person_count = float(row.get("person_count_yolo_detect") or 0)
    subject_height = float(row.get("main_person_height_pct_yolo_detect") or 0)
    subject_area = float(row.get("main_person_bbox_area_pct_yolo_detect") or 0)
    body_coverage = float(row.get("body_coverage_score_yolo_pose") or 0)

    if int(person_count) == 0:
        row["cv_decision"] = "REJECT"
        row["cv_reason_code"] = "NO_PERSON"
        row["cv_reason_summary"] = "No person detected"
    elif person_count > 1:
        row["cv_decision"] = "REJECT"
        row["cv_reason_code"] = "MULTIPLE_PEOPLE"
        row["cv_reason_summary"] = "Multiple people detected"
    elif body_coverage < 66.7:
        row["cv_decision"] = "REJECT"
        row["cv_reason_code"] = "LOW_BODY_COVERAGE"
        row["cv_reason_summary"] = "Too little body visible"
    elif subject_height < 0.50:
        row["cv_decision"] = "REJECT"
        row["cv_reason_code"] = "SUBJECT_TOO_SMALL"
        row["cv_reason_summary"] = "Person too small in frame"
    elif body_coverage >= 75.0 and subject_height >= 0.60 and subject_area >= 0.15:
        row["cv_decision"] = "APPROVE"
        row["cv_reason_code"] = "CLEAR_PASS"
        row["cv_reason_summary"] = "Single person, strong framing, enough body visible"
    elif 66.7 <= body_coverage < 75.0:
        row["cv_decision"] = "REVIEW"
        row["cv_reason_code"] = "BORDERLINE_BODY_COVERAGE"
        row["cv_reason_summary"] = "Body visibility is borderline"
    elif 0.50 <= subject_height < 0.60:
        row["cv_decision"] = "REVIEW"
        row["cv_reason_code"] = "BORDERLINE_SUBJECT_SIZE"
        row["cv_reason_summary"] = "Person size is borderline"
    else:
        row["cv_decision"] = "REVIEW"
        row["cv_reason_code"] = "BORDERLINE_COMPOSITION"
        row["cv_reason_summary"] = "Composition is borderline"


def summarize_material_people(valid_indexes: list[int], boxes: list[list[float]], image_width: int, image_height: int) -> tuple[int, float, float, int | None]:
    """Count only people large enough to create target-wearer ambiguity."""
    person_count, height_pct, area_pct, best_index = summarize_main_person(valid_indexes, boxes, image_width, image_height)
    if best_index is None:
        return person_count, height_pct, area_pct, best_index

    material_count = 0
    main_area = 0.0
    metrics: list[tuple[int, float, float]] = []
    for index in valid_indexes:
        x1, y1, x2, y2 = [float(value) for value in boxes[index]]
        width = max(0.0, x2 - x1)
        height = max(0.0, y2 - y1)
        current_height_pct = height / float(image_height) if image_height else 0.0
        current_area_pct = (width * height) / float(image_width * image_height) if image_width and image_height else 0.0
        metrics.append((index, current_height_pct, current_area_pct))
        if index == best_index:
            main_area = current_area_pct

    for index, current_height_pct, current_area_pct in metrics:
        if index == best_index:
            material_count += 1
        elif current_height_pct >= 0.45 or current_area_pct >= 0.08 or (main_area and current_area_pct >= main_area * 0.35):
            material_count += 1

    return material_count, height_pct, area_pct, best_index


def enrich_rows_with_yolo_gate(
    rows: list[dict[str, object]],
    detect_model,
    pose_model,
    np,
    session,
    Image,
    timeout: float,
    min_person_confidence: float,
    min_pose_keypoint_confidence: float,
    yolo_batch_size: int,
    download_workers: int,
    verbose: bool,
) -> list[dict[str, object]]:
    enriched = [dict(row) for row in rows]

    def download_one(index: int):
        url = str(enriched[index].get("original_url_display") or "").strip()
        return index, fetch_rgb_image(url, timeout, session, Image)

    processed = 0
    batch_size = max(1, yolo_batch_size)
    worker_count = max(1, download_workers)
    indexes = list(range(len(enriched)))

    with ThreadPoolExecutor(max_workers=worker_count) as executor:
        for start in range(0, len(indexes), batch_size):
            batch_indexes = indexes[start : start + batch_size]
            downloaded = []
            futures = {executor.submit(download_one, index): index for index in batch_indexes}
            for future in as_completed(futures):
                index = futures[future]
                try:
                    downloaded.append(future.result())
                except Exception as exc:  # noqa: BLE001
                    row = enriched[index]
                    row["cv_decision"] = "REVIEW"
                    row["cv_reason_code"] = "IMAGE_FETCH_FAILED"
                    row["cv_reason_summary"] = f"Image fetch failed: {exc}"
                    if verbose and os.environ.get("QUIET_FETCH_FAILURES") != "1":
                        print(f"[{index + 1}] fetch failed: {exc}", flush=True)

            downloaded.sort(key=lambda item: item[0])
            if not downloaded:
                continue

            images = [image_rgb for _index, image_rgb in downloaded]
            detect_results = detect_model.predict(images, verbose=False, device="cpu")
            pose_results = pose_model.predict(images, verbose=False, device="cpu")

            for (index, image_rgb), detect_result, pose_result in zip(downloaded, detect_results, pose_results):
                row = enriched[index]
                image_height, image_width = image_rgb.size[1], image_rgb.size[0]

                valid_indexes, boxes = extract_person_metrics(detect_result, image_width, image_height, min_person_confidence)
                person_count, height_pct, area_pct, _best_index = summarize_material_people(valid_indexes, boxes, image_width, image_height)
                row["person_count_yolo_detect"] = person_count
                row["main_person_height_pct_yolo_detect"] = round_or_blank(height_pct)
                row["main_person_bbox_area_pct_yolo_detect"] = round_or_blank(area_pct)

                pose_valid_indexes, pose_boxes = extract_person_metrics(pose_result, image_width, image_height, min_person_confidence)
                _pose_count, _pose_height_pct, _pose_area_pct, pose_best_index = summarize_main_person(
                    pose_valid_indexes,
                    pose_boxes,
                    image_width,
                    image_height,
                )
                body_coverage = 0.0
                if pose_best_index is not None and getattr(pose_result, "keypoints", None) is not None and len(pose_result.keypoints) > pose_best_index:
                    confidences = getattr(pose_result.keypoints, "conf", None)
                    if confidences is not None:
                        body_coverage = compute_yolo_coverage_score(confidences[pose_best_index].tolist(), min_pose_keypoint_confidence)
                row["body_coverage_score_yolo_pose"] = round_or_blank(body_coverage)
                row["has_face_yunet"] = ""
                evaluate_yolo_gate(row)

                processed += 1
                if verbose and processed % 500 == 0:
                    print(f"processed {processed}/{len(rows)}", flush=True)

            time.sleep(0.05)

    return enriched


def chunked(rows: list[dict[str, object]], size: int) -> list[list[dict[str, object]]]:
    return [rows[index : index + size] for index in range(0, len(rows), size)]


def write_options_sheet(workbook: Workbook) -> None:
    sheet = workbook.create_sheet("Rejection Reason Options")
    sheet.sheet_state = "hidden"
    sheet.cell(1, 1).value = "rejection_reason"
    for index, reason in enumerate(REJECTION_REASONS, start=2):
        sheet.cell(index, 1).value = reason


def write_workbook(path: Path, rows: list[dict[str, object]]) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Review"
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(BASE_COLUMNS))}{len(rows) + 1}"
    write_options_sheet(workbook)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    sheet.append(BASE_COLUMNS)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    for index, original_row in enumerate(rows, start=2):
        row = dict(original_row)
        normalize_measurements(row)
        values = []
        for column in BASE_COLUMNS:
            if column == "image_preview":
                values.append(f'=IF(X{index}<>"",IMAGE(X{index}),"")')
            else:
                values.append(row.get(column, ""))
        sheet.append(values)
        sheet.row_dimensions[index].height = 92

    decision_validation = DataValidation(type="list", formula1=f'"{PRODUCTION_DECISIONS}"', allow_blank=True)
    sheet.add_data_validation(decision_validation)
    decision_validation.add(f"A2:A{max(2, len(rows) + 1)}")

    reason_validation = DataValidation(
        type="list",
        formula1=f"'Rejection Reason Options'!$A$2:$A${len(REJECTION_REASONS) + 1}",
        allow_blank=True,
    )
    sheet.add_data_validation(reason_validation)
    reason_validation.add(f"B2:B{max(2, len(rows) + 1)}")

    widths = {
        "A": 22,
        "B": 34,
        "C": 34,
        "D": 22,
        "K": 34,
        "L": 44,
        "M": 18,
        "N": 24,
        "O": 32,
        "X": 60,
        "Y": 60,
        "AA": 60,
        "AB": 60,
        "AD": 40,
        "AS": 55,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width
    for row_cells in sheet.iter_rows(min_row=2, max_row=len(rows) + 1):
        for cell in row_cells:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def write_package(output_dir: Path, buckets: dict[str, list[dict[str, object]]]) -> list[Path]:
    output_dir.mkdir(parents=True, exist_ok=True)
    for stale_file in output_dir.glob("supabase_image_review_*_part_*.xlsx"):
        stale_file.unlink()
    written = []
    for bucket, rows in buckets.items():
        for part, chunk in enumerate(chunked(rows, MAX_ROWS_PER_WORKBOOK), start=1):
            path = output_dir / f"supabase_image_review_{bucket}_part_{part:03d}.xlsx"
            write_workbook(path, chunk)
            written.append(path)
    return written


def write_csv_audit(output_dir: Path, rows: list[dict[str, object]]) -> None:
    path = output_dir / "cv_gate_audit_rows.csv"
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=BASE_COLUMNS + ["source_review_workbook", "source_review_workbook_row"])
        writer.writeheader()
        for row in rows:
            writer.writerow({column: row.get(column, "") for column in writer.fieldnames})


def write_readme(output_dir: Path, buckets: dict[str, list[dict[str, object]]], files: list[Path]) -> None:
    cv_counts = Counter(str(row.get("cv_reason_code") or "") for rows in buckets.values() for row in rows)
    lines = [
        "# Supabase Production Image Review Package - CV Gated",
        "",
        "This package starts from the S3-refresh approval candidates, then runs the YOLO detect and YOLO pose CV gate before assigning review buckets.",
        "",
        "## Buckets",
        "",
    ]
    for bucket, rows in buckets.items():
        lines.append(f"- `{bucket}`: {len(rows)} rows")
    lines.extend(["", "## CV Reason Codes", ""])
    for reason, count in cv_counts.most_common():
        lines.append(f"- `{reason or '<blank>'}`: {count}")
    lines.extend(["", "## Files", ""])
    for path in files:
        lines.append(f"- `{path.name}`")
    lines.append("")
    (output_dir / "README.md").write_text("\n".join(lines), encoding="utf-8")


def main() -> None:
    args = parse_args()
    rows = read_approval_rows(args.input_package, args.limit) if args.read_xlsx_package else load_approval_rows_from_builder(args.limit)
    bootstrap_vendor_paths(args.vendor_dir.resolve())
    np, requests, Image, YOLO = import_yolo_dependencies()
    detect_model = YOLO(str(args.yolo_detect_model.resolve()))
    pose_model = YOLO(str(args.yolo_pose_model.resolve()))
    session = requests.Session()
    ruled = enrich_rows_with_yolo_gate(
        rows=rows,
        detect_model=detect_model,
        pose_model=pose_model,
        np=np,
        session=session,
        Image=Image,
        timeout=args.timeout,
        min_person_confidence=0.35,
        min_pose_keypoint_confidence=0.35,
        yolo_batch_size=args.yolo_batch_size,
        download_workers=args.download_workers,
        verbose=args.verbose,
    )
    buckets = {
        "approve_candidates": [],
        "disapprove_candidates": [],
        "needs_human_review": [],
    }
    for row in ruled:
        buckets[route_row(row)].append(row)
    files = write_package(args.output_package, buckets)
    write_csv_audit(args.output_package, ruled)
    write_readme(args.output_package, buckets, files)
    print(args.output_package)
    for bucket, bucket_rows in buckets.items():
        print(f"{bucket}: {len(bucket_rows)}")
    print(f"workbooks: {len(files)}")


if __name__ == "__main__":
    main()
