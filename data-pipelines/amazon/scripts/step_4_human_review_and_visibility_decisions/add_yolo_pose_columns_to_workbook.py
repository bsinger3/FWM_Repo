#!/usr/bin/env python3
import argparse
import os
import shutil
import sys
from copy import copy
from io import BytesIO
from pathlib import Path


NEW_COLUMNS = [
    "person_count_yolo_pose",
    "main_person_height_pct_yolo_pose",
    "main_person_bbox_area_pct_yolo_pose",
    "body_coverage_score_yolo_pose",
]

POSE_KEYPOINT_WEIGHTS = {
    0: 1.0,   # nose
    5: 1.5,   # left shoulder
    6: 1.5,   # right shoulder
    11: 1.5,  # left hip
    12: 1.5,  # right hip
    13: 1.0,  # left knee
    14: 1.0,  # right knee
    15: 1.5,  # left ankle
    16: 1.5,  # right ankle
}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Add YOLO pose-derived subject-composition columns to a review "
            "workbook using the image in original_url_display."
        )
    )
    parser.add_argument(
        "--workbook",
        type=Path,
        required=True,
        help="Path to the .xlsx workbook to update in place.",
    )
    parser.add_argument(
        "--sheet",
        default="",
        help="Optional worksheet name. Defaults to the workbook's first sheet.",
    )
    parser.add_argument(
        "--model",
        default="yolov8n-pose.pt",
        help="Ultralytics pose model path or model name.",
    )
    parser.add_argument(
        "--url-column",
        default="original_url_display",
        help="Header name containing the image URL.",
    )
    parser.add_argument(
        "--anchor-column",
        default="full_lower_body_visible",
        help="Insert the new columns immediately after this header when possible.",
    )
    parser.add_argument(
        "--timeout",
        type=float,
        default=20.0,
        help="Per-image download timeout in seconds.",
    )
    parser.add_argument(
        "--min-person-confidence",
        type=float,
        default=0.35,
        help="Minimum YOLO confidence used when counting person detections.",
    )
    parser.add_argument(
        "--min-keypoint-confidence",
        type=float,
        default=0.35,
        help="Minimum keypoint confidence used in the body coverage score.",
    )
    parser.add_argument(
        "--limit",
        type=int,
        default=0,
        help="Optional row limit for testing. 0 means process every row.",
    )
    parser.add_argument(
        "--backup-suffix",
        default=".bak",
        help="Suffix for the workbook backup created before saving.",
    )
    parser.add_argument(
        "--verbose",
        action="store_true",
        help="Print progress and failures to stderr.",
    )
    return parser.parse_args()


def import_dependencies():
    try:
        import numpy as np  # type: ignore
        import requests  # type: ignore
        from openpyxl import load_workbook  # type: ignore
        from PIL import Image  # type: ignore
        from ultralytics import YOLO  # type: ignore
    except ModuleNotFoundError as exc:
        missing = exc.name or "required package"
        raise SystemExit(
            "Missing dependency: "
            f"{missing}. Install ultralytics, torch, requests, pillow, numpy, "
            "and openpyxl."
        ) from exc
    return np, requests, load_workbook, Image, YOLO


def find_header_map(worksheet) -> dict[str, int]:
    header_map: dict[str, int] = {}
    for index in range(1, worksheet.max_column + 1):
        value = worksheet.cell(row=1, column=index).value
        if isinstance(value, str) and value.strip():
            header_map[value.strip()] = index
    return header_map


def copy_column_style(worksheet, source_col: int, target_col: int, max_row: int) -> None:
    for row in range(1, max_row + 1):
        source = worksheet.cell(row=row, column=source_col)
        target = worksheet.cell(row=row, column=target_col)
        if source.has_style:
            target._style = copy(source._style)
            target.font = copy(source.font)
            target.fill = copy(source.fill)
            target.border = copy(source.border)
            target.alignment = copy(source.alignment)
            target.protection = copy(source.protection)
            target.number_format = source.number_format


def insert_columns(worksheet, headers: dict[str, int], anchor_column: str) -> dict[str, int]:
    if all(column in headers for column in NEW_COLUMNS):
        return headers
    insert_after = headers.get(anchor_column, worksheet.max_column)
    for offset, column_name in enumerate(NEW_COLUMNS, start=1):
        target_col = insert_after + offset
        worksheet.insert_cols(target_col)
        worksheet.cell(row=1, column=target_col, value=column_name)
        copy_column_style(worksheet, insert_after, target_col, worksheet.max_row)
        insert_after = target_col
    return find_header_map(worksheet)


def fetch_rgb_image(url: str, timeout: float, session, Image):
    response = session.get(
        url,
        timeout=timeout,
        headers={
            "User-Agent": (
                "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/123.0 Safari/537.36"
            )
        },
    )
    response.raise_for_status()
    return Image.open(BytesIO(response.content)).convert("RGB")


def round_or_blank(value):
    if value is None:
        return ""
    return round(float(value), 3)


def compute_body_coverage_score(keypoint_confidences, min_keypoint_confidence: float) -> float:
    total_weight = float(sum(POSE_KEYPOINT_WEIGHTS.values()))
    visible_weight = 0.0
    for index, weight in POSE_KEYPOINT_WEIGHTS.items():
        confidence = float(keypoint_confidences[index])
        if confidence >= min_keypoint_confidence:
            visible_weight += weight
    return round((visible_weight / total_weight) * 100.0, 1)


def analyze_image(image_rgb, model, np, min_person_confidence: float, min_keypoint_confidence: float) -> dict[str, object]:
    result = model.predict(np.array(image_rgb), verbose=False, device="cpu")[0]
    image_height, image_width = image_rgb.size[1], image_rgb.size[0]
    valid_indexes: list[int] = []

    if result.boxes is not None and len(result.boxes):
        for index, (cls_value, confidence) in enumerate(zip(result.boxes.cls.tolist(), result.boxes.conf.tolist())):
            if int(cls_value) == 0 and float(confidence) >= min_person_confidence:
                valid_indexes.append(index)

    if not valid_indexes:
        return {
            "person_count_yolo_pose": 0,
            "main_person_height_pct_yolo_pose": 0.0,
            "main_person_bbox_area_pct_yolo_pose": 0.0,
            "body_coverage_score_yolo_pose": 0.0,
        }

    boxes = result.boxes.xyxy.tolist()
    best_index = max(
        valid_indexes,
        key=lambda idx: max(0.0, float(boxes[idx][2] - boxes[idx][0])) * max(0.0, float(boxes[idx][3] - boxes[idx][1])),
    )
    x1, y1, x2, y2 = [float(value) for value in boxes[best_index]]
    bbox_width = max(0.0, x2 - x1)
    bbox_height = max(0.0, y2 - y1)
    height_pct = bbox_height / float(image_height) if image_height else 0.0
    area_pct = (bbox_width * bbox_height) / float(image_width * image_height) if image_width and image_height else 0.0

    body_coverage_score = 0.0
    if getattr(result, "keypoints", None) is not None and len(result.keypoints) > best_index:
        confidences = getattr(result.keypoints, "conf", None)
        if confidences is not None:
            body_coverage_score = compute_body_coverage_score(
                confidences[best_index].tolist(),
                min_keypoint_confidence=min_keypoint_confidence,
            )

    return {
        "person_count_yolo_pose": len(valid_indexes),
        "main_person_height_pct_yolo_pose": round_or_blank(height_pct),
        "main_person_bbox_area_pct_yolo_pose": round_or_blank(area_pct),
        "body_coverage_score_yolo_pose": round_or_blank(body_coverage_score),
    }


def main() -> None:
    args = parse_args()
    os.environ.setdefault("YOLO_CONFIG_DIR", "/tmp/Ultralytics")
    os.environ.setdefault("MPLCONFIGDIR", "/tmp/matplotlib-codex")
    np, requests, load_workbook, Image, YOLO = import_dependencies()

    workbook_path = args.workbook.resolve()
    if not workbook_path.exists():
        raise SystemExit(f"Workbook not found: {workbook_path}")

    backup_path = workbook_path.with_name(workbook_path.name + args.backup_suffix)
    shutil.copy2(workbook_path, backup_path)
    if args.verbose:
        print(f"Backup created: {backup_path}", file=sys.stderr)

    workbook = load_workbook(workbook_path)
    worksheet = workbook[args.sheet] if args.sheet else workbook[workbook.sheetnames[0]]
    headers = find_header_map(worksheet)
    if args.url_column not in headers:
        raise SystemExit(f"Required URL column not found: {args.url_column}")
    headers = insert_columns(worksheet, headers, args.anchor_column)
    url_col = headers[args.url_column]

    model = YOLO(args.model)
    session = requests.Session()
    processed = 0

    for row in range(2, worksheet.max_row + 1):
        url_value = worksheet.cell(row=row, column=url_col).value
        url = str(url_value or "").strip()
        if not url:
            for column_name in NEW_COLUMNS:
                worksheet.cell(row=row, column=headers[column_name], value="")
            continue

        if args.limit > 0 and processed >= args.limit:
            break

        try:
            image_rgb = fetch_rgb_image(url, args.timeout, session, Image)
            enriched = analyze_image(
                image_rgb=image_rgb,
                model=model,
                np=np,
                min_person_confidence=args.min_person_confidence,
                min_keypoint_confidence=args.min_keypoint_confidence,
            )
            for column_name, value in enriched.items():
                worksheet.cell(row=row, column=headers[column_name], value=value)
        except Exception as exc:  # noqa: BLE001
            for column_name in NEW_COLUMNS:
                worksheet.cell(row=row, column=headers[column_name], value="")
            if args.verbose:
                print(f"[row {row}] failed for {url}: {exc}", file=sys.stderr)

        processed += 1
        if args.verbose and processed % 25 == 0:
            print(f"Processed {processed} images", file=sys.stderr)

    workbook.save(workbook_path)
    if args.verbose:
        print(f"Saved workbook: {workbook_path}", file=sys.stderr)


if __name__ == "__main__":
    main()
