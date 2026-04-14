#!/usr/bin/env python3
import argparse
import shutil
import sys
from copy import copy
from io import BytesIO
from pathlib import Path


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Add an SCRFD-based face-presence column to a review workbook by "
            "running InsightFace against original_url_display."
        )
    )
    parser.add_argument(
        "--workbook",
        type=Path,
        required=True,
        help="Path to the .xlsx workbook to update in place.",
    )
    parser.add_argument(
        "--sheet",
        default="",
        help="Optional worksheet name. Defaults to the workbook's first sheet.",
    )
    parser.add_argument(
        "--url-column",
        default="original_url_display",
        help="Header name containing the image URL.",
    )
    parser.add_argument(
        "--anchor-column",
        default="has_face_yunet",
        help="Insert the new column immediately after this header when possible.",
    )
    parser.add_argument(
        "--output-column",
        default="has_face_scrfd",
        help="Header name for the new face-detection column.",
    )
    parser.add_argument(
        "--timeout",
        type=float,
        default=20.0,
        help="Per-image download timeout in seconds.",
    )
    parser.add_argument(
        "--backup-suffix",
        default=".bak",
        help="Suffix for the workbook backup created before saving.",
    )
    parser.add_argument(
        "--det-size",
        type=int,
        default=640,
        help="Detection input size used by InsightFace SCRFD.",
    )
    parser.add_argument(
        "--min-score",
        type=float,
        default=0.6,
        help="Minimum accepted SCRFD detection score.",
    )
    parser.add_argument(
        "--model-root",
        type=Path,
        default=Path(".codex_vendor/insightface_models"),
        help="Directory used to store InsightFace model downloads.",
    )
    parser.add_argument(
        "--verbose",
        action="store_true",
        help="Print progress and failures to stderr.",
    )
    return parser.parse_args()


def import_dependencies():
    try:
        import numpy as np  # type: ignore
        import requests  # type: ignore
        from insightface.app import FaceAnalysis  # type: ignore
        from openpyxl import load_workbook  # type: ignore
        from PIL import Image  # type: ignore
    except ModuleNotFoundError as exc:
        missing = exc.name or "required package"
        raise SystemExit(
            "Missing dependency: "
            f"{missing}. Install insightface, onnxruntime, numpy, requests, "
            "pillow, and openpyxl."
        ) from exc
    return np, requests, FaceAnalysis, load_workbook, Image


def bool_to_str(value: bool) -> str:
    return "true" if value else "false"


def find_header_map(worksheet) -> dict[str, int]:
    header_map: dict[str, int] = {}
    for index in range(1, worksheet.max_column + 1):
        value = worksheet.cell(row=1, column=index).value
        if isinstance(value, str) and value.strip():
            header_map[value.strip()] = index
    return header_map


def copy_column_style(worksheet, source_col: int, target_col: int, max_row: int) -> None:
    for row in range(1, max_row + 1):
        source = worksheet.cell(row=row, column=source_col)
        target = worksheet.cell(row=row, column=target_col)
        if source.has_style:
            target._style = copy(source._style)
            target.font = copy(source.font)
            target.fill = copy(source.fill)
            target.border = copy(source.border)
            target.alignment = copy(source.alignment)
            target.protection = copy(source.protection)
            target.number_format = source.number_format


def fetch_bgr_image(url: str, timeout: float, session, Image, np):
    response = session.get(
        url,
        timeout=timeout,
        headers={
            "User-Agent": (
                "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/123.0 Safari/537.36"
            )
        },
    )
    response.raise_for_status()
    image = Image.open(BytesIO(response.content)).convert("RGB")
    rgb = np.array(image)
    return rgb[:, :, ::-1]


def detect_face_scrfd(image_bgr, detector, min_score: float) -> bool:
    faces = detector.get(image_bgr)
    if not faces:
        return False
    return any(float(getattr(face, "det_score", 0.0)) >= min_score for face in faces)


def main() -> None:
    args = parse_args()
    np, requests, FaceAnalysis, load_workbook, Image = import_dependencies()

    workbook_path = args.workbook.resolve()
    if not workbook_path.exists():
        raise SystemExit(f"Workbook not found: {workbook_path}")

    backup_path = workbook_path.with_name(workbook_path.name + args.backup_suffix)
    shutil.copy2(workbook_path, backup_path)
    if args.verbose:
        print(f"Backup created: {backup_path}", file=sys.stderr)

    workbook = load_workbook(workbook_path)
    worksheet = workbook[args.sheet] if args.sheet else workbook[workbook.sheetnames[0]]
    headers = find_header_map(worksheet)

    if args.url_column not in headers:
        raise SystemExit(f"Required URL column not found: {args.url_column}")

    if args.output_column in headers:
        output_col = headers[args.output_column]
    else:
        insert_after = headers.get(args.anchor_column, worksheet.max_column)
        worksheet.insert_cols(insert_after + 1)
        output_col = insert_after + 1
        worksheet.cell(row=1, column=output_col, value=args.output_column)
        copy_column_style(worksheet, insert_after, output_col, worksheet.max_row)
        headers = find_header_map(worksheet)

    url_col = headers[args.url_column]
    model_root = args.model_root.resolve()
    model_root.mkdir(parents=True, exist_ok=True)

    detector = FaceAnalysis(
        name="buffalo_l",
        root=str(model_root),
        providers=["CPUExecutionProvider"],
    )
    detector.prepare(ctx_id=-1, det_size=(args.det_size, args.det_size))

    session = requests.Session()
    processed = 0

    for row in range(2, worksheet.max_row + 1):
        url_value = worksheet.cell(row=row, column=url_col).value
        url = str(url_value or "").strip()
        if not url:
            worksheet.cell(row=row, column=output_col, value="")
            continue
        try:
            image_bgr = fetch_bgr_image(url, args.timeout, session, Image, np)
            detected = detect_face_scrfd(image_bgr, detector, args.min_score)
            worksheet.cell(row=row, column=output_col, value=bool_to_str(detected))
        except Exception as exc:  # noqa: BLE001
            worksheet.cell(row=row, column=output_col, value="")
            if args.verbose:
                print(f"[row {row}] failed for {url}: {exc}", file=sys.stderr)
        processed += 1
        if args.verbose and processed % 25 == 0:
            print(f"Processed {processed} images", file=sys.stderr)

    workbook.save(workbook_path)
    if args.verbose:
        print(
            f"Saved workbook with column '{args.output_column}': {workbook_path}",
            file=sys.stderr,
        )


if __name__ == "__main__":
    main()
