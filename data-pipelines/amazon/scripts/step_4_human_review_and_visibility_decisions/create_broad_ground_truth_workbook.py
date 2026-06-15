#!/usr/bin/env python3
"""Create a Google-Sheets-friendly XLSX workbook for broad ground-truth labeling."""

from __future__ import annotations
import sys

import csv
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation


REPO_ROOT = Path(__file__).resolve().parents[4]
PIPELINE_SCRIPTS_DIR = REPO_ROOT / "data-pipelines" / "scripts"
if str(PIPELINE_SCRIPTS_DIR) not in sys.path:
    sys.path.insert(0, str(PIPELINE_SCRIPTS_DIR))

from pipeline_paths import archive_root, cv_annotated_pending_human_review_root  # noqa: E402

LEGACY_OUTPUTS_ARCHIVE = archive_root() / "old_outputs" / "repo_outputs_archive" / "supabase_output_cleanup_2026_05_29"
CV_EXPERIMENTS_DIR = LEGACY_OUTPUTS_ARCHIVE / "cv_experiments"

OUT_DIR = CV_EXPERIMENTS_DIR / "ground_truth_labeling_broad"
SOURCE_CSV = OUT_DIR / "combined_amazon_nonamazon_llm_seeded_ground_truth_queue.csv"
OUTPUT_XLSX = OUT_DIR / "combined_amazon_nonamazon_llm_seeded_ground_truth_queue.xlsx"

REASON_CODES = [
    ("LOW_RESOLUTION", "Image is too small, pixelated, or heavily compressed."),
    ("TOO_DARK", "Image is too dark to see garment fit, silhouette, or details."),
    ("TOO_BRIGHT_OR_WASHED_OUT", "Image is overexposed or washed out enough to hurt fit evaluation."),
    ("BLURRY_OR_MOTION_BLUR", "Image is blurred enough to hurt fit evaluation."),
    ("GRAINY_OR_NOISY", "Image has heavy noise or grain enough to hurt fit evaluation."),
    ("GLARE_OR_REFLECTION", "Mirror glare, flash, lens flare, or reflection hides the garment."),
    ("GARMENT_CUT_OFF", "Relevant worn clothing or fit view is materially cut off by the image boundary."),
    ("GARMENT_TOP_COVERED", "Waistband/top of jeans, pants, or relevant garment is hidden by a shirt, hand, phone, counter, or other object."),
    ("GARMENT_BOTTOM_CUT_OFF", "Hem, ankle, or bottom of pants/jeans is missing, so length and silhouette cannot be judged."),
    ("GARMENT_TOO_PARTIAL", "Only a small detail or narrow garment section is visible, not enough to judge fit."),
    ("GARMENT_OBSCURED", "The garment is blocked by hands, phone, bag, outerwear, hair, furniture, stickers, mirror frame, or other objects."),
    ("NOT_WORN_BY_PERSON", "Clothing is flat-lay, on the floor/bed, hanging, packaging/tag only, or otherwise not shown on a human body."),
    ("TARGET_WEARER_AMBIGUOUS", "Multiple people are visible and it is unclear whose garment should be evaluated."),
    ("PERSON_TOO_FAR", "Person or garment is too small/distant for fit evaluation."),
    ("PERSON_TOO_CLOSE", "The crop is so tight that the shopper cannot judge overall silhouette or fit."),
    ("BAD_ANGLE_TOP_DOWN", "Top-down selfie angle distorts fit, often showing mostly feet/legs from above."),
    ("BAD_ANGLE_SIDE_OR_TWISTED", "Side, twisted, seated, or unusual pose/angle makes the fit hard to understand."),
    ("MIRROR_OR_CAMERA_BLOCKS_VIEW", "Phone, mirror frame, sink/counter, or reflection blocks meaningful garment area."),
    ("ROTATED_OR_WRONG_ORIENTATION", "Image is sideways, upside down, or otherwise needs rotation."),
    ("BAD_ASPECT_RATIO_OR_BARS", "Black bars, letterboxing, screenshot borders, or odd aspect ratio makes image look inconsistent or less usable."),
    ("NEEDS_CROP", "Useful garment exists, but current framing includes excessive distracting or non-useful area."),
    ("BACKGROUND_TOO_CLUTTERED", "Room/object clutter distracts from the clothing and makes the image less shoppable."),
    ("BACKGROUND_VISUALLY_OFFPUTTING", "The setting is likely to deter shoppers even if the garment is visible."),
    ("BATHROOM_OR_PUBLIC_RESTROOM_DISTRACTION", "Bathroom/restroom fixtures dominate the image or make it feel less shoppable."),
    ("MESSY_MIRROR_OR_DIRTY_SURFACE", "Mirror grime, smears, dust, or dirty surface blocks or degrades the image."),
    ("DISTRACTING_OBJECTS", "Large objects compete with the garment or draw attention away from the shopper."),
    ("CATALOG_OR_STOCK_IMAGE", "Professional product/model photo rather than shopper review photo."),
    ("BACKGROUND_REMOVED_OR_ALTERED", "Background has been removed, replaced, or turned white in a way inconsistent with review photos."),
    ("FILTER_OR_EFFECT", "Heavy filter changes color/detail or makes image non-representative."),
    ("STICKERS_OR_DRAWINGS_OVER_IMAGE", "Emojis, drawn circles, arrows, labels, or markup distract from or obscure the garment."),
    ("TEXT_OVERLAY_OR_SCREENSHOT", "Text, screenshot UI, or comparison labels dominate the image."),
    ("COLLAGE_OR_MULTIPANEL", "Multiple images are combined into one, making it unclear which view matters."),
    ("NO_PERSON_VISIBLE", "No human body wearing the garment is visible."),
    ("POSSIBLE_PRODUCT_CONTEXT_MISMATCH", "Image may not match the reviewed product/category, but product context can be missing or wrong; use for review, not automatic rejection."),
    ("UNSUPPORTED_SHOPPER_OR_PRODUCT_TYPE", "Image/product appears outside current site scope or policy, if this remains a business rule."),
    ("DUPLICATE_OR_NEAR_DUPLICATE", "Same or near-identical image already accepted or queued."),
    ("SOLD_OUT_OR_BAD_PRODUCT_LINK", "Product page is unusable, sold out, or bad link; this affects publishing but is not image CV."),
]

REASON_CODE_NAMES = [code for code, _ in REASON_CODES]


def read_rows() -> list[dict[str, str]]:
    with SOURCE_CSV.open(newline="", encoding="utf-8-sig") as handle:
        return list(csv.DictReader(handle))


def style_header(sheet, row: int = 1) -> None:
    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)
    for cell in sheet[row]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(wrap_text=True, vertical="top")


def add_instructions(workbook: Workbook) -> None:
    sheet = workbook.active
    sheet.title = "Instructions"
    rows = [
        ("Purpose", "Label likely rejection reasons for each candidate image. LLM suggestions are only hints; your human label is the ground truth."),
        ("How to label", "Set final_human_decision to APPROVED, REJECTED, or UNSURE. If rejected, fill primary_reason_code and optionally secondary_reason_code."),
        ("Image preview", "Google Sheets should render the image_preview formula. If it does not, open the original_url_display link."),
        ("Reason columns", "The reason-code columns at the far right can be marked with TRUE/x if multiple reasons apply."),
        ("Important", "Do not reject only because product/category metadata seems wrong. Judge whether the photo is useful for fit shopping."),
    ]
    sheet.append(["Field", "Instruction"])
    for row in rows:
        sheet.append(row)
    style_header(sheet)
    sheet.column_dimensions["A"].width = 24
    sheet.column_dimensions["B"].width = 110
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")


def add_reason_codes(workbook: Workbook) -> None:
    sheet = workbook.create_sheet("Reason Codes")
    sheet.append(["reason_code", "description"])
    for code, description in REASON_CODES:
        sheet.append([code, description])
    style_header(sheet)
    sheet.column_dimensions["A"].width = 34
    sheet.column_dimensions["B"].width = 100
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")


def add_queue(workbook: Workbook, rows: list[dict[str, str]]) -> None:
    sheet = workbook.create_sheet("Labeling Queue")
    if not rows:
        return
    source_headers = list(rows[0].keys())
    headers = [header for header in source_headers if header not in REASON_CODE_NAMES]
    insert_after = "cv_reason_code" if "cv_reason_code" in headers else "labeler_notes"
    if insert_after in headers:
        insert_at = headers.index(insert_after) + 1
        headers[insert_at:insert_at] = REASON_CODE_NAMES
    else:
        headers.extend(REASON_CODE_NAMES)
    sheet.append(headers)
    for row in rows:
        values = [row.get(header, "") for header in headers]
        sheet.append(values)
    style_header(sheet)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

    for cell in sheet[1]:
        header = str(cell.value)
        if header in {"image_preview"}:
            sheet.column_dimensions[cell.column_letter].width = 28
        elif header in {"original_url_display", "product_page_url_display"}:
            sheet.column_dimensions[cell.column_letter].width = 42
        elif header in {"llm_summary", "user_comment", "product_title_raw"}:
            sheet.column_dimensions[cell.column_letter].width = 54
        elif header in {"final_human_decision", "primary_reason_code", "secondary_reason_code", "labeler_notes"}:
            sheet.column_dimensions[cell.column_letter].width = 24
        else:
            sheet.column_dimensions[cell.column_letter].width = 20

    for row in sheet.iter_rows(min_row=2):
        sheet.row_dimensions[row[0].row].height = 54
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    decision_validation = DataValidation(type="list", formula1='"APPROVED,REJECTED,UNSURE"', allow_blank=True)
    reason_validation = DataValidation(
        type="list",
        formula1=f"'Reason Codes'!$A$2:$A${len(REASON_CODES) + 1}",
        allow_blank=True,
    )
    sheet.add_data_validation(decision_validation)
    sheet.add_data_validation(reason_validation)
    for idx, header in enumerate(headers, start=1):
        letter = sheet.cell(row=1, column=idx).column_letter
        if header == "final_human_decision":
            decision_validation.add(f"{letter}2:{letter}{sheet.max_row}")
        if header in {"primary_reason_code", "secondary_reason_code"}:
            reason_validation.add(f"{letter}2:{letter}{sheet.max_row}")


def main() -> None:
    rows = read_rows()
    workbook = Workbook()
    add_instructions(workbook)
    add_queue(workbook, rows)
    add_reason_codes(workbook)
    workbook.save(OUTPUT_XLSX)
    print(OUTPUT_XLSX)


if __name__ == "__main__":
    main()
