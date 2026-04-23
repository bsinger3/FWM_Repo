#!/usr/bin/env python3
import argparse
import os
import shutil
import sys
from copy import copy
from io import BytesIO
from pathlib import Path
import urllib.request


PIPELINE_ROOT = Path(__file__).resolve().parents[2]
DEFAULT_VENDOR_DIR = PIPELINE_ROOT.parents[1] / ".codex_vendor"
DEFAULT_MODEL_ROOT = PIPELINE_ROOT.parents[1] / ".codex_vendor" / "insightface_models"


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Add a RetinaFace-based face-presence column to a review workbook by "
            "running InsightFace against original_url_display."
        )
    )
    parser.add_argument(
        "--workbook",
        type=Path,
        required=True,
        help="Path to the .xlsx workbook to update in place.",
    )
    parser.add_argument(
        "--sheet",
        default="",
        help="Optional worksheet name. Defaults to the workbook's first sheet.",
    )
    parser.add_argument(
        "--url-column",
        default="original_url_display",
        help="Header name containing the image URL.",
    )
    parser.add_argument(
        "--anchor-column",
        default="has_face_scrfd",
        help="Insert the new column immediately after this header when possible.",
    )
    parser.add_argument(
        "--fallback-anchor-column",
        default="has_face_yunet",
        help="Secondary anchor used when the primary anchor is absent.",
    )
    parser.add_argument(
        "--output-column",
        default="has_face_retinaface",
        help="Header name for the new face-detection column.",
    )
    parser.add_argument(
        "--timeout",
        type=float,
        default=20.0,
        help="Per-image download timeout in seconds.",
    )
    parser.add_argument(
        "--backup-suffix",
        default=".bak",
        help="Suffix for the workbook backup created before saving.",
    )
    parser.add_argument(
        "--input-size",
        type=int,
        default=640,
        help="Detection input size used by RetinaFace.",
    )
    parser.add_argument(
        "--min-score",
        type=float,
        default=0.6,
        help="Minimum accepted RetinaFace detection score.",
    )
    parser.add_argument(
        "--model-file",
        default="retinaface_mnet025_v2.onnx",
        help="RetinaFace ONNX filename to cache under the model root.",
    )
    parser.add_argument(
        "--model-path",
        type=Path,
        help="Optional existing RetinaFace ONNX file to use directly.",
    )
    parser.add_argument(
        "--download-url",
        default="",
        help="Optional direct URL to a RetinaFace ONNX file when official auto-download is unavailable.",
    )
    parser.add_argument(
        "--model-root",
        type=Path,
        default=DEFAULT_MODEL_ROOT,
        help="Directory used to store InsightFace model downloads.",
    )
    parser.add_argument(
        "--vendor-dir",
        type=Path,
        default=DEFAULT_VENDOR_DIR,
        help="Directory containing the vendored scrfd_test runtime.",
    )
    parser.add_argument(
        "--limit",
        type=int,
        default=0,
        help="Optional row limit for smoke tests. 0 means process every row.",
    )
    parser.add_argument(
        "--verbose",
        action="store_true",
        help="Print progress and failures to stderr.",
    )
    return parser.parse_args()


def bootstrap_vendor_paths(vendor_dir: Path) -> None:
    runtime_dir = vendor_dir / "scrfd_test"
    if runtime_dir.exists():
        sys.path.insert(0, str(runtime_dir))


def import_dependencies():
    os.environ.setdefault("MPLCONFIGDIR", "/tmp/matplotlib-codex")
    try:
        import numpy as np  # type: ignore
        import requests  # type: ignore
        from insightface.model_zoo.retinaface import RetinaFace  # type: ignore
        from insightface.utils.storage import download_onnx  # type: ignore
        from openpyxl import load_workbook  # type: ignore
        from PIL import Image  # type: ignore
    except ModuleNotFoundError as exc:
        missing = exc.name or "required package"
        raise SystemExit(
            "Missing dependency: "
            f"{missing}. This script expects the vendored scrfd_test runtime under .codex_vendor."
        ) from exc
    return np, requests, RetinaFace, download_onnx, load_workbook, Image


def bool_to_str(value: bool) -> str:
    return "true" if value else "false"


def find_header_map(worksheet) -> dict[str, int]:
    header_map: dict[str, int] = {}
    for index in range(1, worksheet.max_column + 1):
        value = worksheet.cell(row=1, column=index).value
        if isinstance(value, str) and value.strip():
            header_map[value.strip()] = index
    return header_map


def copy_column_style(worksheet, source_col: int, target_col: int, max_row: int) -> None:
    for row in range(1, max_row + 1):
        source = worksheet.cell(row=row, column=source_col)
        target = worksheet.cell(row=row, column=target_col)
        if source.has_style:
            target._style = copy(source._style)
            target.font = copy(source.font)
            target.fill = copy(source.fill)
            target.border = copy(source.border)
            target.alignment = copy(source.alignment)
            target.protection = copy(source.protection)
            target.number_format = source.number_format


def ensure_output_column(
    worksheet,
    headers: dict[str, int],
    output_column: str,
    anchor_column: str,
    fallback_anchor_column: str,
) -> tuple[dict[str, int], int]:
    if output_column in headers:
        return headers, headers[output_column]

    insert_after = headers.get(anchor_column) or headers.get(fallback_anchor_column) or worksheet.max_column
    worksheet.insert_cols(insert_after + 1)
    output_col = insert_after + 1
    worksheet.cell(row=1, column=output_col, value=output_column)
    copy_column_style(worksheet, insert_after, output_col, worksheet.max_row)
    headers = find_header_map(worksheet)
    return headers, headers[output_column]


def fetch_bgr_image(url: str, timeout: float, session, Image, np):
    response = session.get(
        url,
        timeout=timeout,
        headers={
            "User-Agent": (
                "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/123.0 Safari/537.36"
            )
        },
    )
    response.raise_for_status()
    image = Image.open(BytesIO(response.content)).convert("RGB")
    rgb = np.array(image)
    return rgb[:, :, ::-1]


def build_detector(model_root: Path, model_file: str, RetinaFace, download_onnx, input_size: int, min_score: float):
    model_root = model_root.resolve()
    model_root.mkdir(parents=True, exist_ok=True)
    model_path = download_onnx("models", model_file, root=str(model_root.parent))
    detector = RetinaFace(model_file=model_path)
    detector.prepare(
        ctx_id=-1,
        det_thresh=min_score,
        input_size=(input_size, input_size),
    )
    return detector


def detect_face_retinaface(image_bgr, detector, input_size: int) -> bool:
    detections, _kps = detector.detect(image_bgr, input_size=(input_size, input_size))
    return detections is not None and len(detections) > 0


def resolve_model_path(args, download_onnx) -> str:
    if args.model_path:
        model_path = args.model_path.resolve()
        if not model_path.exists():
            raise SystemExit(f"RetinaFace model file not found: {model_path}")
        return str(model_path)

    model_root = args.model_root.resolve()
    model_root.mkdir(parents=True, exist_ok=True)
    target_path = model_root / "models" / args.model_file
    if target_path.exists():
        return str(target_path)

    if args.download_url:
        target_path.parent.mkdir(parents=True, exist_ok=True)
        urllib.request.urlretrieve(args.download_url, str(target_path))
        return str(target_path)

    try:
        return download_onnx("models", args.model_file, root=str(model_root.parent))
    except RuntimeError as exc:
        raise SystemExit(
            "RetinaFace model download failed. The current official InsightFace v0.7 release "
            "does not appear to expose this RetinaFace ONNX asset. "
            "Provide --model-path to a local RetinaFace ONNX file or --download-url to a direct model URL."
        ) from exc


def main() -> None:
    args = parse_args()
    bootstrap_vendor_paths(args.vendor_dir.resolve())
    np, requests, RetinaFace, download_onnx, load_workbook, Image = import_dependencies()

    workbook_path = args.workbook.resolve()
    if not workbook_path.exists():
        raise SystemExit(f"Workbook not found: {workbook_path}")

    model_path = resolve_model_path(args, download_onnx)

    backup_path = workbook_path.with_name(workbook_path.name + args.backup_suffix)
    shutil.copy2(workbook_path, backup_path)
    if args.verbose:
        print(f"Backup created: {backup_path}", file=sys.stderr)

    workbook = load_workbook(workbook_path)
    worksheet = workbook[args.sheet] if args.sheet else workbook[workbook.sheetnames[0]]
    headers = find_header_map(worksheet)

    if args.url_column not in headers:
        raise SystemExit(f"Required URL column not found: {args.url_column}")

    headers, output_col = ensure_output_column(
        worksheet,
        headers,
        args.output_column,
        args.anchor_column,
        args.fallback_anchor_column,
    )
    url_col = headers[args.url_column]

    detector = build_detector(
        model_root=args.model_root,
        model_file=Path(model_path).name,
        RetinaFace=RetinaFace,
        download_onnx=lambda *a, **k: model_path,
        input_size=args.input_size,
        min_score=args.min_score,
    )

    session = requests.Session()
    processed = 0

    for row in range(2, worksheet.max_row + 1):
        url_value = worksheet.cell(row=row, column=url_col).value
        url = str(url_value or "").strip()
        if not url:
            worksheet.cell(row=row, column=output_col, value="")
            continue

        if args.limit > 0 and processed >= args.limit:
            break

        try:
            image_bgr = fetch_bgr_image(url, args.timeout, session, Image, np)
            detected = detect_face_retinaface(image_bgr, detector, args.input_size)
            worksheet.cell(row=row, column=output_col, value=bool_to_str(detected))
        except Exception as exc:  # noqa: BLE001
            worksheet.cell(row=row, column=output_col, value="")
            if args.verbose:
                print(f"[row {row}] failed for {url}: {exc}", file=sys.stderr)

        processed += 1
        if args.verbose and processed % 25 == 0:
            print(f"Processed {processed} images", file=sys.stderr)

    workbook.save(workbook_path)
    if args.verbose:
        print(
            f"Saved workbook with column '{args.output_column}': {workbook_path}",
            file=sys.stderr,
        )


if __name__ == "__main__":
    main()
