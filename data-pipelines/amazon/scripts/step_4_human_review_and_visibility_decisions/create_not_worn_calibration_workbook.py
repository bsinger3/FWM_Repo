#!/usr/bin/env python3
"""Create a one-question NOT_WORN_BY_PERSON calibration workbook."""

from __future__ import annotations
import sys

import csv
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation


REPO_ROOT = Path(__file__).resolve().parents[4]
PIPELINE_SCRIPTS_DIR = REPO_ROOT / "data-pipelines" / "scripts"
if str(PIPELINE_SCRIPTS_DIR) not in sys.path:
    sys.path.insert(0, str(PIPELINE_SCRIPTS_DIR))

from pipeline_paths import archive_root, cv_annotated_pending_human_review_root  # noqa: E402

LEGACY_OUTPUTS_ARCHIVE = archive_root() / "old_outputs" / "repo_outputs_archive" / "supabase_output_cleanup_2026_05_29"
CV_EXPERIMENTS_DIR = LEGACY_OUTPUTS_ARCHIVE / "cv_experiments"

EXP_DIR = CV_EXPERIMENTS_DIR / "yolo_segmentation_crop_reasons_broad_2026_05_25"
SOURCE_CSV = EXP_DIR / "yolo_segmentation_crop_reason_rows.csv"
OUTPUT_XLSX = EXP_DIR / "not_worn_by_person_yes_no_review_queue.xlsx"

ANSWER_CHOICES = ["YES", "NO", "UNSURE"]

HEADERS = [
    "not_worn_by_person_yes_no",
    "image_preview",
    "question",
    "original_url_display",
    "product_page_url_display",
    "source_site_display",
    "current_primary_reason_code",
    "current_secondary_reason_code",
    "current_labeler_notes",
    "final_human_decision",
    "model_predicted_not_worn",
    "model_rule",
    "seg_person_count",
    "seg_person_confidence",
    "seg_mask_area_pct",
    "llm_summary",
    "review_row_key",
]


def read_rows() -> list[dict[str, str]]:
    with SOURCE_CSV.open(newline="", encoding="utf-8-sig") as handle:
        return list(csv.DictReader(handle))


def truthy(value: object) -> bool:
    return str(value or "").strip().lower() in {"true", "yes", "1", "x", "checked"}


def to_float(value: object, default: float = 0.0) -> float:
    try:
        if value is None or str(value).strip() == "":
            return default
        return float(value)
    except ValueError:
        return default


def model_predicted_not_worn(row: dict[str, str]) -> bool:
    return to_float(row.get("seg_person_count"), default=999.0) == 0


def current_truth_not_worn(row: dict[str, str]) -> bool:
    return truthy(row.get("truth_NOT_WORN_BY_PERSON"))


def decision(row: dict[str, str]) -> str:
    return str(row.get("final_human_decision_norm") or row.get("final_human_decision") or "").upper()


def build_candidates(rows: list[dict[str, str]]) -> list[dict[str, str]]:
    positives = [
        row
        for row in rows
        if current_truth_not_worn(row) or model_predicted_not_worn(row)
    ]
    approved_controls = [
        row
        for row in rows
        if decision(row) == "APPROVED"
        and not current_truth_not_worn(row)
        and not model_predicted_not_worn(row)
    ][:20]
    rejected_controls = [
        row
        for row in rows
        if decision(row) == "REJECTED"
        and not current_truth_not_worn(row)
        and not model_predicted_not_worn(row)
    ][:25]

    seen: set[str] = set()
    candidates: list[dict[str, str]] = []
    for row in positives + approved_controls + rejected_controls:
        key = row.get("review_row_key") or row.get("original_url_display") or str(len(candidates))
        if key in seen:
            continue
        seen.add(key)
        candidates.append(row)

    return sorted(
        candidates,
        key=lambda row: (
            not (current_truth_not_worn(row) or model_predicted_not_worn(row)),
            decision(row) != "APPROVED",
            to_float(row.get("seg_person_count"), default=999.0),
            row.get("review_row_key", ""),
        ),
    )


def value_for(row: dict[str, str], header: str) -> object:
    if header == "not_worn_by_person_yes_no":
        return ""
    if header == "question":
        return "Is the clothing NOT being worn by a visible person? YES = flat-lay, hanger, packaging, product-only, or no visible wearer."
    if header == "current_primary_reason_code":
        return row.get("primary_reason_code", "")
    if header == "current_secondary_reason_code":
        return row.get("secondary_reason_code", "")
    if header == "current_labeler_notes":
        return row.get("labeler_notes", "")
    if header == "final_human_decision":
        return decision(row)
    if header == "model_predicted_not_worn":
        return model_predicted_not_worn(row)
    if header == "model_rule":
        return "YES if seg_person_count == 0"
    return row.get(header, "")


def style_header(sheet) -> None:
    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)
    for cell in sheet[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(wrap_text=True, vertical="top")


def main() -> None:
    candidates = build_candidates(read_rows())
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Not Worn Review"
    sheet.append(HEADERS)
    for row in candidates:
        sheet.append([value_for(row, header) for header in HEADERS])

    style_header(sheet)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:Q{sheet.max_row}"

    validation = DataValidation(type="list", formula1=f'"{",".join(ANSWER_CHOICES)}"', allow_blank=True)
    sheet.add_data_validation(validation)
    validation.add(f"A2:A{sheet.max_row}")

    widths = {
        "A": 24,
        "B": 26,
        "C": 76,
        "D": 42,
        "E": 42,
        "F": 22,
        "G": 24,
        "H": 24,
        "I": 42,
        "J": 20,
        "K": 24,
        "L": 28,
        "M": 16,
        "N": 18,
        "O": 16,
        "P": 64,
        "Q": 28,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width

    for row in sheet.iter_rows(min_row=2):
        sheet.row_dimensions[row[0].row].height = 72
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    workbook.save(OUTPUT_XLSX)
    print(OUTPUT_XLSX)
    print(f"review rows: {len(candidates)}")


if __name__ == "__main__":
    main()
