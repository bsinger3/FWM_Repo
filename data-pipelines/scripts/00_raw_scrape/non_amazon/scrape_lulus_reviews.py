#!/usr/bin/env python
"""Scrape Lulus PDP review images from server-rendered Nuxt review payloads."""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import math
import re
import shutil
import subprocess
import time
import urllib.error
import urllib.request
from collections import OrderedDict
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime, timezone
import os
from pathlib import Path
from typing import Any

import openpyxl


CSV_COLUMNS = [
    "created_at_display",
    "id",
    "original_url_display",
    "product_page_url_display",
    "monetized_product_url_display",
    "height_raw",
    "weight_raw",
    "user_comment",
    "date_review_submitted_raw",
    "height_in_display",
    "review_date",
    "source_site_display",
    "status_code",
    "content_type",
    "bytes",
    "width",
    "height",
    "hash_md5",
    "fetched_at",
    "updated_at",
    "brand",
    "waist_raw_display",
    "hips_raw",
    "age_raw",
    "waist_in",
    "hips_in_display",
    "age_years_display",
    "search_fts",
    "weight_display_display",
    "weight_raw_needs_correction",
    "clothing_type_id",
    "reviewer_profile_url",
    "reviewer_name_raw",
    "inseam_inches_display",
    "color_canonical",
    "color_display",
    "size_display",
    "bust_in_number_display",
    "cupsize_display",
    "weight_lbs_display",
    "weight_lbs_raw_issue",
    "product_title_raw",
    "product_subtitle_raw",
    "product_description_raw",
    "product_detail_raw",
    "product_category_raw",
    "product_variant_raw",
]

PRODUCT_URL_RE = re.compile(r"https://www\.lulus\.com/products/[^\s\"']+?\.html")
REVIEW_IMAGE_RE = re.compile(r"https://www\.lulus\.com/images/review/[^\s\"'<>),\]]+")
PRODUCT_ID_RE = re.compile(r"/(\d+)\.html(?:$|\?)")
APPAREL_HINT_RE = re.compile(
    r"\b(dress|top|sweater|shirt|blouse|bodysuit|skirt|pant|jean|short|jumpsuit|romper|"
    r"coat|jacket|blazer|vest|cardigan|leggings|trouser|set|gown|maxi|midi|mini)\b",
    re.I,
)

NODE_EXTRACTOR = r"""
let html = "";
process.stdin.setEncoding("utf8");
process.stdin.on("data", chunk => html += chunk);
process.stdin.on("end", () => {
  const vm = require("vm");
  const match = html.match(/<script[^>]*>window\.__NUXT__=([\s\S]*?)<\/script>/);
  if (!match) {
    process.stdout.write(JSON.stringify({error: "nuxt_payload_not_found"}));
    return;
  }
  const context = {window: {}, console: {log(){}, warn(){}, error(){}}};
  vm.createContext(context);
  try {
    vm.runInContext("window.__NUXT__=" + match[1], context, {timeout: 1500});
  } catch (error) {
    process.stdout.write(JSON.stringify({error: String(error && error.message || error)}));
    return;
  }
  const nuxt = context.window.__NUXT__ || {};
  const firstData = Array.isArray(nuxt.data) ? (nuxt.data[0] || {}) : {};
  const product = firstData.product || {};
  const reviewStore = (nuxt.state && nuxt.state.reviewStore) || {};
  process.stdout.write(JSON.stringify({
    product,
    reviewStore: {
      productId: reviewStore.productId,
      productGroupId: reviewStore.productGroupId,
      productReviewCount: reviewStore.productReviewCount,
      reviewsTotal: reviewStore.reviewsTotal,
      reviewImagesTotal: reviewStore.reviewImagesTotal,
      pageIndex: reviewStore.pageIndex,
      pageSize: reviewStore.pageSize,
      reviews: reviewStore.reviews || []
    }
  }));
});
"""


def data_root() -> Path:
    if os.environ.get("FWM_DATA_DIR"):
        return Path(os.environ["FWM_DATA_DIR"])
    cwd_candidate = Path.cwd() / "FWM_Data"
    if cwd_candidate.exists():
        return cwd_candidate
    return Path.cwd().parent / "FWM_Data"


def utc_now() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z")


def fetch_html(url: str, timeout: int = 30) -> tuple[int | None, str]:
    headers = [
        "User-Agent: Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
        "(KHTML, like Gecko) Chrome/135.0.0.0 Safari/537.36",
        "Accept: text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "Accept-Language: en-US,en;q=0.9",
    ]
    curl_bin = shutil.which("curl.exe") or shutil.which("curl") or "curl"
    cmd = [curl_bin, "-sS", "-L", url, "--max-time", str(timeout), "-w", "\n__HTTP_STATUS__:%{http_code}"]
    for header in headers:
        cmd.extend(["-H", header])
    proc = subprocess.run(
        cmd,
        capture_output=True,
        text=True,
        encoding="utf-8",
        errors="replace",
        timeout=timeout + 5,
        check=False,
    )
    output = proc.stdout or ""
    marker = "\n__HTTP_STATUS__:"
    if marker in output:
        html, status_text = output.rsplit(marker, 1)
        try:
            status = int(status_text.strip()[:3])
        except ValueError:
            status = None
        return status, html
    return None, output


def fetch_html_with_urllib(url: str, timeout: int = 30) -> tuple[int | None, str]:
    """Kept for local debugging; Lulus blocks this client on review URLs."""
    req = urllib.request.Request(
        url,
        headers={
            "User-Agent": (
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                "(KHTML, like Gecko) Chrome/135.0.0.0 Safari/537.36"
            ),
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
            "Accept-Language": "en-US,en;q=0.9",
        },
    )
    try:
        with urllib.request.urlopen(req, timeout=timeout) as resp:
            return resp.status, resp.read().decode("utf-8", errors="replace")
    except urllib.error.HTTPError as exc:
        try:
            body = exc.read().decode("utf-8", errors="replace")
        except Exception:
            body = ""
        return exc.code, body
    except Exception:
        return None, ""


def extract_payload(html: str) -> dict[str, Any]:
    proc = subprocess.run(
        ["node", "-e", NODE_EXTRACTOR],
        input=html,
        text=True,
        encoding="utf-8",
        errors="replace",
        capture_output=True,
        timeout=5,
        check=False,
    )
    if proc.returncode != 0:
        return {"error": proc.stderr.strip() or f"node_exit_{proc.returncode}"}
    try:
        return json.loads(proc.stdout)
    except json.JSONDecodeError as exc:
        return {"error": f"json_decode_error: {exc}"}


def product_id_from_url(url: str) -> str | None:
    match = PRODUCT_ID_RE.search(url)
    return match.group(1) if match else None


def review_page_url(product_id: str, page: int, page_size: int) -> str:
    return f"https://www.lulus.com/products/reviews/{product_id}?page={page}&ps={page_size}"


def read_product_urls(workbook: Path, limit: int | None = None) -> list[str]:
    wb = openpyxl.load_workbook(workbook, read_only=True, data_only=True)
    urls: OrderedDict[str, None] = OrderedDict()
    for ws in wb.worksheets:
        for row in ws.iter_rows(values_only=True):
            for value in row:
                if isinstance(value, str):
                    for url in PRODUCT_URL_RE.findall(value):
                        urls.setdefault(url, None)
                        if limit and len(urls) >= limit:
                            return list(urls)
    return list(urls)


def classify_product(product: dict[str, Any], url: str) -> tuple[bool, str]:
    text_parts = [
        str(product.get("name") or ""),
        str(product.get("description") or ""),
        " ".join(str(c.get("name") or "") for c in product.get("categories") or [] if isinstance(c, dict)),
        url,
    ]
    text = " ".join(text_parts)
    if APPAREL_HINT_RE.search(text):
        category = ""
        cats = product.get("categories") or []
        if cats and isinstance(cats[0], dict):
            category = str(cats[0].get("name") or "")
        return True, category
    return False, ""


def normalize_date(value: Any) -> str:
    if not value:
        return ""
    text = str(value)
    if not re.match(r"\d{4}-\d{2}-\d{2}", text):
        return ""
    return text[:10] if len(text) >= 10 else text


def row_id(review: dict[str, Any], image: dict[str, Any], product_url: str) -> str:
    basis = f"lulus|{product_url}|{review.get('id')}|{image.get('id')}|{image.get('url')}"
    return hashlib.md5(basis.encode("utf-8")).hexdigest()


def build_rows(product_url: str, product: dict[str, Any], reviews: list[dict[str, Any]], category: str) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    fetched_at = utc_now()
    product_name = str(product.get("name") or "")
    product_detail = str(product.get("description") or "")
    for review in reviews:
        images = review.get("images") or []
        for image in images:
            image_url = image.get("url")
            if not image_url:
                continue
            body_parts = [str(review.get("title") or ""), str(review.get("bodyText") or "")]
            body = " ".join(part for part in body_parts if part).strip()
            color = str(review.get("colorName") or product.get("color") or "")
            size = str(review.get("sizeOrdered") or "")
            date_raw = str(review.get("datePublished") or image.get("created") or "")
            rows.append(
                {
                    "created_at_display": image.get("created") or date_raw,
                    "id": row_id(review, image, product_url),
                    "original_url_display": image_url,
                    "product_page_url_display": product_url,
                    "monetized_product_url_display": "",
                    "height_raw": review.get("height") or "",
                    "weight_raw": review.get("weight") or "",
                    "user_comment": body,
                    "date_review_submitted_raw": date_raw,
                    "height_in_display": "",
                    "review_date": normalize_date(date_raw),
                    "source_site_display": "https://www.lulus.com/",
                    "status_code": "",
                    "content_type": "",
                    "bytes": "",
                    "width": "",
                    "height": "",
                    "hash_md5": "",
                    "fetched_at": fetched_at,
                    "updated_at": fetched_at,
                    "brand": "Lulus",
                    "waist_raw_display": review.get("waist") or "",
                    "hips_raw": review.get("hip") or "",
                    "age_raw": "",
                    "waist_in": "",
                    "hips_in_display": review.get("hip") or "",
                    "age_years_display": "",
                    "search_fts": " ".join(x for x in [product_name, body] if x),
                    "weight_display_display": review.get("weight") or "",
                    "weight_raw_needs_correction": "",
                    "clothing_type_id": "",
                    "reviewer_profile_url": "",
                    "reviewer_name_raw": review.get("authorName") or "",
                    "inseam_inches_display": "",
                    "color_canonical": review.get("commonColor") or "",
                    "color_display": color,
                    "size_display": size,
                    "bust_in_number_display": review.get("bust") or "",
                    "cupsize_display": review.get("cup") or "",
                    "weight_lbs_display": review.get("weight") or "",
                    "weight_lbs_raw_issue": "",
                    "product_title_raw": product_name or review.get("productName") or "",
                    "product_subtitle_raw": "",
                    "product_description_raw": product_detail,
                    "product_detail_raw": product_detail,
                    "product_category_raw": category,
                    "product_variant_raw": " / ".join(x for x in [color, size] if x),
                }
            )
    return rows


def scrape_product(url: str, page_size: int, delay: float) -> dict[str, Any]:
    product_id = product_id_from_url(url)
    if not product_id:
        return {"url": url, "status": "bad_product_url", "rows": []}

    status, html = fetch_html(review_page_url(product_id, 1, page_size))
    if delay:
        time.sleep(delay)
    if status != 200 or not html:
        return {"url": url, "status": f"http_{status}", "rows": [], "status_code": status}

    payload = extract_payload(html)
    if payload.get("error"):
        return {"url": url, "status": "parse_error", "error": payload["error"], "rows": []}

    product = payload.get("product") or {}
    review_store = payload.get("reviewStore") or {}
    is_apparel, category = classify_product(product, url)
    if not is_apparel:
        return {
            "url": url,
            "status": "skipped_non_apparel",
            "rows": [],
            "review_images_total": review_store.get("reviewImagesTotal") or 0,
        }

    total = int(review_store.get("reviewsTotal") or review_store.get("productReviewCount") or 0)
    image_total = int(review_store.get("reviewImagesTotal") or 0)
    reviews_by_id: OrderedDict[str, dict[str, Any]] = OrderedDict()
    for review in review_store.get("reviews") or []:
        reviews_by_id[str(review.get("id"))] = review

    if image_total:
        page_count = max(1, math.ceil(total / page_size))
        for page in range(2, page_count + 1):
            status, html = fetch_html(review_page_url(product_id, page, page_size))
            if delay:
                time.sleep(delay)
            if status != 200 or not html:
                continue
            payload = extract_payload(html)
            next_store = payload.get("reviewStore") or {}
            for review in next_store.get("reviews") or []:
                reviews_by_id[str(review.get("id"))] = review

    rows = build_rows(url, product, list(reviews_by_id.values()), category)
    return {
        "url": url,
        "status": "ok",
        "rows": rows,
        "reviews_seen": len(reviews_by_id),
        "review_images_total": image_total,
        "rows_found": len(rows),
    }


def write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as fh:
        writer = csv.DictWriter(fh, fieldnames=CSV_COLUMNS)
        writer.writeheader()
        for row in rows:
            writer.writerow({column: row.get(column, "") for column in CSV_COLUMNS})


def row_unique_key(row: dict[str, Any]) -> str:
    product_url = str(row.get("product_page_url_display") or "")
    image_url = str(row.get("original_url_display") or "")
    if product_url and image_url:
        return f"{product_url}|{image_url}"
    return str(row.get("id") or "")


def title_from_url(url: str) -> str:
    match = re.search(r"/products/([^/]+)/\d+\.html", url)
    if not match:
        return ""
    words = [word for word in match.group(1).split("-") if not word.isdigit()]
    return " ".join(word.capitalize() for word in words)


def parse_profile(profile: str) -> dict[str, str]:
    result = {"height": "", "weight": "", "bust": "", "cup": "", "waist": "", "hip": ""}
    if not profile:
        return result
    height = re.search(r"(\d+'\s*\d+\")\s*tall", profile)
    weight = re.search(r"(\d+)\s*lbs", profile)
    bust = re.search(r"\b(\d{2,3})([A-Z]{1,4})\b", profile)
    waist = re.search(r"waist\s+(\d+(?:\.\d+)?\")", profile)
    hip = re.search(r"(\d+(?:\.\d+)?\")\s*hip", profile)
    if height:
        result["height"] = height.group(1).replace(" ", "")
    if weight:
        result["weight"] = f"{weight.group(1)}lbs"
    if bust:
        result["bust"] = bust.group(1)
        result["cup"] = bust.group(2)
    if waist:
        result["waist"] = waist.group(1)
    if hip:
        result["hip"] = hip.group(1)
    return result


def normalize_workbook_image(url: str) -> str:
    return url.replace("/images/review/w_110/", "/images/review/").split("?")[0]


def workbook_row_id(product_url: str, image_url: str, title: str, author: str) -> str:
    basis = f"lulus-workbook|{product_url}|{image_url}|{title}|{author}"
    return hashlib.md5(basis.encode("utf-8")).hexdigest()


def rows_from_review_workbook(workbook: Path) -> list[dict[str, Any]]:
    wb = openpyxl.load_workbook(workbook, read_only=True, data_only=True)
    rows: OrderedDict[str, dict[str, Any]] = OrderedDict()
    fetched_at = utc_now()
    for ws in wb.worksheets:
        if ws.title not in {"bigImages", "oldFormat"}:
            continue
        iterator = ws.iter_rows(values_only=True)
        headers = [str(value or "") for value in next(iterator, [])]
        for raw in iterator:
            values = list(raw)
            product_url = next((str(v) for v in values if isinstance(v, str) and PRODUCT_URL_RE.search(v)), "")
            if not product_url:
                continue
            image_urls: list[str] = []
            for value in values:
                if isinstance(value, str):
                    image_urls.extend(REVIEW_IMAGE_RE.findall(value))
            image_urls = list(OrderedDict.fromkeys(normalize_workbook_image(url) for url in image_urls))
            if not image_urls:
                continue

            title = str(values[3] or "") if ws.title == "bigImages" and len(values) > 3 else ""
            body = str(values[4] or "") if ws.title == "bigImages" and len(values) > 4 else ""
            date_raw = str(values[2] or "") if ws.title == "bigImages" and len(values) > 2 else ""
            author = str(values[7] or "") if ws.title == "bigImages" and len(values) > 7 else ""
            profile = str(values[8] or "") if ws.title == "bigImages" and len(values) > 8 else ""
            fit = str(values[9] or "") if ws.title == "bigImages" and len(values) > 9 else ""
            size = str(values[10] or "") if ws.title == "bigImages" and len(values) > 10 else ""

            if ws.title == "oldFormat":
                date_raw = str(values[3] or "") if len(values) > 3 else ""
                title = str(values[4] or "") if len(values) > 4 else ""
                body = str(values[5] or "") if len(values) > 5 else ""
                author = str(values[7] or "") if len(values) > 7 else ""
                profile = str(values[8] or "") if len(values) > 8 else ""
                fit = str(values[9] or "") if len(values) > 9 else ""
                size = str(values[10] or "") if len(values) > 10 else ""

            color = ""
            event = ""
            body_type = ""
            for idx, value in enumerate(values):
                if not isinstance(value, str) or not value:
                    continue
                label = str(values[idx + 1] or "") if idx + 1 < len(values) else ""
                if label == "Color:":
                    color = value
                elif label == "Event:":
                    event = value
                elif label == "Body Type:":
                    body_type = value

            stats = parse_profile(profile)
            product_title = title_from_url(product_url)
            category = "Dresses" if "dress" in product_url.lower() else "Apparel"
            if "sweater" in product_url.lower() or "top" in product_url.lower():
                category = "Tops"
            if "coat" in product_url.lower() or "jacket" in product_url.lower():
                category = "Outerwear"
            for image_url in image_urls:
                row = {
                    "created_at_display": date_raw,
                    "id": workbook_row_id(product_url, image_url, title, author),
                    "original_url_display": image_url,
                    "product_page_url_display": product_url,
                    "monetized_product_url_display": "",
                    "height_raw": stats["height"],
                    "weight_raw": stats["weight"],
                    "user_comment": " ".join(part for part in [title, body] if part).strip(),
                    "date_review_submitted_raw": date_raw,
                    "height_in_display": "",
                    "review_date": normalize_date(date_raw),
                    "source_site_display": "https://www.lulus.com/",
                    "status_code": "",
                    "content_type": "",
                    "bytes": "",
                    "width": "",
                    "height": "",
                    "hash_md5": "",
                    "fetched_at": fetched_at,
                    "updated_at": fetched_at,
                    "brand": "Lulus",
                    "waist_raw_display": stats["waist"],
                    "hips_raw": stats["hip"],
                    "age_raw": "",
                    "waist_in": "",
                    "hips_in_display": stats["hip"],
                    "age_years_display": "",
                    "search_fts": " ".join(part for part in [product_title, title, body] if part),
                    "weight_display_display": stats["weight"],
                    "weight_raw_needs_correction": "",
                    "clothing_type_id": "",
                    "reviewer_profile_url": "",
                    "reviewer_name_raw": author,
                    "inseam_inches_display": "",
                    "color_canonical": color.lower(),
                    "color_display": color,
                    "size_display": size,
                    "bust_in_number_display": stats["bust"],
                    "cupsize_display": stats["cup"],
                    "weight_lbs_display": stats["weight"],
                    "weight_lbs_raw_issue": "",
                    "product_title_raw": product_title,
                    "product_subtitle_raw": "",
                    "product_description_raw": "",
                    "product_detail_raw": " / ".join(part for part in [fit, event, body_type] if part),
                    "product_category_raw": category,
                    "product_variant_raw": " / ".join(part for part in [color, size] if part),
                }
                rows[str(row["id"])] = row
    return list(rows.values())


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--workbook", type=Path)
    parser.add_argument("--output-prefix", default="lulus_com")
    parser.add_argument("--limit-products", type=int)
    parser.add_argument("--workers", type=int, default=4)
    parser.add_argument("--page-size", type=int, default=40)
    parser.add_argument("--delay", type=float, default=0.15)
    parser.add_argument("--skip-live", action="store_true", help="Only convert workbook review-photo sheets.")
    parser.add_argument("--live-only", action="store_true", help="Do not merge workbook review-photo rows into the output.")
    args = parser.parse_args()

    root = data_root()
    workbook = args.workbook or root / "non-amazon/data/step_1_raw_scraping_data/lulus/Lulus_ProdLinks_March2026.xlsx"
    output_dir = root / "non-amazon/data/step_1_raw_scraping_data/lulus_com"
    csv_path = output_dir / f"{args.output_prefix}_reviews_matching_intake_schema.csv"
    summary_path = output_dir / f"{args.output_prefix}_scrape_summary.json"

    urls = read_product_urls(workbook, args.limit_products)
    all_rows: list[dict[str, Any]] = []
    product_results: list[dict[str, Any]] = []
    started = utc_now()

    if not args.skip_live:
        with ThreadPoolExecutor(max_workers=args.workers) as executor:
            futures = {executor.submit(scrape_product, url, args.page_size, args.delay): url for url in urls}
            for idx, future in enumerate(as_completed(futures), 1):
                result = future.result()
                product_results.append({k: v for k, v in result.items() if k != "rows"})
                all_rows.extend(result.get("rows") or [])
                if idx % 20 == 0 or idx == len(urls):
                    print(f"processed {idx}/{len(urls)} products; rows={len(all_rows)}", flush=True)

    unique_rows: OrderedDict[str, dict[str, Any]] = OrderedDict()
    for row in all_rows:
        unique_rows[row_unique_key(row)] = row
    live_rows = len(unique_rows)
    workbook_rows = [] if args.live_only else rows_from_review_workbook(workbook)
    for row in workbook_rows:
        unique_rows.setdefault(row_unique_key(row), row)
    all_rows = list(unique_rows.values())
    used_workbook_rows = bool(workbook_rows)
    write_csv(csv_path, all_rows)

    products_with_rows = sum(1 for result in product_results if result.get("rows_found"))
    if used_workbook_rows:
        products_with_rows = len({row["product_page_url_display"] for row in all_rows})
    if args.skip_live:
        product_results = [{"status": "live_fetch_skipped", "reason": "converted workbook review-photo sheets only"}]
    summary = {
        "site": "lulus.com",
        "retailer": "lulus_com",
        "status": (
            "completed_live_only"
            if args.live_only
            else "completed_from_workbook_review_sheets"
            if args.skip_live or (used_workbook_rows and not live_rows)
            else "completed_from_live_and_workbook_review_sheets"
            if used_workbook_rows
            else "completed"
        ),
        "started_at": started,
        "finished_at": utc_now(),
        "source_workbook": str(workbook),
        "product_urls_discovered": len(urls),
        "product_urls_processed": 0 if args.skip_live else len(product_results),
        "products_with_review_image_rows": products_with_rows,
        "rows": len(all_rows),
        "live_rows": live_rows,
        "workbook_rows": len(workbook_rows),
        "distinct_image_rows": len({row["id"] for row in all_rows}),
        "distinct_product_urls": len({row["product_page_url_display"] for row in all_rows}),
        "rows_with_size": sum(1 for row in all_rows if row.get("size_display")),
        "rows_with_measurement": sum(
            1
            for row in all_rows
            if row.get("height_raw")
            or row.get("weight_raw")
            or row.get("waist_raw_display")
            or row.get("hips_raw")
            or row.get("bust_in_number_display")
            or row.get("cupsize_display")
        ),
        "output_csv": str(csv_path),
        "notes": [
            "Used March 2026 Lulus PDP workbook as product discovery seed because lulus.com does not expose Shopify products.json.",
            "Review records are extracted from server-rendered Nuxt review pages under /products/reviews/{product_id}?page=N&ps=40.",
            "Rows are one per public customer review image URL.",
            "If live Lulus review pages return PerimeterX captcha/403, the scraper falls back to the workbook's bigImages and oldFormat review-photo sheets.",
        ],
        "used_workbook_rows": used_workbook_rows,
        "product_results": product_results,
    }
    with summary_path.open("w", encoding="utf-8") as fh:
        json.dump(summary, fh, indent=2, ensure_ascii=False)
    print(json.dumps({k: summary[k] for k in ["rows", "product_urls_processed", "products_with_review_image_rows"]}, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
