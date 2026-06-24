#!/usr/bin/env python3
"""Build an image-review dashboard package from AWIN affiliate link outputs."""

from __future__ import annotations

import argparse
import csv
import sys
from pathlib import Path


REPO_ROOT = Path(__file__).resolve().parents[4]
PIPELINE_SCRIPTS_DIR = REPO_ROOT / "data-pipelines" / "scripts"
if str(PIPELINE_SCRIPTS_DIR) not in sys.path:
    sys.path.insert(0, str(PIPELINE_SCRIPTS_DIR))

from pipeline_paths import cv_annotated_pending_human_review_root  # noqa: E402


OUTPUT_COLUMNS = [
    "production_decision",
    "review_notes",
    "sovrn_has_payout",
    "sovrn_priority",
    "sovrn_payout_priority_rank",
    "sovrn_estimated_commission_per_click",
    "sovrn_pricing",
    "sovrn_merchant_group",
    "sorter_recommendation",
    "sorter_reason_codes",
    "source_family",
    "source_site_display",
    "review_row_key",
    "image_url_to_use",
    "raw_scraped_image_url",
    "needs_url_update",
    "product_page_url_display",
    "monetized_product_url_display",
    "brand",
    "product_title_raw",
    "product_category_raw",
    "product_variant_raw",
    "clothing_type_id",
    "size_display",
    "height_in_display",
    "weight_lbs_display",
    "weight_display_display",
    "waist_in",
    "hips_in_display",
    "bust_in_display",
    "bra_band_in_display",
    "cupsize_display",
    "inseam_inches_display",
    "user_comment",
    "source_file",
    "source_row_number",
]
MAX_ROWS_PER_WORKBOOK = 1000


def read_csv(path: Path) -> list[dict[str, str]]:
    with path.open(newline="", encoding="utf-8-sig") as handle:
        return list(csv.DictReader(handle))


def write_workbook(path: Path, rows: list[dict[str, str]]) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.datavalidation import DataValidation
    except ModuleNotFoundError as exc:
        raise SystemExit("openpyxl is required to build review workbooks.") from exc

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Review"
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(OUTPUT_COLUMNS))}{len(rows) + 1}"

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    sheet.append(OUTPUT_COLUMNS)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    for row in rows:
        sheet.append([row.get(column, "") for column in OUTPUT_COLUMNS])

    validation = DataValidation(type="list", formula1='"APPROVE,REJECT,NEEDS_MORE_REVIEW,SKIP"', allow_blank=True)
    sheet.add_data_validation(validation)
    validation.add(f"A2:A{max(2, len(rows) + 1)}")

    for column_index in range(1, len(OUTPUT_COLUMNS) + 1):
        sheet.column_dimensions[get_column_letter(column_index)].width = 18
    for column, width in {
        "A": 24,
        "B": 34,
        "L": 28,
        "M": 55,
        "N": 55,
        "O": 55,
        "Q": 55,
        "R": 55,
        "T": 40,
        "U": 36,
        "AH": 55,
        "AI": 55,
    }.items():
        sheet.column_dimensions[column].width = width

    for row_cells in sheet.iter_rows(min_row=2, max_row=len(rows) + 1):
        for cell in row_cells:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def row_at(csv_path: Path, source_row_number: int) -> dict[str, str]:
    with csv_path.open(newline="", encoding="utf-8-sig") as handle:
        for row_number, row in enumerate(csv.DictReader(handle), start=2):
            if row_number == source_row_number:
                return row
    return {}


def chunked(rows: list[dict[str, str]], size: int) -> list[list[dict[str, str]]]:
    return [rows[index : index + size] for index in range(0, len(rows), size)]


def build_rows(candidates_csv: Path, link_map_csv: Path) -> list[dict[str, str]]:
    tracking_by_url = {
        row["normalized_product_url"]: row.get("tracking_url", "")
        for row in read_csv(link_map_csv)
        if row.get("status") == "generated" and row.get("tracking_url")
    }
    rows: list[dict[str, str]] = []
    seen_keys: set[str] = set()
    for candidate in read_csv(candidates_csv):
        if candidate.get("status") != "eligible":
            continue
        tracking_url = tracking_by_url.get(candidate.get("normalized_product_url", ""))
        if not tracking_url:
            continue
        source_file = Path(candidate["source_csv"])
        source_row_number = int(candidate["source_row_number"])
        source = row_at(source_file, source_row_number)
        raw_image = (source.get("original_url_display") or "").strip()
        product_url = (source.get("product_page_url_display") or candidate.get("destination_url") or "").strip()
        source_token = source_file.stem.replace("_reviews_matching_intake_schema", "")
        review_row_key = (
            f"awin::{candidate.get('matched_domain') or candidate.get('product_domain')}::"
            f"{source_token}::{source.get('id') or 'row'}::{source_row_number}"
        )
        dedupe_key = f"{review_row_key}::{raw_image}::{product_url}"
        if dedupe_key in seen_keys:
            continue
        seen_keys.add(dedupe_key)
        rows.append(
            {
                "production_decision": "",
                "review_notes": "",
                "sorter_recommendation": "NEEDS_HUMAN_REVIEW_AWIN_QUALIFIED_ROW",
                "sorter_reason_codes": "AWIN_SUPABASE_QUALIFIED_LINKED",
                "source_family": "non_amazon_awin",
                "source_site_display": source.get("source_site_display") or candidate.get("product_domain", ""),
                "review_row_key": review_row_key,
                "image_url_to_use": raw_image,
                "raw_scraped_image_url": raw_image,
                "needs_url_update": "",
                "product_page_url_display": product_url,
                "monetized_product_url_display": tracking_url,
                "brand": source.get("brand", ""),
                "product_title_raw": source.get("product_title_raw", ""),
                "product_category_raw": source.get("product_category_raw", ""),
                "product_variant_raw": source.get("product_variant_raw", ""),
                "clothing_type_id": source.get("clothing_type_id", ""),
                "size_display": source.get("size_display") or source.get("size_ordered_raw_display", ""),
                "height_in_display": source.get("height_in_display", ""),
                "weight_lbs_display": source.get("weight_lbs_display") or source.get("weight_lb", ""),
                "weight_display_display": source.get("weight_display_display", ""),
                "waist_in": source.get("waist_in", ""),
                "hips_in_display": source.get("hips_in_display", ""),
                "bust_in_display": source.get("bust_in_display") or source.get("bust_in_number_display", ""),
                "bra_band_in_display": source.get("bra_band_in_display", ""),
                "cupsize_display": source.get("cupsize_display", ""),
                "inseam_inches_display": source.get("inseam_inches_display", ""),
                "user_comment": source.get("user_comment", ""),
                "source_file": str(source_file),
                "source_row_number": str(source_row_number),
            }
        )
    rows.sort(key=lambda row: (row.get("source_site_display", ""), row.get("product_title_raw", ""), row.get("review_row_key", "")))
    return rows


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--candidates-csv", required=True, type=Path)
    parser.add_argument("--link-map-csv", required=True, type=Path)
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=cv_annotated_pending_human_review_root() / "awin_supabase_qualified_linked_20260616",
    )
    args = parser.parse_args()

    rows = build_rows(args.candidates_csv, args.link_map_csv)
    args.output_dir.mkdir(parents=True, exist_ok=True)
    written = []
    for part, chunk in enumerate(chunked(rows, MAX_ROWS_PER_WORKBOOK), start=1):
        path = args.output_dir / f"supabase_image_review_needs_human_review_part_{part:03d}.xlsx"
        write_workbook(path, chunk)
        written.append(path)

    readme = [
        "# AWIN Supabase-Qualified Linked Image Review Package",
        "",
        f"- Rows: `{len(rows)}`",
        f"- Workbooks: `{len(written)}`",
        f"- Candidates CSV: `{args.candidates_csv}`",
        f"- Link map CSV: `{args.link_map_csv}`",
        "",
        "All rows are AWIN applied-queue rows that passed the broad Supabase-qualified image rule and have generated AWIN tracking URLs.",
    ]
    (args.output_dir / "README.md").write_text("\n".join(readme) + "\n", encoding="utf-8")
    print(args.output_dir)
    print(f"rows: {len(rows)}")
    print(f"workbooks: {len(written)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
