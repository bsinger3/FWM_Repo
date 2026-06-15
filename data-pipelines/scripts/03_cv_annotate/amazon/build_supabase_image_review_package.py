#!/usr/bin/env python3
"""Build Google-Sheets-friendly review workbooks for Supabase image candidates."""

from __future__ import annotations

import csv
import argparse
import re
import sys
from collections import Counter
from pathlib import Path
from urllib.parse import urlsplit


REPO_ROOT = Path(__file__).resolve().parents[4]
PIPELINE_SCRIPTS_DIR = REPO_ROOT / "data-pipelines" / "scripts"
if str(PIPELINE_SCRIPTS_DIR) not in sys.path:
    sys.path.insert(0, str(PIPELINE_SCRIPTS_DIR))

from pipeline_paths import (  # noqa: E402
    archive_root,
    cv_annotated_pending_human_review_root,
    raw_scraped_data_root,
)

AMAZON_MANUAL_CHUNKS = (
    cv_annotated_pending_human_review_root()
    / "amazon_legacy_step_4_human_review_and_visibility_decisions"
    / "manual_chunks"
)
NON_AMAZON_RAW = raw_scraped_data_root()
SORTER_RESULTS = (
    archive_root()
    / "old_outputs"
    / "repo_outputs_archive"
    / "supabase_output_cleanup_2026_05_29"
    / "cv_experiments"
    / "fit_image_sorter_2026_05_27"
    / "fit_image_sorter_results.csv"
)
OUT_DIR = cv_annotated_pending_human_review_root() / "supabase_production_image_review_2026_05_28_s3_refresh_sovrn_prioritized"
SOVRN_CANDIDATES = REPO_ROOT / "data-pipelines/docs/sovrn_commerce/sovrn_commerce_scrape_triage_candidates.csv"
SOVRN_TRACKER = REPO_ROOT / "data-pipelines/docs/sovrn_commerce/sovrn_commerce_apparel_triage_tracker.csv"

MAX_ROWS_PER_WORKBOOK = 1000

Workbook = None
Alignment = None
Font = None
PatternFill = None
get_column_letter = None
DataValidation = None


def require_openpyxl() -> None:
    global Workbook, Alignment, Font, PatternFill, get_column_letter, DataValidation
    if Workbook is not None:
        return
    try:
        from openpyxl import Workbook as openpyxl_workbook
        from openpyxl.styles import Alignment as openpyxl_alignment
        from openpyxl.styles import Font as openpyxl_font
        from openpyxl.styles import PatternFill as openpyxl_pattern_fill
        from openpyxl.utils import get_column_letter as openpyxl_get_column_letter
        from openpyxl.worksheet.datavalidation import DataValidation as openpyxl_data_validation
    except ModuleNotFoundError as exc:
        raise SystemExit("openpyxl is required to build review workbooks. Install project dependencies and rerun.") from exc

    Workbook = openpyxl_workbook
    Alignment = openpyxl_alignment
    Font = openpyxl_font
    PatternFill = openpyxl_pattern_fill
    get_column_letter = openpyxl_get_column_letter
    DataValidation = openpyxl_data_validation

OUTPUT_COLUMNS = [
    "production_decision",
    "review_notes",
    "sovrn_has_payout",
    "sovrn_priority",
    "sovrn_payout_priority_rank",
    "sovrn_estimated_commission_per_click",
    "sovrn_pricing",
    "sovrn_merchant_group",
    "sorter_recommendation",
    "sorter_reason_codes",
    "source_family",
    "source_site_display",
    "review_row_key",
    "image_url_to_use",
    "raw_scraped_image_url",
    "needs_url_update",
    "product_page_url_display",
    "monetized_product_url_display",
    "brand",
    "product_title_raw",
    "product_category_raw",
    "product_variant_raw",
    "clothing_type_id",
    "size_display",
    "height_in_display",
    "weight_lbs_display",
    "weight_display_display",
    "waist_in",
    "hips_in_display",
    "bust_in_display",
    "bra_band_in_display",
    "cupsize_display",
    "inseam_inches_display",
    "user_comment",
    "source_file",
    "source_row_number",
]

MEASUREMENT_COLUMNS = [
    "height_in_display",
    "weight_lbs_display",
    "weight_display_display",
    "waist_in",
    "hips_in_display",
    "bust_in_display",
    "bra_band_in_display",
    "cupsize_display",
    "inseam_inches_display",
]


def read_csv(path: Path) -> list[dict[str, str]]:
    with path.open(newline="", encoding="utf-8-sig") as handle:
        return list(csv.DictReader(handle))


def valid_http_url(value: str) -> bool:
    try:
        parts = urlsplit(str(value or "").strip())
    except ValueError:
        return False
    return parts.scheme in {"http", "https"} and bool(parts.netloc)


def normalize_domain(value: str) -> str:
    text = str(value or "").strip().lower()
    if not text:
        return ""
    if "://" not in text:
        text = "https://" + text
    try:
        host = urlsplit(text).netloc.lower()
    except ValueError:
        return ""
    if host.startswith("www."):
        host = host[4:]
    return host


def domain_candidates(value: str) -> list[str]:
    host = normalize_domain(value)
    if not host:
        return []
    parts = host.split(".")
    candidates = [host]
    if len(parts) > 2:
        candidates.append(".".join(parts[-2:]))
    return list(dict.fromkeys(candidates))


def money_to_float(value: str) -> float:
    text = str(value or "").replace("$", "").replace(",", "").strip()
    try:
        return float(text)
    except ValueError:
        return 0.0


def rank_to_int(value: str) -> int:
    try:
        return int(float(str(value or "").strip()))
    except ValueError:
        return 999_999


def truthy(value: object) -> bool:
    return str(value or "").strip().lower() in {"true", "yes", "1", "y"}


def blank(value: object) -> bool:
    return str(value or "").strip() == ""


def has_measurement(row: dict[str, str]) -> bool:
    return any(not blank(row.get(column)) for column in MEASUREMENT_COLUMNS)


def has_size(row: dict[str, str]) -> bool:
    return not blank(row.get("size_display"))


def product_url(row: dict[str, str]) -> str:
    return str(row.get("product_page_url_display") or row.get("monetized_product_url_display") or "").strip()


def normalized_site_from_path(path: Path) -> str:
    return path.parent.name


def non_amazon_review_files() -> list[Path]:
    patterns = [
        "*/*_reviews_matching_intake_schema.csv",
        "*/*_reviews_matching_amazon_schema.csv",
    ]
    files: dict[Path, Path] = {}
    for pattern in patterns:
        for path in NON_AMAZON_RAW.glob(pattern):
            files[path] = path
    return sorted(files)


def load_sorter_lookup() -> dict[str, dict[str, str]]:
    lookup: dict[str, dict[str, str]] = {}
    if not SORTER_RESULTS.exists():
        return lookup
    for row in read_csv(SORTER_RESULTS):
        raw = row.get("raw_scraped_image_url") or row.get("original_url_display") or ""
        key = raw.strip()
        if key:
            lookup[key] = row
    return lookup


def load_sovrn_lookup() -> dict[str, dict[str, str]]:
    lookup: dict[str, dict[str, str]] = {}
    for path in [SOVRN_TRACKER, SOVRN_CANDIDATES]:
        if not path.exists():
            continue
        for row in read_csv(path):
            domains = row.get("primary_domains") or row.get("primary_domain") or ""
            for domain in re.split(r"[;|,]", domains):
                normalized = normalize_domain(domain)
                if not normalized:
                    continue
                current = lookup.get(normalized)
                if current is None or rank_to_int(row.get("payout_priority_rank", "")) < rank_to_int(current.get("payout_priority_rank", "")):
                    lookup[normalized] = row
    return lookup


def sovrn_for_row(row: dict[str, str], lookup: dict[str, dict[str, str]]) -> dict[str, str]:
    for value in [row.get("product_page_url_display", ""), row.get("monetized_product_url_display", ""), row.get("source_site_display", "")]:
        for domain in domain_candidates(value):
            if domain in lookup:
                match = lookup[domain]
                pricing = match.get("pricing", "")
                has_payout = "CPA" in pricing.upper() or "CPC" in pricing.upper()
                return {
                    "sovrn_has_payout": "TRUE" if has_payout else "",
                    "sovrn_priority": match.get("priority", ""),
                    "sovrn_payout_priority_rank": match.get("payout_priority_rank", ""),
                    "sovrn_estimated_commission_per_click": match.get("estimated_commission_per_click", ""),
                    "sovrn_pricing": pricing,
                    "sovrn_merchant_group": match.get("merchant_group", ""),
                }
    return {
        "sovrn_has_payout": "",
        "sovrn_priority": "",
        "sovrn_payout_priority_rank": "",
        "sovrn_estimated_commission_per_click": "",
        "sovrn_pricing": "",
        "sovrn_merchant_group": "",
    }


def base_output_row(
    row: dict[str, str],
    source_family: str,
    source_file: Path,
    row_number: int,
    review_row_key: str,
    sorter_lookup: dict[str, dict[str, str]],
    sovrn_lookup: dict[str, dict[str, str]],
) -> dict[str, str]:
    raw_url = str(row.get("original_url_display") or "").strip()
    sorter = sorter_lookup.get(raw_url, {})
    image_url_to_use = sorter.get("sort_image_url") or raw_url
    output = {
        "production_decision": "",
        "review_notes": "",
        "sorter_recommendation": sorter.get("primary_action", ""),
        "sorter_reason_codes": sorter.get("reason_codes", ""),
        "source_family": source_family,
        "source_site_display": row.get("source_site_display") or ("https://www.amazon.com/" if source_family == "amazon" else ""),
        "review_row_key": review_row_key,
        "image_url_to_use": image_url_to_use,
        "raw_scraped_image_url": raw_url,
        "needs_url_update": sorter.get("needs_url_update", ""),
        "product_page_url_display": row.get("product_page_url_display", ""),
        "monetized_product_url_display": row.get("monetized_product_url_display", ""),
        "brand": row.get("brand", ""),
        "product_title_raw": row.get("product_title_raw", ""),
        "product_category_raw": row.get("product_category_raw", ""),
        "product_variant_raw": row.get("product_variant_raw", ""),
        "clothing_type_id": row.get("clothing_type_id", ""),
        "size_display": row.get("size_display", ""),
        "height_in_display": row.get("height_in_display", ""),
        "weight_lbs_display": row.get("weight_lbs_display", ""),
        "weight_display_display": row.get("weight_display_display", ""),
        "waist_in": row.get("waist_in", ""),
        "hips_in_display": row.get("hips_in_display", ""),
        "bust_in_display": row.get("bust_in_display") or row.get("bust_in_number_display", ""),
        "bra_band_in_display": row.get("bra_band_in_display", ""),
        "cupsize_display": row.get("cupsize_display", ""),
        "inseam_inches_display": row.get("inseam_inches_display", ""),
        "user_comment": row.get("user_comment", ""),
        "source_file": str(source_file),
        "source_row_number": str(row_number),
    }
    output.update(sovrn_for_row(output, sovrn_lookup))
    return output


def recommend_amazon(row: dict[str, str], output: dict[str, str]) -> str:
    reasons = []
    if not truthy(row.get("has_person")):
        reasons.append("NO_PERSON_EXISTING_FLAG")
    if row.get("lighting_ok") and not truthy(row.get("lighting_ok")):
        reasons.append("LIGHTING_NOT_OK_EXISTING_FLAG")

    clothing = str(row.get("clothing_type_id") or "").lower()
    lower_body = bool(re.search(r"jean|pant|short|legging|skirt|trouser|bottom", clothing))
    if lower_body and row.get("full_lower_body_visible") and not truthy(row.get("full_lower_body_visible")):
        reasons.append("LOWER_BODY_NOT_FULLY_VISIBLE_EXISTING_FLAG")

    if reasons:
        output["sorter_recommendation"] = "NEEDS_HUMAN_REVIEW"
        output["sorter_reason_codes"] = ";".join(reasons)
        return "needs_human_review"

    output["sorter_recommendation"] = "APPROVE_CANDIDATE"
    output["sorter_reason_codes"] = "EXISTING_AMAZON_FLAGS_PASS"
    return "approve_candidates"


def recommend_non_amazon(output: dict[str, str]) -> str:
    if output["sorter_recommendation"] in {"QUALITY_REVIEW", "LLM_SEMANTIC_REVIEW", "LLM_NOT_WORN_REVIEW", "CROP_REVIEW_PRIORITY"}:
        return "needs_human_review"
    if output["sorter_recommendation"] == "LLM_APPROVAL_CONFIRMATION":
        output["sorter_recommendation"] = "APPROVE_CANDIDATE_NEEDS_FINAL_MANUAL_CONFIRMATION"
        return "approve_candidates"
    output["sorter_recommendation"] = "NEEDS_HUMAN_REVIEW_NO_CV_SORT_YET"
    output["sorter_reason_codes"] = output["sorter_reason_codes"] or "NON_AMAZON_QUALIFIED_ROW_NOT_IN_CV_SAMPLE"
    return "needs_human_review"


def gather_candidates() -> dict[str, list[dict[str, str]]]:
    sorter_lookup = load_sorter_lookup()
    sovrn_lookup = load_sovrn_lookup()
    buckets: dict[str, list[dict[str, str]]] = {
        "approve_candidates": [],
        "disapprove_candidates": [],
        "needs_human_review": [],
    }
    seen: set[tuple[str, str, str]] = set()

    for path in sorted(AMAZON_MANUAL_CHUNKS.glob("images_to_approve_part_*.csv")):
        for row_number, row in enumerate(read_csv(path), start=2):
            raw_url = str(row.get("original_url_display") or "").strip()
            if not valid_http_url(raw_url) or not valid_http_url(product_url(row)):
                continue
            if not has_size(row) or not has_measurement(row):
                continue
            key = ("amazon", raw_url, product_url(row))
            if key in seen:
                continue
            seen.add(key)
            review_key = f"amazon::{path.stem}::{row_number}"
            output = base_output_row(row, "amazon", path, row_number, review_key, sorter_lookup, sovrn_lookup)
            bucket = recommend_amazon(row, output)
            buckets[bucket].append(output)

    for path in non_amazon_review_files():
        for row_number, row in enumerate(read_csv(path), start=2):
            raw_url = str(row.get("original_url_display") or "").strip()
            if not valid_http_url(raw_url) or not valid_http_url(product_url(row)):
                continue
            if not has_size(row) or not has_measurement(row):
                continue
            site = row.get("source_site_display") or normalized_site_from_path(path)
            key = ("non_amazon", raw_url, product_url(row))
            if key in seen:
                continue
            seen.add(key)
            review_key = f"nonamazon::{site}::{row.get('id') or row_number}"
            output = base_output_row(row, "non_amazon", path, row_number, review_key, sorter_lookup, sovrn_lookup)
            if not output["source_site_display"]:
                output["source_site_display"] = site
            bucket = recommend_non_amazon(output)
            buckets[bucket].append(output)

    for rows in buckets.values():
        rows.sort(key=priority_sort_key)
    return buckets


def priority_sort_key(row: dict[str, str]) -> tuple[object, ...]:
    has_payout = row.get("sovrn_has_payout") == "TRUE"
    priority = str(row.get("sovrn_priority") or "P9").upper()
    priority_number = rank_to_int(priority[1:] if priority.startswith("P") else priority)
    return (
        0 if has_payout else 1,
        rank_to_int(row.get("sovrn_payout_priority_rank", "")),
        priority_number,
        -money_to_float(row.get("sovrn_estimated_commission_per_click", "")),
        row.get("source_family", ""),
        row.get("source_site_display", ""),
    )


def write_workbook(path: Path, rows: list[dict[str, str]]) -> None:
    require_openpyxl()
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Review"
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(OUTPUT_COLUMNS))}{len(rows) + 1}"

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    sheet.append(OUTPUT_COLUMNS)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    for row in rows:
        sheet.append([row.get(column, "") for column in OUTPUT_COLUMNS])

    validation = DataValidation(
        type="list",
        formula1='"APPROVE,REJECT,NEEDS_MORE_REVIEW,SKIP"',
        allow_blank=True,
    )
    sheet.add_data_validation(validation)
    validation.add(f"A2:A{max(2, len(rows) + 1)}")

    widths = {
        "A": 24,
        "B": 34,
        "C": 38,
        "D": 44,
        "E": 16,
        "F": 24,
        "G": 34,
        "H": 60,
        "I": 60,
        "J": 16,
        "K": 55,
        "L": 55,
        "M": 18,
        "N": 34,
        "O": 24,
        "P": 28,
        "Q": 18,
        "R": 16,
        "S": 14,
        "T": 16,
        "U": 12,
        "V": 14,
        "W": 14,
        "X": 16,
        "Y": 14,
        "Z": 55,
        "AA": 55,
        "AB": 16,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width

    for row_cells in sheet.iter_rows(min_row=2, max_row=len(rows) + 1):
        for cell in row_cells:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def chunked(rows: list[dict[str, str]], size: int) -> list[list[dict[str, str]]]:
    return [rows[index : index + size] for index in range(0, len(rows), size)]


def write_package(buckets: dict[str, list[dict[str, str]]]) -> list[Path]:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    written = []
    for bucket, rows in buckets.items():
        for part, chunk in enumerate(chunked(rows, MAX_ROWS_PER_WORKBOOK), start=1):
            path = OUT_DIR / f"supabase_image_review_{bucket}_part_{part:03d}.xlsx"
            write_workbook(path, chunk)
            written.append(path)
    return written


def write_readme(buckets: dict[str, list[dict[str, str]]], files: list[Path]) -> None:
    source_counts = Counter()
    for rows in buckets.values():
        for row in rows:
            source_counts[row["source_family"]] += 1
    lines = [
        "# Supabase Production Image Review Package",
        "",
        "Open these `.xlsx` files in Google Sheets. Fill in only `production_decision` and, when useful, `review_notes`.",
        "",
        "Allowed production decisions: `APPROVE`, `REJECT`, `NEEDS_MORE_REVIEW`, `SKIP`.",
        "",
        "## Buckets",
        "",
    ]
    for bucket, rows in buckets.items():
        lines.append(f"- `{bucket}`: {len(rows)} rows")
    lines.extend(["", "## Source Families", ""])
    for source, count in source_counts.most_common():
        lines.append(f"- `{source}`: {count} rows")
    payout_count = sum(1 for rows in buckets.values() for row in rows if row.get("sovrn_has_payout") == "TRUE")
    lines.extend(["", "## Sovrn Priority", "", f"- Rows matched to Sovrn payout merchants: `{payout_count}`"])
    lines.append("- Workbooks are sorted so Sovrn payout rows come first, then lower payout rank, then higher estimated commission per click.")
    lines.extend(["", "## Files", ""])
    for path in files:
        lines.append(f"- `{path.name}`")
    lines.extend(
        [
            "",
            "## Notes",
            "",
            "- Rows with blank or non-http image URLs are excluded.",
            "- `image_url_to_use` is the URL to put into the production database if the row is approved.",
            "- `raw_scraped_image_url` preserves the original source URL.",
            "- `needs_url_update` indicates that the sorting logic found a larger replacement image URL.",
            "- The bucket is a starting recommendation; your `production_decision` is the final manual gate.",
            "",
        ]
    )
    (OUT_DIR / "README.md").write_text("\n".join(lines), encoding="utf-8")


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.parse_args()

    buckets = gather_candidates()
    files = write_package(buckets)
    write_readme(buckets, files)
    print(OUT_DIR)
    for bucket, rows in buckets.items():
        print(f"{bucket}: {len(rows)}")
    print(f"workbooks: {len(files)}")


if __name__ == "__main__":
    main()
